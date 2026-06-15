from ntpath import join
import os
import sys
import subprocess
import tkinter as tk
from tkinter import messagebox, Canvas
from PIL import Image, ImageTk
import cv2
from getmac import get_mac_address
from datetime import datetime
import getpass
from pathlib import Path
import tempfile
import atexit
import random 

#------------------------------------------------------------------------------------------------------------------------------------------------------------------


def single_instance_check():
    lock_file = os.path.join(tempfile.gettempdir(), "EDA_APP.lock")
    if os.path.exists(lock_file):
        messagebox.showwarning(
            "Already Running",
            "⚠️ EDA application is already running."
        )
        try:
            os.remove(lock_file)
        except FileNotFoundError:
            pass
        sys.exit(0)

    with open(lock_file, "w") as f:
        f.write(str(os.getpid()))

    def cleanup():
        try:
            if os.path.exists(lock_file):
                os.remove(lock_file)
        except FileNotFoundError:
            pass

    atexit.register(cleanup)

single_instance_check()


#------------------------------------------------------------------------------------------------------------------------------------------------------------------


def get_global_groups(username):
    command = f"net user {username} /domain"
    result = subprocess.run(command, capture_output=True, text=True, shell=True)
    if result.returncode != 0:
        return None, None

    output = result.stdout
    full_name = None
    groups = []
    capturing_groups = False

    for line in output.splitlines():
        if "Full Name" in line and full_name is None:
            full_name = line.split("Full Name")[1].strip()
        if "Local Group Memberships" in line:
            capturing_groups = True
        elif capturing_groups and line.strip() == "":
            break
        elif capturing_groups:
            groups.append(line.strip())

    return full_name, groups


username = os.getlogin()
full_name, groups = get_global_groups(username) 
mac_address = get_mac_address()

PassProgram = "UnBlock"

if groups:
    group_str = " ".join(", ".join(groups).split())
    if "designappgrp_rw" in group_str:
        PassProgram = "UnBlock"

        GroupList = []
        PurposeList = []
        if "App_Design" in group_str:
            GroupList.append("Design")
        if "App_Simulation" in group_str:
            GroupList.append("Simulation")
        if "App_Application" in group_str:
            GroupList.append("Application")
        if "App_Documentation" in group_str:
            GroupList.append("Documentation")
        if "App_QC" in group_str:
            GroupList.append("QC")

        GroupListStr = ",".join(GroupList)

if PassProgram == "UnBlock":
    def resource_path(relative_path):
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    cap = None
    after_id = None
    window = tk.Tk()
    window.iconbitmap(resource_path("icon.ico"))
    window.geometry("1280x720")
    window.title(f"EDA-v29 : Hello..  {full_name}")
    window.configure(bg='black')

#------------------------------------------------------------------------------------------------------------------------------------------------------------------


    class MainPage:
        def __init__(self, window, text, y, command, border_style="circuit"):
            self.window = window
            self.text = text
            self.y = y
            self.command = command
            self.is_hovered = False
            self.matrix_chars = []
            self.drops = []
            self.border_style = border_style
            self.animation_running = True
            self.after_id = None
            
            self.canvas = Canvas(window, width=400, height=60, 
                                highlightthickness=0, bd=0, bg="#051542")
            self.canvas.place(relx=0.5, rely=y, anchor="center")
            
            self.create_base()
            
            self.text_bg = self.canvas.create_rectangle(
                60, 18, 340, 42,  
                fill="#020638", outline="#CAE4EB", width=1
            )
            
            self.text_shadow = self.canvas.create_text(
                202, 32, text=text, 
                font=("Consolas", 14, "bold"),
                fill="#000000"  
            )
            self.btn_text = self.canvas.create_text(
                200, 30, text=text, 
                font=("Consolas", 14, "bold"),
                fill="#CAE4EB"
            )
            
            self.init_matrix_rain()
            

            self.canvas.bind("<Enter>", self.on_enter)
            self.canvas.bind("<Leave>", self.on_leave)
            self.canvas.bind("<Button-1>", self.on_click)
            
            self.animate_matrix()
        
        def create_base(self):
            if self.border_style == "circuit":
                self.create_circuit_border()
            elif self.border_style == "glitch":
                self.create_glitch_border()
            elif self.border_style == "segmented":
                self.create_segmented_border()
            elif self.border_style == "corner":
                self.create_corner_border()
            else:
                self.create_default_border()
        
        def create_circuit_border(self):
            self.border_elements = []
            
            self.border_elements.append(self.canvas.create_line(30, 10, 370, 10, fill="#CAE4EB", width=2))
            for x in range(50, 350, 40):
                node = self.canvas.create_oval(x-3, 7, x+3, 13, fill="#020638", outline="#CAE4EB")
                self.border_elements.append(node)
            
            self.border_elements.append(self.canvas.create_line(30, 50, 370, 50, fill="#CAE4EB", width=2))
            for x in range(50, 350, 40):
                node = self.canvas.create_oval(x-3, 47, x+3, 53, fill="#020638", outline="#CAE4EB")
                self.border_elements.append(node)
            
            self.border_elements.append(self.canvas.create_line(20, 20, 20, 40, fill="#CAE4EB", width=2))
            for y in range(25, 35, 5):
                node = self.canvas.create_oval(17, y-3, 23, y+3, fill="#020638", outline="#CAE4EB")
                self.border_elements.append(node)
            
            self.border_elements.append(self.canvas.create_line(380, 20, 380, 40, fill="#CAE4EB", width=2))
            for y in range(25, 35, 5):
                node = self.canvas.create_oval(377, y-3, 383, y+3, fill="#020638", outline="#CAE4EB")
                self.border_elements.append(node)
        
        def create_glitch_border(self):
            self.border_elements = []
            
            segments = [(20, 10), (80, 10), (120, 10), (180, 10), (240, 10), (320, 10), (380, 10)]
            for i in range(len(segments) - 1):
                if random.random() > 0.3:  
                    line = self.canvas.create_line(
                        segments[i][0], segments[i][1],
                        segments[i+1][0], segments[i+1][1],
                        fill="#CAE4EB", width=2
                    )
                    self.border_elements.append(line)
            
            segments = [(20, 50), (70, 50), (130, 50), (190, 50), (250, 50), (330, 50), (380, 50)]
            for i in range(len(segments) - 1):
                if random.random() > 0.3:  
                    line = self.canvas.create_line(
                        segments[i][0], segments[i][1],
                        segments[i+1][0], segments[i+1][1],
                        fill="#CAE4EB", width=2
                    )
                    self.border_elements.append(line)
            
            segments = [(20, 10), (20, 20), (20, 30), (20, 40), (20, 50)]
            for i in range(len(segments) - 1):
                if random.random() > 0.3:  
                    line = self.canvas.create_line(
                        segments[i][0], segments[i][1],
                        segments[i+1][0], segments[i+1][1],
                        fill="#CAE4EB", width=2
                    )
                    self.border_elements.append(line)
            
            segments = [(380, 10), (380, 20), (380, 30), (380, 40), (380, 50)]
            for i in range(len(segments) - 1):
                if random.random() > 0.3:  
                    line = self.canvas.create_line(
                        segments[i][0], segments[i][1],
                        segments[i+1][0], segments[i+1][1],
                        fill="#CAE4EB", width=2
                    )
                    self.border_elements.append(line)
        
        def create_segmented_border(self):
            self.border_elements = []
            
            for x in range(20, 380, 15):
                line = self.canvas.create_line(x, 10, min(x+10, 380), 10, fill="#CAE4EB", width=2)
                self.border_elements.append(line)
            
            for x in range(20, 380, 15):
                line = self.canvas.create_line(x, 50, min(x+10, 380), 50, fill="#CAE4EB", width=2)
                self.border_elements.append(line)
            
            for y in range(10, 50, 10):
                line = self.canvas.create_line(20, y, 20, min(y+5, 50), fill="#CAE4EB", width=2)
                self.border_elements.append(line)
            
            for y in range(10, 50, 10):
                line = self.canvas.create_line(380, y, 380, min(y+5, 50), fill="#CAE4EB", width=2)
                self.border_elements.append(line)
        
        def create_corner_border(self):
            self.border_elements = []
            
            self.border_elements.append(self.canvas.create_line(20, 10, 60, 10, fill="#CAE4EB", width=2))
            self.border_elements.append(self.canvas.create_line(20, 10, 20, 50, fill="#CAE4EB", width=2))
            
            self.border_elements.append(self.canvas.create_line(340, 10, 380, 10, fill="#CAE4EB", width=2))
            self.border_elements.append(self.canvas.create_line(380, 10, 380, 50, fill="#CAE4EB", width=2))
            
            self.border_elements.append(self.canvas.create_line(20, 50, 60, 50, fill="#CAE4EB", width=2))
            
            self.border_elements.append(self.canvas.create_line(340, 50, 380, 50, fill="#CAE4EB", width=2))
        
        def create_default_border(self):
            self.border_elements = []
            self.border_elements.append(self.canvas.create_line(20, 10, 380, 10, fill="#CAE4EB", width=2))
            self.border_elements.append(self.canvas.create_line(20, 50, 380, 50, fill="#CAE4EB", width=2))
            self.border_elements.append(self.canvas.create_line(20, 10, 20, 50, fill="#CAE4EB", width=2))
            self.border_elements.append(self.canvas.create_line(380, 10, 380, 50, fill="#CAE4EB", width=2))
            
            corner_size = 10
            self.corners = []
            corners = [
                (20, 10, 20+corner_size, 10+corner_size),
                (380-corner_size, 10, 380, 10+corner_size),
                (20, 50-corner_size, 20+corner_size, 50),
                (380-corner_size, 50-corner_size, 380, 50)
            ]
            for x1, y1, x2, y2 in corners:
                corner = self.canvas.create_arc(x1, y1, x2, y2, start=0, extent=90, outline="#CAE4EB", width=2, style="arc")
                self.corners.append(corner)
                self.border_elements.append(corner)
        
        def init_matrix_rain(self):
            tamil_chars = "01அஆஇஈஉஊஎஏஐஒஓக்ஙசஞடணதநபமயரறலளழவ"
            for i in range(12):  
                x = 70 + i * 25  
                y = 10
                char = self.canvas.create_text(
                    x, y, text=random.choice(tamil_chars),
                    font=("Consolas", 8), fill="#CAE4EB"
                )
                self.matrix_chars.append(char)
                self.drops.append({'char': char, 'y': y, 'speed': random.uniform(0.8, 2.5)})  
        
        def stop_animation(self):
            """Stop all animations for this button"""
            self.animation_running = False
            if self.after_id:
                self.window.after_cancel(self.after_id)
                self.after_id = None
        
        def on_enter(self, event):
            if not self.is_hovered:
                self.is_hovered = True
                self.canvas.itemconfig(self.btn_text, fill="#007BB4")
                self.canvas.itemconfig(self.text_bg, fill="#0A1929", outline="#007BB4")
                for element in self.border_elements:
                    if self.canvas.type(element) == "line":
                        self.canvas.itemconfig(element, fill="#007BB4", width=3)
                    elif self.canvas.type(element) == "oval":
                        self.canvas.itemconfig(element, outline="#007BB4")
                    elif self.canvas.type(element) == "arc":
                        self.canvas.itemconfig(element, outline="#007BB4", width=3)
        
        def on_leave(self, event):
            self.is_hovered = False
            self.canvas.itemconfig(self.btn_text, fill="#CAE4EB")
            self.canvas.itemconfig(self.text_bg, fill="#020638", outline="#CAE4EB")
            for element in self.border_elements:
                if self.canvas.type(element) == "line":
                    self.canvas.itemconfig(element, fill="#CAE4EB", width=2)
                elif self.canvas.type(element) == "oval":
                    self.canvas.itemconfig(element, outline="#CAE4EB")
                elif self.canvas.type(element) == "arc":
                    self.canvas.itemconfig(element, outline="#CAE4EB", width=2)
        
        def on_click(self, event):
            self.stop_animation()
            
            for drop in self.drops:
                try:
                    self.canvas.itemconfig(drop['char'], fill="#FFFFFF")
                except:
                    pass  
            
            for element in self.border_elements:
                try:
                    if self.canvas.type(element) == "line":
                        self.canvas.itemconfig(element, fill="#FFFFFF", width=4)
                    elif self.canvas.type(element) == "oval":
                        self.canvas.itemconfig(element, outline="#FFFFFF")
                    elif self.canvas.type(element) == "arc":
                        self.canvas.itemconfig(element, outline="#FFFFFF", width=4)
                except:
                    pass  
            
            try:
                self.canvas.itemconfig(self.text_bg, fill="#1E3A5F", outline="#FFFFFF")
            except:
                pass  
            
            self.window.after(100, self.command)
        
        def animate_matrix(self):
            if not self.animation_running:
                return
                
            for drop in self.drops:
                drop['y'] += drop['speed']
                
                if drop['y'] > 50:
                    drop['y'] = 10
                    tamil_chars = "அஆஇஈஉஊஎஏஐஒஓக்ஙசஞடணதநபமயரறலளழவ"
                    try:
                        self.canvas.itemconfig(drop['char'], 
                                            text=random.choice(tamil_chars))
                    except:
                        pass  
                
                try:
                    x = 70 + self.drops.index(drop) * 25  
                    self.canvas.coords(drop['char'], x, drop['y'])
                except:
                    pass  
                
                fade = max(0, min(255, int(255 * (1 - (drop['y'] - 10) / 40))))
                if self.is_hovered:
                    color = f"#{0:02x}{fade:02x}{fade:02x}"  
                else:
                    color = f"#{0:02x}{int(fade*0.8):02x}{int(fade*0.8):02x}"  
                try:
                    self.canvas.itemconfig(drop['char'], fill=color)
                except:
                    pass  
            
            if self.animation_running:
                self.after_id = self.window.after(30, self.animate_matrix)


#------------------------------------------------------------------------------------------------------------------------------------------------------------------
  

    def on_close():
        lock_file = os.path.join(tempfile.gettempdir(), "EDA_APP.lock")
        if os.path.exists(lock_file):
            os.remove(lock_file)
        global cap, after_id

        if cap:
            cap.release()
            cap = None

        if after_id:
            window.after_cancel(after_id)
            after_id = None

        PurposeListStr = ",".join(PurposeList)

        def record_user_txt():
            try:
                log_dir = Path(
                    r"\\sng-psrvr06\Automation-Public\MiscFiles\SKILL_Logs\Application_Logs"
                )
                log_dir.mkdir(parents=True, exist_ok=True)

                username = getpass.getuser()
                now = datetime.now()
                date_str = now.strftime("%Y-%m-%d")
                time_str = now.strftime("%H-%M-%S")

                file_name = f"{username}_{date_str}_{time_str}.txt"
                file_path = log_dir / file_name

                with open(file_path, "w", encoding="utf-8") as f:
                    f.write(f"Name      : {full_name}\n")
                    f.write(f"Team      : {GroupListStr}\n")
                    f.write(f"Username  : {username}\n")
                    f.write(f"Date      : {date_str}\n")
                    f.write(f"Time      : {time_str}\n")
                    f.write("Purpose   :\n")

                    for idx, purpose in enumerate(PurposeList, start=1):
                        f.write(f" {idx}. {purpose}\n")

                print(f"✅ TXT log created: {file_path}")

            except Exception as e:
                print(f"❌ Error writing TXT log: {e}")

        record_user_txt()
        window.destroy()


#------------------------------------------------------------------------------------------------------------------------------------------------------------------


    def update_video_frame(label, video_source):
        global after_id
        if cap is None: return
        ret, frame = cap.read()
        if not ret:
            cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
            ret, frame = cap.read()
        if ret:
            try:
                frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                width = window.winfo_width() if window.winfo_width() > 1 else 1280
                height = window.winfo_height() if window.winfo_height() > 1 else 720
                frame = cv2.resize(frame, (width, height))
                frame_tk = ImageTk.PhotoImage(image=Image.fromarray(frame))
                label.config(image=frame_tk)
                label.image = frame_tk
            except tk.TclError:
                return
        after_id = window.after(30, update_video_frame, label, video_source)
        
                                 
#------------------------------------------------------------------------------------------------------------------------------------------------------------------


    def DesignFunc():
        global  cap, after_id

        UsedDt = "Opened - Design Menu"
        PurposeList.append(UsedDt)
        if after_id:
            window.after_cancel(after_id)
            after_id = None
        if cap:
            cap.release()
        for widget in window.winfo_children():
            widget.destroy()

        cap_path = resource_path("VidMain1.mp4")
        cap = cv2.VideoCapture(cap_path)
        bg_label = tk.Label(window)
        bg_label.place(relwidth=1, relheight=1)
        update_video_frame(bg_label, cap_path)

#------------------------------------------------------------------------------------------------------------------------------------------------------------------         
        
        def ExcelComp():
            UsedDt = "Design Menu - Used : Excel Comparison"
            PurposeList.append(UsedDt)
            import openpyxl
            from openpyxl.styles import PatternFill
            from tkinter import filedialog
            from tkinter import messagebox
            import tkinter as tk
            def choose_file(title="Choose a file"):
                root = tk.Tk()
                root.withdraw()  
                file_path = filedialog.askopenfilename(title=title, filetypes=[("Excel files", "*.xlsx")])
                return file_path

            def choose_save_location():
                root = tk.Tk()
                root.withdraw()  
                save_path = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[("Excel files", "*.xlsx")], title="Save the output file as...")
                return save_path

            messagebox.showinfo("Excel Comparison", "Please Select The Old Excel File..!")
            file1_path = choose_file("Select file1.xlsx")  
            if not file1_path:
                print("No file selected for file1. Exiting.")
                exit()
                
            messagebox.showinfo("Excel Comparison", "Please Select The New Excel File..!")
            file2_path = choose_file("Select file2.xlsx")  
            if not file2_path:
                print("No file selected for file2. Exiting.")
                exit()

            try:
                wb1 = openpyxl.load_workbook(file1_path)
                wb2 = openpyxl.load_workbook(file2_path)
            except FileNotFoundError as e:
                print(f"Error: {e}")
                exit()

            highlight_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            differences_found = False

            for sheet_name in wb1.sheetnames:
                if sheet_name in wb2.sheetnames:
                    sheet1 = wb1[sheet_name]
                    sheet2 = wb2[sheet_name]
                    
                    for row in range(1, sheet1.max_row + 1):
                        for col in range(1, sheet1.max_column + 1):
                            data1 = str(sheet1.cell(row=row, column=col).value)
                            data2 = str(sheet2.cell(row=row, column=col).value)
                            
                            if data1 != data2:
                                sheet2.cell(row=row, column=col).fill = highlight_fill
                                differences_found = True

            save_path = choose_save_location()
            if not save_path:
                print("No save location selected. Exiting.")
                exit()

            wb2.save(save_path)

            if differences_found:
                messagebox.showinfo("Excel Comparison", "Comparison Completed! Please check the generated Excel file. Differences have been highlighted in red.")
            else:
                messagebox.showinfo("Excel Comparison", "No differences found! The files are identical.")
            
#------------------------------------------------------------------------------------------------------------------------------------------------------------------     
            
        def CHMapATE1():
            UsedDt = "Design Menu - Used : Channel Mappling ATE-1"
            PurposeList.append(UsedDt)
            for widget in window.winfo_children():
                if isinstance(widget, tk.Label) and hasattr(widget, 'image'):
                    continue  # Keep background video label
                widget.destroy()

            from tkinter import filedialog
            import numpy as np
            import csv
            from tkinter.filedialog import asksaveasfilename
            from tkinter import messagebox
            import pandas as pd
            
            def CHMapping():
                #input1 = entry1.get()
                #NNSkip = input1.upper()
                #if ',' in NNSkip:
                    #NNSkipList = NNSkip.split(",")
                #else:
                    #NNSkipList = []
                    #NNSkipList.append(NNSkip)
                    
                input2 = entry1.get()
                JumpName = input2.upper()
                if ',' in JumpName:
                    JumpNameList = JumpName.split(",")
                else:
                    JumpNameList = []
                    JumpNameList.append(JumpName)
                
                input3 = entry2.get()
                DutName = input3.upper()
                if ',' in DutName:
                    DutNameList = DutName.split(",")
                else:
                    DutNameList = []
                    DutNameList.append(DutName)
                
                import pandas as pd
                import tkinter as tk
                from tkinter import filedialog

                messagebox.showinfo("Channel Mappling ATE-1", "Please Select The Symbol The Pin Report..!")
                def get_file_path():
                    root = tk.Tk()
                    root.withdraw()  
                    file_path = filedialog.askopenfilename(
                        title="Select Excel file", 
                        filetypes=[("Excel files", "*.xlsx")]
                    )
                    return file_path
                file_path = get_file_path()
                df = pd.read_excel(file_path, header=None)
                data = df.iloc[4:, [0, 1, 7]].values.tolist()
                formatted_data = [f"{row[0]}.{row[1]},{row[2]}" for row in data]
                
                import tkinter as tk
                from tkinter import filedialog
                messagebox.showinfo("Channel Mapping ATE-1", "Please Select The GND Skip list..!")

                def Gnd_file():
                    root = tk.Tk()
                    root.withdraw()  
                    Gndfile_path = filedialog.askopenfilename(title="Select a TXT File", filetypes=[("TXT Files", "*.txt")])
                    return Gndfile_path

                Gndfile_path = Gnd_file()
                if Gndfile_path:
                    try:
                        with open(Gndfile_path, 'r') as file:
                            NNSkipListFlt = [line.strip() for line in file.readlines()]
                    except Exception as e:
                        messagebox.showerror("Error", f"Error reading the file: {e}")
                else:
                    messagebox.showwarning("No File Selected", "You did not select any file.")
                
                        
                import tkinter as tk
                from tkinter import filedialog
                messagebox.showinfo("Channel Mapping ATE-1", "Please Select The Connection list If Needed Otherwise Click ok and Click Cancel..!")
                
                def Con_file():
                    root = tk.Tk()
                    root.withdraw()  
                    Confile_path = filedialog.askopenfilename(title="Select a TXT File", filetypes=[("TXT Files", "*.txt")])
                    return Confile_path

                Confile_path = Con_file()
                if Confile_path:
                    try:
                        with open(Confile_path, 'r') as file:
                            ConListUp = [line.strip() for line in file.readlines()]
                    except Exception as e:
                        messagebox.showerror("Error", f"Error reading the file: {e}")
                else:
                    ConData = "None"
                    ConListUp = []
                    ConListUp.append(ConData)
                
                ConList = []
                for ConListlpFlt in ConListUp:
                    if ConListlpFlt != "":
                        ConList.append(ConListlpFlt)
                        

                NNSkipList = []
                for NNSkipListFltlp in NNSkipListFlt:
                    if NNSkipListFltlp != "":
                        NNSkipUp = NNSkipListFltlp.upper()
                        NNSkipList.append(NNSkipUp)
                
                ExcelData = []
                for formatted_datalp in formatted_data:
                    DutNameLst = formatted_datalp.split(",")
                    DutNameDt = DutNameLst[0]
                    for DutNameListlp in DutNameList:
                        if DutNameListlp in DutNameDt:
                            ExcelData.append(formatted_datalp)
                
                ExcelFiltList = []
                for ExcelDatalp in ExcelData:
                    ExNNList = ExcelDatalp.split(",")
                    ExNNDt = ExNNList[1]
                    PassDt = "NO"
                    for NNSkipListlp in NNSkipList:
                        if NNSkipListlp == ExNNDt:
                            PassDt = "YES"
                    if PassDt == "NO":       
                        ExcelFiltList.append(ExcelDatalp) 
                    
                                    
                
                from tkinter import filedialog
                import numpy as np
                messagebox.showinfo("Channel Mappling ATE-1", "Please Select The Design Netlist..!")
                def open_file():
                    file_path = filedialog.askopenfilename(title="Select a TXT File", filetypes=[("TXT Files", "*.txt")])
                    if file_path:
                        try:
                            with open(file_path, 'r') as file:
                                content_list = [line.strip() for line in file.readlines()]
                                net_list = []
                                capturing = False  
                                for line in content_list:
                                    if line == "$NETS":
                                        capturing = True 
                                    elif line == "$PACKAGES": 
                                        capturing = False 
                                    elif line == "$A_PROPERTIES":
                                        capturing = False
                                    if capturing:
                                        net_list.append(line)
                                        
                                #net_list = net_list[1:-1]
                                my_array = np.array(net_list)
                                LenVal = len(net_list)
                                
                                OutputList = []
                                IncVal = 0
                                for net_listlp in net_list:
                                    net_Filt = net_listlp.replace(",", "")
                                    if ";" in net_Filt:
                                        StoreStr = ""
                                        StoreStr = net_Filt
                                        NextIncVal = IncVal + 1
                                        while NextIncVal <= LenVal - 1:
                                            if ";" in my_array[NextIncVal] or "$NETS" in my_array[NextIncVal]:
                                                OutputList.append(StoreStr)
                                                break                                       
                                            else:
                                                StoreStr = StoreStr + " " + my_array[NextIncVal]
                                                StoreStr = StoreStr.replace(",", "") 
                                            NextIncVal += 1
                                        
                                    IncVal = IncVal + 1
                            
                                FinalList = []
                                FinalList = [entry.replace("  ", " ") for entry in OutputList]
                                
                                ResultList = []   
                                for ExcelFiltListlp in ExcelFiltList:
                                    ExcelSplitList = ExcelFiltListlp.split(",")
                                    DutNameEx = ExcelSplitList[0]
                                    for FinalListlp in FinalList:
                                        NLSplitList = FinalListlp.split(";")
                                        NLNetName = NLSplitList[0]
                                        NLPinDt = NLSplitList[1]
                                        NLPinList = NLPinDt.split(" ")
                                        for NLPinListlp in NLPinList:
                                            if NLPinListlp == DutNameEx:
                                                PassDt1 = "NO"
                                                for NNSkipListlp1 in NNSkipList:
                                                    if NNSkipListlp1 == NLNetName: 
                                                        PassDt1 = "YES"
                                                
                                                if PassDt1 == "NO":
                                                    NLName = NLSplitList[1]
                                                    NLNameList = NLName.split(" ")
                                                    PassData = "NO"
                                                    CompList = []
                                                    for NLNameListlp in NLNameList:
                                                        if NLNameListlp != "":  
                                                            if NLNameListlp[0].isdigit() or "POWER_HEADER" in NLNameListlp:
                                                                TesterName = NLNameListlp
                                                                PassData = "YES"
                                                            else:
                                                                if NLNameListlp != DutNameEx:
                                                                    CompList.append(NLNameListlp)
                                                    
                                                    CompData = " ".join(CompList)
                                                    if PassData == "YES":
                                                        NLNetName = NLNetName.replace(" ", "")
                                                        PassString = f"{DutNameEx},{NLNetName},----,{CompData},{TesterName},----"
                                                        PassString_Filt = PassString.replace("'", "")
                                                        ResultList.append(PassString_Filt)
                                                        
                                                    
                                                    if PassData == "NO":
                                                        for JumpNameListlp in JumpNameList:
                                                            JumpNameLen = len(JumpNameListlp)
                                                            JumpNameUp = JumpNameListlp.upper()
                                                            for NLNameListlp1 in NLNameList:
                                                                NLNameLen = (NLNameListlp1[:JumpNameLen])
                                                                NLNameLen = NLNameLen.upper()
                                                                if JumpNameUp == NLNameLen:
                                                                    PassData2 = "NO"
                                                                    for ConListlp in ConList:
                                                                        ConPrse = ConListlp.split(",")
                                                                        ConPrseFirst = ConPrse[0]
                                                                        ConPrseFirst = ConPrseFirst.upper()
                                                                        NLNameListlp1 = NLNameListlp1.upper()
                                                                        if ConPrseFirst == NLNameListlp1:
                                                                            JumperName = ConPrse[1]
                                                                            PassData2 = "Yes"
                                                                            
                                                                    if PassData2 == "NO":
                                                                        JumperList = NLNameListlp1.split(".")
                                                                        JumperName = JumperList[0]
                                                                            
                                                                    for FinalListlp1 in FinalList:
                                                                        if JumperName in FinalListlp1: 
                                                                            if NLNameListlp1 not in FinalListlp1:
                                                                                JNLSplitList = FinalListlp1.split(";")
                                                                                JNLNetName = JNLSplitList[0]
                                                                                JNLName = JNLSplitList[1]
                                                                                JNLNameList = JNLName.split(" ")
                                                                                for JNLNameListlp in JNLNameList:
                                                                                    if JNLNameListlp != "":  
                                                                                        if JNLNameListlp[0].isdigit():
                                                                                            NLNetName = NLNetName.replace(" ", "")
                                                                                            JNLNetName = JNLNetName.replace(" ", "")
                                                                                            JPassString = f"{DutNameEx},{NLNetName},{JNLNetName},{NLNameListlp1},----,{JNLNameListlp}"
                                                                                            JPassString_Filt = JPassString.replace("'", "")
                                                                                            ResultList.append(JPassString_Filt)



                                ResultList = sorted(set(ResultList))
                                    
                                import csv
                                from tkinter import Tk
                                from tkinter.filedialog import asksaveasfilename
                                Tk().withdraw()  
                                SaveFile = asksaveasfilename(defaultextension='.csv', filetypes=[('CSV Files', '*.csv'), ('All Files', '*.*')],
                                                            title="Save CSV File")

                                if SaveFile:
                                    with open(SaveFile, mode='w', newline='') as file:
                                        writer = csv.writer(file)
                                        for line in ResultList:
                                            writer.writerow(line.split(','))  
                                        
                                    print(f"File saved to {SaveFile}")
                                else:
                                    print("No file selected.")

                        except Exception as e:
                            print(f"Error opening the file: {e}")
                    else:
                        print("No file was selected.")
                open_file()
                
            def create_button(text, relx, rely, command):
                btn = tk.Label(window, text=text, font=("Trebuchet MS", 20), fg="#F6FAFC", bg="#020638", bd=0)
                btn.bind("<Enter>", lambda e: btn.config(fg="#1F5186"))
                btn.bind("<Leave>", lambda e: btn.config(fg="#F6FAFC"))
                btn.bind("<Button-1>", lambda e: command())
                btn.place(relx=relx, rely=rely, anchor="center")
                return btn
            
            form_start_y = 0.25
            spacing = 0.12
            entry_width = 30

            def create_label_entry(text, rel_y):
                label = tk.Label(window, text=text, font=("Didot", 12), bg="#f0f0f0")
                label.place(relx=0.5, rely=rel_y, anchor="center")
                entry = tk.Entry(window, font=("Didot", 12), width=entry_width)
                entry.place(relx=0.5, rely=rel_y + 0.05, anchor="center")
                return entry

            entry1 = create_label_entry("[ ENTER JUMP NAMES ]", form_start_y)
            entry2 = create_label_entry("[ ENTER DUT COMMON REDES ]", form_start_y + spacing)
            
            create_button("[ CLICK HERE ]", 0.5, form_start_y + spacing * 2.5, CHMapping)
            create_button("[<-- BACK ]", 0.1, 0.05, DesignFunc)
        
#------------------------------------------------------------------------------------------------------------------------------------------------------------------
    
        def CompPinRepFilt():
            UsedDt = "Design Menu - Used : PIN Report MLO"
            PurposeList.append(UsedDt)
            
            import pandas as pd
            import tkinter as tk
            from tkinter import filedialog, messagebox
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment

            # Show initial message
            messagebox.showinfo("PIN Report MLO", "Please Select The Symbol The Pin Report..!")

            def get_file_path():
                root = tk.Tk()
                root.withdraw()
                file_path = filedialog.askopenfilename(
                    title="Select Excel file",
                    filetypes=[("Excel files", "*.xlsx")]
                )
                return file_path

            file_path = get_file_path()
            if not file_path:
                messagebox.showerror("Error", "No file selected. Program will exit.")
                exit()

            # Read data from Excel
            df = pd.read_excel(file_path, header=None)
            data = df.iloc[1:, [0, 1, 4]].values.tolist()
            formatted_data = [f"{row[0]}.{row[1]},{row[2]}" for row in data]
            Net_data = [f"{row[2]}" for row in data]
            Net_data_Unq = sorted(set(Net_data))

            UpdatedList = []
            for Net_data_Unqlp in Net_data_Unq:
                TempDieList = []
                TempBGAList = []
                for formatted_datalp in formatted_data:
                    ListNetName = formatted_datalp.split(",")
                    ListPinName = formatted_datalp.split(".")
                    if ListNetName[1] == Net_data_Unqlp:
                        if "DIE" in ListPinName[0]:
                            TempDieList.append(ListNetName[0])
                        if "BGA" in ListPinName[0]:
                            TempBGAList.append(ListNetName[0])

                DIELen = len(TempDieList)
                BGALen = len(TempBGAList)
                if DIELen > BGALen:
                    for _ in range(DIELen - BGALen):
                        TempBGAList.append("---")
                else:
                    for _ in range(BGALen - DIELen):
                        TempDieList.append("---")

                for k in range(len(TempBGAList)):
                    PassData = f"{Net_data_Unqlp},{TempDieList[k]},{TempBGAList[k]},{Net_data_Unqlp}"
                    UpdatedList.append(PassData)

            if len(UpdatedList) > 0:
                df1 = pd.DataFrame([item.split(',') for item in UpdatedList],
                                columns=['NET_NAME_DIE', 'DIE_PIN', 'BGA_PIN', 'NET_NAME_BGA'])

                # Replace '---' in BGA_PIN column (Column C) with empty string
                df1['BGA_PIN'] = df1['BGA_PIN'].replace('---', '')

                # If DIE_PIN is '---', empty DIE_PIN and NET_NAME_DIE
                df1.loc[df1['DIE_PIN'] == '---', ['DIE_PIN', 'NET_NAME_DIE']] = ''

                # Move rows where A and B are empty to the bottom (C and D values only)
                move_rows = df1[(df1['NET_NAME_DIE'] == '') & (df1['DIE_PIN'] == '')][['BGA_PIN', 'NET_NAME_BGA']].copy()
                df1 = df1[~((df1['NET_NAME_DIE'] == '') & (df1['DIE_PIN'] == ''))]  # Remove them from main DataFrame

                for _, row in move_rows.iterrows():
                    new_row = pd.Series({'NET_NAME_DIE': '', 'DIE_PIN': '', 'BGA_PIN': row['BGA_PIN'], 'NET_NAME_BGA': row['NET_NAME_BGA']})
                    df1 = pd.concat([df1, pd.DataFrame([new_row])], ignore_index=True)

                # Optionally simplify NET_NAME_BGA display
                previous_value = None
                for idx, row in df1.iterrows():
                    current_value = row['NET_NAME_BGA']
                    if current_value == previous_value:
                        df1.at[idx, 'NET_NAME_BGA'] = current_value
                    else:
                        previous_value = current_value

                # Save file dialog
                root = tk.Tk()
                root.withdraw()
                save_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                        filetypes=[("Excel files", "*.xlsx")])
                if save_path:
                    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                        df1.to_excel(writer, sheet_name='BGA LIST', index=False)

                    wb = load_workbook(save_path)
                    ws = wb['BGA LIST']

                    # Merge NET_NAME_BGA cells with same value
                    previous_value = None
                    start_row = 2
                    for row in range(2, len(df1) + 2):
                        current_value = ws.cell(row=row, column=4).value
                        if current_value != previous_value:
                            if row > start_row:
                                ws.merge_cells(start_row=start_row, start_column=4, end_row=row - 1, end_column=4)
                                ws.cell(row=start_row, column=4).alignment = Alignment(horizontal='left', vertical='top')
                            start_row = row
                            previous_value = current_value

                    if start_row <= len(df1) + 1:
                        ws.merge_cells(start_row=start_row, start_column=4, end_row=len(df1) + 1, end_column=4)
                        ws.cell(row=start_row, column=4).alignment = Alignment(horizontal='left', vertical='top')

                    wb.save(save_path)
                    messagebox.showinfo("PIN Report MLO", f"Result successfully saved to {save_path}")
                else:
                    messagebox.showinfo("PIN Report MLO", "No file selected. Data not saved.")
            else:
                messagebox.showinfo("PIN Report MLO", "No Error Found...!")

#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def MenuPass():
            from tkinter import messagebox
            messagebox.showinfo("Allegro Menu", "Please Choose Allegro Skill Menu According To Your Team..")

#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def ExcelFunc():
            UsedDt = "Design Menu - Added : Excel Menu"
            PurposeList.append(UsedDt)
            
            import os
            import shutil
            import getpass
            import time
            import win32com.client
            from tkinter import messagebox

            SOURCE_PATH = r"\\sng-psrvr06\Cadence_Skills\Menu-enc\MyCustomMenu.xlam"

            USERNAME = getpass.getuser()

            DEST_FOLDER = rf"C:\Users\{USERNAME}\AppData\Roaming\Microsoft\AddIns\Menu"
            DEST_PATH = os.path.join(DEST_FOLDER, os.path.basename(SOURCE_PATH))

            if not os.path.exists(SOURCE_PATH):
                raise FileNotFoundError(f"Source add-in not found: {SOURCE_PATH}")
            print(f"✅ Source file found: {SOURCE_PATH}")

            os.makedirs(DEST_FOLDER, exist_ok=True)
            print(f"✅ Destination folder ready: {DEST_FOLDER}")

            if os.path.exists(DEST_PATH):
                print(f"⚠️ Existing add-in found, replacing: {DEST_PATH}")
                os.remove(DEST_PATH)  
            shutil.copy2(SOURCE_PATH, DEST_PATH)
            print(f"✅ Copied new add-in to: {DEST_PATH}")

            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False 

            excel.Workbooks.Add()
            time.sleep(1)

            addin_name = os.path.basename(DEST_PATH).lower()
            existing_addin = None

            for a in excel.AddIns:
                if os.path.basename(a.FullName).lower() == addin_name:
                    existing_addin = a
                    break

            if existing_addin:
                try:
                    print(f"🔄 Disabling old version of add-in: {existing_addin.FullName}")
                    existing_addin.Installed = False
                    del existing_addin
                    time.sleep(0.5)
                except Exception as e:
                    print(f"⚠️ Warning: could not disable old version: {e}")


            try:
                print(f"➕ Registering new add-in: {DEST_PATH}")
                new_addin = excel.AddIns.Add(DEST_PATH)
                new_addin.Installed = True
                print(f"✅ '{new_addin.Name}' successfully installed and enabled in Excel!")

            except Exception as e:
                print(f"❌ Failed to add or enable add-in: {e}")
                print("➡️ Check if the add-in is in a Trusted Location or not blocked (Properties → Unblock).")

            finally:
                excel.Quit()

            messagebox.showinfo("Excel Update", "✅ EDA Excel Menu.\n -- Updated Successfully --")

#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def AlignFunc():
            UsedDt = "Design Menu - Added : Align"
            PurposeList.append(UsedDt)
            from pathlib import Path
            import shutil
            import getpass
            from tkinter import messagebox

            base_path = Path(r"C:\Cadence")
            source_path = Path(r"\\sng-psrvr06\Cadence_Skills\Menu-enc")

            if not source_path.exists():
                messagebox.showerror("Menu Update", "Error: Network path not found. Contact Automation Team..!")
                return

            form_file = "align.form"
            icon_files = [
                "align-bottom.bmp", "align-centerH.bmp", "align-centerV.bmp",
                "align-left.bmp", "align-right.bmp", "align-top.bmp",
                "distributeH.bmp", "distributeV.bmp", "pickH.bmp", "pickV.bmp"
            ]

            for folder in base_path.iterdir():
                if folder.is_dir() and "SPB_" in folder.name:
                    forms_path = folder / "share" / "pcb" / "text" / "forms"
                    icons_path = folder / "share" / "pcb" / "text" / "icons"

                    if forms_path.exists():
                        src_form = source_path / form_file
                        dest_form = forms_path / form_file
                        try:
                            if src_form.exists():
                                shutil.copy2(src_form, dest_form)
                        except Exception as e:
                            messagebox.showerror("Menu Update", f"Error copying {form_file}: {e}")

                    if icons_path.exists():
                        for icon in icon_files:
                            src_icon = source_path / icon
                            dest_icon = icons_path / icon
                            try:
                                if src_icon.exists():
                                    shutil.copy2(src_icon, dest_icon)
                            except Exception as e:
                                messagebox.showerror("Menu Update", f"Error copying {icon}: {e}")
                                
            messagebox.showinfo("Align Update", "Please Restart the Tool and Check -- If Align was not updated Contact Automation Team..!")
    
#------------------------------------------------------------------------------------------------------------------------------------------------------------------    

        def ATE1Func():
            UsedDt = "Design Menu - Added : ATE - 1 Menu"
            PurposeList.append(UsedDt)
            from pathlib import Path
            from datetime import datetime
            import shutil
            import getpass
            import os
            import tkinter as tk
            from tkinter import messagebox

            # --- Initialize Tkinter safely (to avoid 'no default root' errors) ---
            root = tk.Tk()
            root.withdraw()  # Hide the main window

            base_path = Path(r"C:\Cadence")

            insert_text = """POPUP "&EDA_ind"
        BEGIN
            MENUITEM "add_menu",           "add_menu"
        END"""

            insert_text1 = """POPUP "&ATE-1"
        BEGIN
            MENUITEM "ate1_menu",           "ate1_menu"
        END"""

            # --- Update Allegro menu definitions ---
            for folder in base_path.iterdir():
                if folder.is_dir() and "SPB_" in folder.name:
                    link_val = folder / "share" / "pcb" / "text" / "cuimenus"
                    allegro_file = link_val / "allegro.men"

                    if allegro_file.exists():
                        content = allegro_file.read_text(encoding="utf-8")

                        has_EDA = 'POPUP "&EDA_ind"' in content
                        has_ate1 = 'POPUP "&ATE-1"' in content

                        if has_EDA and has_ate1:
                            continue  # Skip already updated files

                        # Backup original
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        backup_file = link_val / f"allegro_Backup_{timestamp}.txt"
                        shutil.copy2(allegro_file, backup_file)

                        # Find insertion point
                        lines = content.splitlines()
                        insert_index = next((i for i in range(len(lines)-1, -1, -1) if lines[i].strip() == "END"), len(lines))

                        updated_lines = lines[:insert_index]

                        if not has_EDA:
                            updated_lines += insert_text.splitlines()
                        if not has_ate1:
                            updated_lines += insert_text1.splitlines()

                        updated_lines.append("END")

                        allegro_file.write_text("\n".join(updated_lines) + "\n", encoding="utf-8")

            # --- Source and destination paths ---
            source_path = Path(r"\\sng-psrvr06\Cadence_Skills\Menu-enc")

            allegro_env = os.environ.get("ALLEGRO_PCBENV")
            home_env = os.environ.get("HOME") or os.environ.get("USERPROFILE")

            # Ensure we have a valid destination path
            if allegro_env:
                dest_path = Path(allegro_env)
            elif home_env:
                dest_path = Path(home_env) / "pcbenv"
            else:
                messagebox.showerror("Menu Update", "Cannot determine user environment path.")
                return

            files_to_copy = ["allegro.ilinit", "env"]

            allowed_lines = [
                'load("//sng-psrvr06/Cadence_Skills/Menu-enc/add0_menu0.ile")',
                'load("//sng-psrvr06/Cadence_Skills/Menu-enc/ate1_menu1.ile")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align.ile" "EncFileKar")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align_do.ile" "EncFileKar")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align_utils.ile" "EncFileKar")',
            ]

            # --- Copy and clean up files ---
            if not source_path.exists():
                messagebox.showerror("Menu Update", "Error: Source path not found. Contact Automation Team.")
                return

            dest_path.mkdir(parents=True, exist_ok=True)

            for file_name in files_to_copy:
                src_file = source_path / file_name
                dest_file = dest_path / file_name

                if src_file.exists():
                    # Backup existing file
                    if dest_file.exists():
                        backup_file = dest_file.with_suffix(dest_file.suffix + ".bak")
                        shutil.copy2(dest_file, backup_file)

                    # Copy new file
                    shutil.copy2(src_file, dest_file)

                    # Clean up allegro.ilinit after copy
                    if file_name.lower() == "allegro.ilinit":
                        try:
                            lines = dest_file.read_text(encoding="utf-8").splitlines()
                            filtered_lines = [line for line in lines if line.strip() in allowed_lines]
                            dest_file.write_text("\n".join(filtered_lines) + "\n", encoding="utf-8")
                        except Exception as e:
                            messagebox.showerror("Menu Update", f"Error updating allegro.ilinit: {e}")
                else:
                    messagebox.showerror("Menu Update", f"Missing file: {src_file}. Contact Automation Team.")

            messagebox.showinfo("Menu Update", "✅ Please restart Allegro.\nIf SKILL was not updated, contact Automation Team.")

            root.destroy()

#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def ATE2Func():
            UsedDt = "Design Menu - Added : ATE - 2 Menu"
            PurposeList.append(UsedDt)
            from pathlib import Path
            from datetime import datetime
            import shutil
            import getpass
            import os
            import tkinter as tk
            from tkinter import messagebox

            # --- Initialize Tkinter safely (to avoid 'no default root' errors) ---
            root = tk.Tk()
            root.withdraw()  # Hide the main window

            base_path = Path(r"C:\Cadence")

            insert_text = """POPUP "&EDA_ind"
        BEGIN
            MENUITEM "add_menu",           "add_menu"
        END"""

            insert_text1 = """POPUP "&ATE-2"
        BEGIN
            MENUITEM "ate2_menu",           "ate2_menu"
        END"""

            # --- Update Allegro menu definitions ---
            for folder in base_path.iterdir():
                if folder.is_dir() and "SPB_" in folder.name:
                    link_val = folder / "share" / "pcb" / "text" / "cuimenus"
                    allegro_file = link_val / "allegro.men"

                    if allegro_file.exists():
                        content = allegro_file.read_text(encoding="utf-8")

                        has_EDA = 'POPUP "&EDA_ind"' in content
                        has_ate1 = 'POPUP "&ATE-2"' in content

                        if has_EDA and has_ate1:
                            continue  # Skip already updated files

                        # Backup original
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        backup_file = link_val / f"allegro_Backup_{timestamp}.txt"
                        shutil.copy2(allegro_file, backup_file)

                        # Find insertion point
                        lines = content.splitlines()
                        insert_index = next((i for i in range(len(lines)-1, -1, -1) if lines[i].strip() == "END"), len(lines))

                        updated_lines = lines[:insert_index]

                        if not has_EDA:
                            updated_lines += insert_text.splitlines()
                        if not has_ate1:
                            updated_lines += insert_text1.splitlines()

                        updated_lines.append("END")

                        allegro_file.write_text("\n".join(updated_lines) + "\n", encoding="utf-8")

            # --- Source and destination paths ---
            source_path = Path(r"\\sng-psrvr06\Cadence_Skills\Menu-enc")

            allegro_env = os.environ.get("ALLEGRO_PCBENV")
            home_env = os.environ.get("HOME") or os.environ.get("USERPROFILE")

            # Ensure we have a valid destination path
            if allegro_env:
                dest_path = Path(allegro_env)
            elif home_env:
                dest_path = Path(home_env) / "pcbenv"
            else:
                messagebox.showerror("Menu Update", "Cannot determine user environment path.")
                return

            files_to_copy = ["allegro.ilinit", "env"]

            allowed_lines = [
                'load("//sng-psrvr06/Cadence_Skills/Menu-enc/add0_menu0.ile")',
                'load("//sng-psrvr06/Cadence_Skills/Menu-enc/ate2_menu2.ile")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align.ile" "EncFileKar")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align_do.ile" "EncFileKar")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align_utils.ile" "EncFileKar")',
            ]

            # --- Copy and clean up files ---
            if not source_path.exists():
                messagebox.showerror("Menu Update", "Error: Source path not found. Contact Automation Team.")
                return

            dest_path.mkdir(parents=True, exist_ok=True)

            for file_name in files_to_copy:
                src_file = source_path / file_name
                dest_file = dest_path / file_name

                if src_file.exists():
                    # Backup existing file
                    if dest_file.exists():
                        backup_file = dest_file.with_suffix(dest_file.suffix + ".bak")
                        shutil.copy2(dest_file, backup_file)

                    # Copy new file
                    shutil.copy2(src_file, dest_file)

                    # Clean up allegro.ilinit after copy
                    if file_name.lower() == "allegro.ilinit":
                        try:
                            lines = dest_file.read_text(encoding="utf-8").splitlines()
                            filtered_lines = [line for line in lines if line.strip() in allowed_lines]
                            dest_file.write_text("\n".join(filtered_lines) + "\n", encoding="utf-8")
                        except Exception as e:
                            messagebox.showerror("Menu Update", f"Error updating allegro.ilinit: {e}")
                else:
                    messagebox.showerror("Menu Update", f"Missing file: {src_file}. Contact Automation Team.")

            messagebox.showinfo("Menu Update", "✅ Please restart Allegro.\nIf SKILL was not updated, contact Automation Team.")

            root.destroy()

#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def ATE3Func():
            UsedDt = "Design Menu - Added : ATE - 3 Menu"
            PurposeList.append(UsedDt)
            from pathlib import Path
            from datetime import datetime
            import shutil
            import getpass
            import os
            import tkinter as tk
            from tkinter import messagebox

            # --- Initialize Tkinter safely (to avoid 'no default root' errors) ---
            root = tk.Tk()
            root.withdraw()  # Hide the main window

            base_path = Path(r"C:\Cadence")

            insert_text = """POPUP "&EDA_ind"
        BEGIN
            MENUITEM "add_menu",           "add_menu"
        END"""

            insert_text1 = """POPUP "&ATE-3"
        BEGIN
            MENUITEM "ate3_menu",           "ate3_menu"
        END"""

            # --- Update Allegro menu definitions ---
            for folder in base_path.iterdir():
                if folder.is_dir() and "SPB_" in folder.name:
                    link_val = folder / "share" / "pcb" / "text" / "cuimenus"
                    allegro_file = link_val / "allegro.men"

                    if allegro_file.exists():
                        content = allegro_file.read_text(encoding="utf-8")

                        has_EDA = 'POPUP "&EDA_ind"' in content
                        has_ate1 = 'POPUP "&ATE-3"' in content

                        if has_EDA and has_ate1:
                            continue  # Skip already updated files

                        # Backup original
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        backup_file = link_val / f"allegro_Backup_{timestamp}.txt"
                        shutil.copy2(allegro_file, backup_file)

                        # Find insertion point
                        lines = content.splitlines()
                        insert_index = next((i for i in range(len(lines)-1, -1, -1) if lines[i].strip() == "END"), len(lines))

                        updated_lines = lines[:insert_index]

                        if not has_EDA:
                            updated_lines += insert_text.splitlines()
                        if not has_ate1:
                            updated_lines += insert_text1.splitlines()

                        updated_lines.append("END")

                        allegro_file.write_text("\n".join(updated_lines) + "\n", encoding="utf-8")

            # --- Source and destination paths ---
            source_path = Path(r"\\sng-psrvr06\Cadence_Skills\Menu-enc")

            allegro_env = os.environ.get("ALLEGRO_PCBENV")
            home_env = os.environ.get("HOME") or os.environ.get("USERPROFILE")

            # Ensure we have a valid destination path
            if allegro_env:
                dest_path = Path(allegro_env)
            elif home_env:
                dest_path = Path(home_env) / "pcbenv"
            else:
                messagebox.showerror("Menu Update", "Cannot determine user environment path.")
                return

            files_to_copy = ["allegro.ilinit", "env"]

            allowed_lines = [
                'load("//sng-psrvr06/Cadence_Skills/Menu-enc/add0_menu0.ile")',
                'load("//sng-psrvr06/Cadence_Skills/Menu-enc/ate3_menu3.ile")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align.ile" "EncFileKar")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align_do.ile" "EncFileKar")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align_utils.ile" "EncFileKar")',
            ]

            # --- Copy and clean up files ---
            if not source_path.exists():
                messagebox.showerror("Menu Update", "Error: Source path not found. Contact Automation Team.")
                return

            dest_path.mkdir(parents=True, exist_ok=True)

            for file_name in files_to_copy:
                src_file = source_path / file_name
                dest_file = dest_path / file_name

                if src_file.exists():
                    # Backup existing file
                    if dest_file.exists():
                        backup_file = dest_file.with_suffix(dest_file.suffix + ".bak")
                        shutil.copy2(dest_file, backup_file)

                    # Copy new file
                    shutil.copy2(src_file, dest_file)

                    # Clean up allegro.ilinit after copy
                    if file_name.lower() == "allegro.ilinit":
                        try:
                            lines = dest_file.read_text(encoding="utf-8").splitlines()
                            filtered_lines = [line for line in lines if line.strip() in allowed_lines]
                            dest_file.write_text("\n".join(filtered_lines) + "\n", encoding="utf-8")
                        except Exception as e:
                            messagebox.showerror("Menu Update", f"Error updating allegro.ilinit: {e}")
                else:
                    messagebox.showerror("Menu Update", f"Missing file: {src_file}. Contact Automation Team.")

            messagebox.showinfo("Menu Update", "✅ Please restart Allegro.\nIf SKILL was not updated, contact Automation Team.")

            root.destroy()

#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def ATE4Func():
            UsedDt = "Design Menu - Added : ATE - 4 Menu"
            PurposeList.append(UsedDt)
            from pathlib import Path
            from datetime import datetime
            import shutil
            import getpass
            import os
            import tkinter as tk
            from tkinter import messagebox

            # --- Initialize Tkinter safely (to avoid 'no default root' errors) ---
            root = tk.Tk()
            root.withdraw()  # Hide the main window

            base_path = Path(r"C:\Cadence")

            insert_text = """POPUP "&EDA_ind"
        BEGIN
            MENUITEM "add_menu",           "add_menu"
        END"""

            insert_text1 = """POPUP "&ATE-4"
        BEGIN
            MENUITEM "ate4_menu",           "ate4_menu"
        END"""

            # --- Update Allegro menu definitions ---
            for folder in base_path.iterdir():
                if folder.is_dir() and "SPB_" in folder.name:
                    link_val = folder / "share" / "pcb" / "text" / "cuimenus"
                    allegro_file = link_val / "allegro.men"

                    if allegro_file.exists():
                        content = allegro_file.read_text(encoding="utf-8")

                        has_EDA = 'POPUP "&EDA_ind"' in content
                        has_ate1 = 'POPUP "&ATE-4"' in content

                        if has_EDA and has_ate1:
                            continue  # Skip already updated files

                        # Backup original
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        backup_file = link_val / f"allegro_Backup_{timestamp}.txt"
                        shutil.copy2(allegro_file, backup_file)

                        # Find insertion point
                        lines = content.splitlines()
                        insert_index = next((i for i in range(len(lines)-1, -1, -1) if lines[i].strip() == "END"), len(lines))

                        updated_lines = lines[:insert_index]

                        if not has_EDA:
                            updated_lines += insert_text.splitlines()
                        if not has_ate1:
                            updated_lines += insert_text1.splitlines()

                        updated_lines.append("END")

                        allegro_file.write_text("\n".join(updated_lines) + "\n", encoding="utf-8")

            # --- Source and destination paths ---
            source_path = Path(r"\\sng-psrvr06\Cadence_Skills\Menu-enc")

            allegro_env = os.environ.get("ALLEGRO_PCBENV")
            home_env = os.environ.get("HOME") or os.environ.get("USERPROFILE")

            # Ensure we have a valid destination path
            if allegro_env:
                dest_path = Path(allegro_env)
            elif home_env:
                dest_path = Path(home_env) / "pcbenv"
            else:
                messagebox.showerror("Menu Update", "Cannot determine user environment path.")
                return

            files_to_copy = ["allegro.ilinit", "env"]

            allowed_lines = [
                'load("//sng-psrvr06/Cadence_Skills/Menu-enc/add0_menu0.ile")',
                'load("//sng-psrvr06/Cadence_Skills/Menu-enc/ate4_menu4.ile")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align.ile" "EncFileKar")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align_do.ile" "EncFileKar")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align_utils.ile" "EncFileKar")',
            ]

            # --- Copy and clean up files ---
            if not source_path.exists():
                messagebox.showerror("Menu Update", "Error: Source path not found. Contact Automation Team.")
                return

            dest_path.mkdir(parents=True, exist_ok=True)

            for file_name in files_to_copy:
                src_file = source_path / file_name
                dest_file = dest_path / file_name

                if src_file.exists():
                    # Backup existing file
                    if dest_file.exists():
                        backup_file = dest_file.with_suffix(dest_file.suffix + ".bak")
                        shutil.copy2(dest_file, backup_file)

                    # Copy new file
                    shutil.copy2(src_file, dest_file)

                    # Clean up allegro.ilinit after copy
                    if file_name.lower() == "allegro.ilinit":
                        try:
                            lines = dest_file.read_text(encoding="utf-8").splitlines()
                            filtered_lines = [line for line in lines if line.strip() in allowed_lines]
                            dest_file.write_text("\n".join(filtered_lines) + "\n", encoding="utf-8")
                        except Exception as e:
                            messagebox.showerror("Menu Update", f"Error updating allegro.ilinit: {e}")
                else:
                    messagebox.showerror("Menu Update", f"Missing file: {src_file}. Contact Automation Team.")

            messagebox.showinfo("Menu Update", "✅ Please restart Allegro.\nIf SKILL was not updated, contact Automation Team.")

            root.destroy()

#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def ExpFunc():
            UsedDt = "Design Menu - Added : EXP Menu"
            PurposeList.append(UsedDt)
            from pathlib import Path
            from datetime import datetime
            import shutil
            import getpass
            import os
            import tkinter as tk
            from tkinter import messagebox

            # --- Initialize Tkinter safely (to avoid 'no default root' errors) ---
            root = tk.Tk()
            root.withdraw()  # Hide the main window

            base_path = Path(r"C:\Cadence")

            insert_text = """POPUP "&EDA_ind"
        BEGIN
            MENUITEM "add_menu",           "add_menu"
        END"""

            insert_text1 = """POPUP "&EXP-1"
        BEGIN
            MENUITEM "exp1_menu",           "exp1_menu"
        END"""

            # --- Update Allegro menu definitions ---
            for folder in base_path.iterdir():
                if folder.is_dir() and "SPB_" in folder.name:
                    link_val = folder / "share" / "pcb" / "text" / "cuimenus"
                    allegro_file = link_val / "allegro.men"

                    if allegro_file.exists():
                        content = allegro_file.read_text(encoding="utf-8")

                        has_EDA = 'POPUP "&EDA_ind"' in content
                        has_ate1 = 'POPUP "&EXP-1"' in content

                        if has_EDA and has_ate1:
                            continue  # Skip already updated files

                        # Backup original
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        backup_file = link_val / f"allegro_Backup_{timestamp}.txt"
                        shutil.copy2(allegro_file, backup_file)

                        # Find insertion point
                        lines = content.splitlines()
                        insert_index = next((i for i in range(len(lines)-1, -1, -1) if lines[i].strip() == "END"), len(lines))

                        updated_lines = lines[:insert_index]

                        if not has_EDA:
                            updated_lines += insert_text.splitlines()
                        if not has_ate1:
                            updated_lines += insert_text1.splitlines()

                        updated_lines.append("END")

                        allegro_file.write_text("\n".join(updated_lines) + "\n", encoding="utf-8")

            # --- Source and destination paths ---
            source_path = Path(r"\\sng-psrvr06\Cadence_Skills\Menu-enc")

            allegro_env = os.environ.get("ALLEGRO_PCBENV")
            home_env = os.environ.get("HOME") or os.environ.get("USERPROFILE")

            # Ensure we have a valid destination path
            if allegro_env:
                dest_path = Path(allegro_env)
            elif home_env:
                dest_path = Path(home_env) / "pcbenv"
            else:
                messagebox.showerror("Menu Update", "Cannot determine user environment path.")
                return

            files_to_copy = ["allegro.ilinit", "env"]

            allowed_lines = [
                'load("//sng-psrvr06/Cadence_Skills/Menu-enc/add0_menu0.ile")',
                'load("//sng-psrvr06/Cadence_Skills/Menu-enc/exp1_menu1.ile")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align.ile" "EncFileKar")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align_do.ile" "EncFileKar")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align_utils.ile" "EncFileKar")',
            ]

            # --- Copy and clean up files ---
            if not source_path.exists():
                messagebox.showerror("Menu Update", "Error: Source path not found. Contact Automation Team.")
                return

            dest_path.mkdir(parents=True, exist_ok=True)

            for file_name in files_to_copy:
                src_file = source_path / file_name
                dest_file = dest_path / file_name

                if src_file.exists():
                    # Backup existing file
                    if dest_file.exists():
                        backup_file = dest_file.with_suffix(dest_file.suffix + ".bak")
                        shutil.copy2(dest_file, backup_file)

                    # Copy new file
                    shutil.copy2(src_file, dest_file)

                    # Clean up allegro.ilinit after copy
                    if file_name.lower() == "allegro.ilinit":
                        try:
                            lines = dest_file.read_text(encoding="utf-8").splitlines()
                            filtered_lines = [line for line in lines if line.strip() in allowed_lines]
                            dest_file.write_text("\n".join(filtered_lines) + "\n", encoding="utf-8")
                        except Exception as e:
                            messagebox.showerror("Menu Update", f"Error updating allegro.ilinit: {e}")
                else:
                    messagebox.showerror("Menu Update", f"Missing file: {src_file}. Contact Automation Team.")

            messagebox.showinfo("Menu Update", "✅ Please restart Allegro.\nIf SKILL was not updated, contact Automation Team.")

            root.destroy()

#------------------------------------------------------------------------------------------------------------------------------------------------------------------  

        def ApdFunc():
            UsedDt = "Design Menu - Added : APD Menu"
            PurposeList.append(UsedDt)
            from pathlib import Path
            from datetime import datetime
            import shutil
            import getpass
            import os
            import tkinter as tk
            from tkinter import messagebox

            # --- Initialize Tkinter safely (to avoid 'no default root' errors) ---
            root = tk.Tk()
            root.withdraw()  # Hide the main window

            base_path = Path(r"C:\Cadence")

            insert_text = """POPUP "&EDA_ind"
        BEGIN
            MENUITEM "add_menu",           "add_menu"
        END"""

            insert_text1 = """POPUP "&APD"
        BEGIN
            MENUITEM "apd_menu",           "apd_menu"
        END"""

            # --- Update Allegro menu definitions ---
            for folder in base_path.iterdir():
                if folder.is_dir() and "SPB_" in folder.name:
                    link_val = folder / "share" / "pcb" / "text" / "cuimenus"
                    allegro_file = link_val / "apd.men"

                    if allegro_file.exists():
                        content = allegro_file.read_text(encoding="utf-8")

                        has_EDA = 'POPUP "&EDA_ind"' in content
                        has_ate1 = 'POPUP "&APD"' in content

                        if has_EDA and has_ate1:
                            continue  # Skip already updated files

                        # Backup original
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        backup_file = link_val / f"allegro_Backup_{timestamp}.txt"
                        shutil.copy2(allegro_file, backup_file)

                        # Find insertion point
                        lines = content.splitlines()
                        insert_index = next((i for i in range(len(lines)-1, -1, -1) if lines[i].strip() == "END"), len(lines))

                        updated_lines = lines[:insert_index]

                        if not has_EDA:
                            updated_lines += insert_text.splitlines()
                        if not has_ate1:
                            updated_lines += insert_text1.splitlines()

                        updated_lines.append("END")

                        allegro_file.write_text("\n".join(updated_lines) + "\n", encoding="utf-8")

            # --- Source and destination paths ---
            source_path = Path(r"\\sng-psrvr06\Cadence_Skills\Menu-enc")

            allegro_env = os.environ.get("ALLEGRO_PCBENV")
            home_env = os.environ.get("HOME") or os.environ.get("USERPROFILE")

            # Ensure we have a valid destination path
            if allegro_env:
                dest_path = Path(allegro_env)
            elif home_env:
                dest_path = Path(home_env) / "pcbenv"
            else:
                messagebox.showerror("Menu Update", "Cannot determine user environment path.")
                return

            files_to_copy = ["allegro.ilinit", "env"]

            allowed_lines = [
                'load("//sng-psrvr06/Cadence_Skills/Menu-enc/add0_menu0.ile")',
                'load("//sng-psrvr06/Cadence_Skills/Menu-enc/apd_menu.il")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align.ile" "EncFileKar")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align_do.ile" "EncFileKar")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align_utils.ile" "EncFileKar")',
            ]

            # --- Copy and clean up files ---
            if not source_path.exists():
                messagebox.showerror("Menu Update", "Error: Source path not found. Contact Automation Team.")
                return

            dest_path.mkdir(parents=True, exist_ok=True)

            for file_name in files_to_copy:
                src_file = source_path / file_name
                dest_file = dest_path / file_name

                if src_file.exists():
                    # Backup existing file
                    if dest_file.exists():
                        backup_file = dest_file.with_suffix(dest_file.suffix + ".bak")
                        shutil.copy2(dest_file, backup_file)

                    # Copy new file
                    shutil.copy2(src_file, dest_file)

                    # Clean up allegro.ilinit after copy
                    if file_name.lower() == "allegro.ilinit":
                        try:
                            lines = dest_file.read_text(encoding="utf-8").splitlines()
                            filtered_lines = [line for line in lines if line.strip() in allowed_lines]
                            dest_file.write_text("\n".join(filtered_lines) + "\n", encoding="utf-8")
                        except Exception as e:
                            messagebox.showerror("Menu Update", f"Error updating allegro.ilinit: {e}")
                else:
                    messagebox.showerror("Menu Update", f"Missing file: {src_file}. Contact Automation Team.")

            messagebox.showinfo("Menu Update", "✅ Please restart Allegro.\nIf SKILL was not updated, contact Automation Team.")

            root.destroy()
        
#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def LibFunc():
            UsedDt = "Design Menu - Added : LIB Menu"
            PurposeList.append(UsedDt)
            from pathlib import Path
            from datetime import datetime
            import shutil
            import getpass
            import os
            import tkinter as tk
            from tkinter import messagebox

            # --- Initialize Tkinter safely (to avoid 'no default root' errors) ---
            root = tk.Tk()
            root.withdraw()  # Hide the main window

            base_path = Path(r"C:\Cadence")

            insert_text1 = """POPUP "&LIBRARY"
        BEGIN
            MENUITEM "lib_menu",           "lib_menu"
        END"""

            # --- Update Allegro menu definitions ---
            for folder in base_path.iterdir():
                if folder.is_dir() and "SPB_" in folder.name:
                    link_val = folder / "share" / "pcb" / "text" / "cuimenus"
                    allegro_file = link_val / "pcb_symbol.men"

                    if allegro_file.exists():
                        content = allegro_file.read_text(encoding="utf-8")
                        has_ate1 = 'POPUP "&LIBRARY"' in content

                        if has_ate1:
                            continue  # Skip already updated files

                        # Backup original
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        backup_file = link_val / f"allegro_Backup_{timestamp}.txt"
                        shutil.copy2(allegro_file, backup_file)

                        # Find insertion point
                        lines = content.splitlines()
                        insert_index = next((i for i in range(len(lines)-1, -1, -1) if lines[i].strip() == "END"), len(lines))

                        updated_lines = lines[:insert_index]

                        if not has_ate1:
                            updated_lines += insert_text1.splitlines()

                        updated_lines.append("END")

                        allegro_file.write_text("\n".join(updated_lines) + "\n", encoding="utf-8")

            # --- Source and destination paths ---
            source_path = Path(r"\\sng-psrvr06\Cadence_Skills\Menu-enc")

            allegro_env = os.environ.get("ALLEGRO_PCBENV")
            home_env = os.environ.get("HOME") or os.environ.get("USERPROFILE")

            # Ensure we have a valid destination path
            if allegro_env:
                dest_path = Path(allegro_env)
            elif home_env:
                dest_path = Path(home_env) / "pcbenv"
            else:
                messagebox.showerror("Menu Update", "Cannot determine user environment path.")
                return

            files_to_copy = ["allegro.ilinit", "env"]

            allowed_lines = [
                'load("//sng-psrvr06/Cadence_Skills/Menu-enc/lib_menu.il")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align.ile" "EncFileKar")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align_do.ile" "EncFileKar")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align_utils.ile" "EncFileKar")',
            ]

            # --- Copy and clean up files ---
            if not source_path.exists():
                messagebox.showerror("Menu Update", "Error: Source path not found. Contact Automation Team.")
                return

            dest_path.mkdir(parents=True, exist_ok=True)

            for file_name in files_to_copy:
                src_file = source_path / file_name
                dest_file = dest_path / file_name

                if src_file.exists():
                    # Backup existing file
                    if dest_file.exists():
                        backup_file = dest_file.with_suffix(dest_file.suffix + ".bak")
                        shutil.copy2(dest_file, backup_file)

                    # Copy new file
                    shutil.copy2(src_file, dest_file)

                    # Clean up allegro.ilinit after copy
                    if file_name.lower() == "allegro.ilinit":
                        try:
                            lines = dest_file.read_text(encoding="utf-8").splitlines()
                            filtered_lines = [line for line in lines if line.strip() in allowed_lines]
                            dest_file.write_text("\n".join(filtered_lines) + "\n", encoding="utf-8")
                        except Exception as e:
                            messagebox.showerror("Menu Update", f"Error updating allegro.ilinit: {e}")
                else:
                    messagebox.showerror("Menu Update", f"Missing file: {src_file}. Contact Automation Team.")

            messagebox.showinfo("Menu Update", "✅ Please restart Allegro.\nIf SKILL was not updated, contact Automation Team.")

            root.destroy()

#------------------------------------------------------------------------------------------------------------------------------------------------------------------     

        def CamMen():
            from tkinter import messagebox
            messagebox.showinfo("CAM Menu", "Please Check CAM Automation..")

#------------------------------------------------------------------------------------------------------------------------------------------------------------------    

        def PanelFunc():
            UsedDt = "Design Menu - Used : Cam Panelling Multi Brd"
            PurposeList.append(UsedDt)

            import subprocess, time, pyautogui, pygetwindow as gw, os, tempfile
            from tkinter import messagebox

            # ---- SCR CONTENT ----
            scr_content = r'''
            view_toolbar@
            view_statbar@
            view_lyrbar@
            view_SuppressGrphcs@ 1
            setgrid@ 2,1.0000,1.0000

            setlayer@ 0
            layer_alloff@ 1

            topBoxX# = 0
            topBoxY# = 0
            bottomBoxX# = 0
            bottomBoxY# = 0
            CreateLayerCount% = 0
            DrillLayerCount% = 1
            PanelCount% = 0
            XSides% = 0
            Airgap% = 100

            OpenForm #1, 70, 10, "Bounding Box Coordinates", character, okcancel

            AddForm #1, 3, 0, "Top Box X Point                         -->", input, topBoxX#, -999999 to 999999
            AddForm #1, 3, 1, "Top Box Y Point                         -->", input, topBoxY#, -999999 to 999999
            AddForm #1, 3, 2, "Bot Box X Point                         -->", input, bottomBoxX#, -999999 to 999999
            AddForm #1, 3, 3, "Bot Box Y Point                         -->", input, bottomBoxY#, -999999 to 999999
            AddForm #1, 3, 4, "Total Num Of ALL Layers In Single Brd   -->", input, CreateLayerCount%, 0 to 999999
            AddForm #1, 3, 5, "Total Num Of Drill Layers In Single Brd -->", input, DrillLayerCount%, 0 to 999999
            AddForm #1, 3, 6, "Total Num Of Boards For Panelling       -->", input, PanelCount%, 0 to 999999
            AddForm #1, 3, 7, "Total Num Of Boards On X Sides          -->", input, XSides%, 0 to 999999
            AddForm #1, 3, 8, "Brd to Brd Airgap Distance              -->", input, Airgap%, 0 to 999999

            DisplayForm #1


            If formcancel(1) Then
                DeleteForm #1
                goto EndScript
            End If

            For Z% = 1 To 100
                view_zoomin@
                axy@ 0.00,0.00
                end@
            Next


            AddValM# = -15
            AddValP# = 15
            If bottomBoxX# < topBoxX# Then
                minX# = bottomBoxX# + AddValM#
                maxX# = topBoxX# + AddValP
            Else
                minX# = topBoxX# + AddValM#
                maxX# = bottomBoxX# + AddValP#
            End If

            If bottomBoxY# < topBoxY# Then
                minY# = bottomBoxY# + AddValM#
                maxY# = topBoxY# + AddValP#
            Else
                minY# = topBoxY# + AddValM#
                maxY# = bottomBoxY# + AddValP#
            End If



            totalLayers% = numlayers!
            TotalLayerList% = totalLayers% 

            BoardWidth# = maxX# - minX#
            BoardHeight# = maxY# - minY#

            Airgap% = Airgap% - 30

            GraphicLayer% = CreateLayerCount% - DrillLayerCount%

            For srcLayer% = 0 To GraphicLayer% - 1
                setlayer@ srcLayer%
                layer_alloff@ 1
                CreateLyrName$ = Layername!
                PassMove# = BoardWidth#
                newLayer% = TotalLayerList% + srcLayer%
                edit_layer@ newLayer%, 0, 32, 4, 0, CreateLyrName$

                subLayer% = srcLayer%
                For i% = 1 To PanelCount% 
                    setlayer@ subLayer%
                    layer_alloff@ 1
                    edit_copy@
                    setsnap@ 1
                    edit_group@
                    axy@ minX#, minY#
                    axy@ maxX#, maxY#
                    back@
                    setsnap@ 0
                    setcoplayer@ newLayer%, 1
                    copy_to_layer@
                    setlayer@ newLayer%
                    layer_alloff@ 1
                    edit_move@
                    setsnap@ 1
                    edit_group@
                    axy@ minX#, minY#
                    axy@ maxX#, maxY#
                    back@
                    axy@ 0.0000,0.0000
                    axy@ PassMove#,0.0000
                    PassMove# = PassMove# + BoardWidth#
                    subLayer% = CreateLayerCount% + subLayer%
                Next
            Next

            XSidesInc% = XSides%


            For srcLayer% = 0 To GraphicLayer% - 1
                setlayer@ srcLayer%
                layer_alloff@ 1
                CreateLyrName = Layername!
                XSides% = XSidesInc%
                AirgapY% = Airgap%
                YLyrInc% = 1
                newLayer% = TotalLayerList% + srcLayer%
                edit_layer@ newLayer%, 0, 32, 4, 0, CreateLyrName
                MoveObject# = BoardWidth#
                YInc# = 0.0000
                XInc# = BoardWidth#
                For j% = 1 To PanelCount%
                    setlayer@ newLayer%
                    layer_alloff@ 1
                    edit_move@
                    setsnap@ 1
                    edit_group@
                    axy@ minX# + MoveObject#, minY#
                    axy@ maxX# + MoveObject#, maxY#
                    back@
                    axy@ 0.0000,0.0000
                    axy@ -XInc#,YInc#
                    XInc# = XInc# - Airgap%
                    MoveObject# = MoveObject# + BoardWidth#
                    If j% = XSides% Then
                        YIncPass# = YLyrInc% * BoardHeight#
                        YInc# = YIncPass# + AirgapY% 
                        YInc# = -YInc#
                        XSides% = XSides% + XSidesInc%
                        AirgapY% = AirgapY% + Airgap%
                        YLyrInc% = YLyrInc% + 1

                        XVal% = j% + 1
                        XInc# = XVal% * BoardWidth#  
                    End If
                Next
            Next


            For DrillLyrlp% = 1 To DrillLayerCount%
                DrillLyrMain% = CreateLayerCount% - DrillLyrlp%
                setlayer@ DrillLyrMain%
                layer_alloff@ 1
                edit_move@
                setsnap@ 1
                edit_group@
                axy@ minX#, minY#
                axy@ maxX#, maxY#
                back@
                axy@ 0.0000,0.0000
                axy@ BoardWidth#,0.0000

                For k% = 2 To PanelCount%
                    DrillLyr% = CreateLayerCount% * k% 
                    DrillLyr% = DrillLyr% - DrillLyrlp%
                    setlayer@ DrillLyr%
                    layer_alloff@ 1
                    edit_move@
                    setsnap@ 1
                    edit_group@
                    axy@ minX#, minY#
                    axy@ maxX#, maxY#
                    back@
                    setsnap@ 0
                    setcoplayer@ DrillLyrMain%, 1
                    copy_to_layer@
                    setlayer@ DrillLyrMain%
                    layer_alloff@ 1
                    edit_move@
                    setsnap@ 1
                    edit_group@
                    axy@ minX#, minY#
                    axy@ maxX#, maxY#
                    back@
                    axy@ 0.0000,0.0000
                    axy@ BoardWidth# * k%,0.0000
                Next
            Next

            For DrillLyrlp1% = 1 To DrillLayerCount%
                DrillLyrMain1% = CreateLayerCount% - DrillLyrlp1%
                XSides% = XSidesInc%
                AirgapY% = Airgap%
                YLyrInc% = 1
                MoveObject# = BoardWidth#
                YInc# = 0.0000
                XInc# = BoardWidth#
                For l% = 1 To PanelCount%
                    setlayer@ DrillLyrMain1%
                    layer_alloff@ 1
                    edit_move@
                    setsnap@ 1
                    edit_group@
                    axy@ minX# + MoveObject#, minY#
                    axy@ maxX# + MoveObject#, maxY#
                    back@
                    axy@ 0.0000,0.0000
                    axy@ -XInc#,YInc#
                    XInc# = XInc# - Airgap%
                    MoveObject# = MoveObject# + BoardWidth#
                    If l% = XSides% Then
                        YIncPass# = YLyrInc% * BoardHeight#
                        YInc# = YIncPass# + AirgapY% 
                        YInc# = -YInc#
                        XSides% = XSides% + XSidesInc%
                        AirgapY% = AirgapY% + Airgap%
                        YLyrInc% = YLyrInc% + 1

                        XVal% = l% + 1
                        XInc# = XVal% * BoardWidth#  
                    End If
                Next
            Next

            DelLyr% = CreateLayerCount% * PanelCount%

            For m% = CreateLayerCount% To DelLyr% - 1
                edit_removelyr@ m%
            Next

            For n% = 0 To GraphicLayer% - 1
                edit_removelyr@ n%
            Next

            RenameLyr% = DelLyr%

            init_layer_reorder@ 
            For P% = 0 To GraphicLayer% - 1
                UpRenameLyr% = RenameLyr% + P%
                change_layer_position@ P%,UpRenameLyr%,P%
            Next
            apply_layer_order@ 

            setlayer@ 0
            layer_alloff@ 1

            view_all@
            print "Multi Board Panelling Has Been Completed"

            DeleteForm #1

            EndScript:
            '''  # >> Keep your full SCR here <<

            # ---- Create temp .scr file ----
            temp_file = os.path.join(tempfile.gettempdir(), "AutoPanelling.scr")
            with open(temp_file, "w") as file:
                file.write(scr_content)

            def bring_cam_to_front():
                windows = gw.getWindowsWithTitle("CAM350")
                if windows:
                    cam_window = windows[0]
                    if cam_window.isMinimized:
                        cam_window.restore()
                        time.sleep(1)
                    cam_window.activate()
                    cam_window.maximize()
                    time.sleep(1)
                else:
                    messagebox.showinfo("CAM Panel", "Error : Please Open CAM350 and Try Again")

            def run_macro_with_scr():
                pyautogui.hotkey('alt', 'm') # open macro menu
                time.sleep(1)
                pyautogui.press('p')         # play macro
                time.sleep(1)

                pyautogui.write(temp_file)   # pass temp file path
                pyautogui.press('enter')
                time.sleep(2)

            bring_cam_to_front()
            run_macro_with_scr()

            # ---- Wait for macro to finish ----
            time.sleep(5)

            # ---- Delete temp .scr file ----
            try:
                os.remove(temp_file)
            except Exception as e:
                print("Warning: temp script not deleted:", e)

#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def PanelAltFunc():
            UsedDt = "Design Menu - Used : Cam Panelling Multi Brd Altium"
            PurposeList.append(UsedDt)

            import subprocess, time, pyautogui, pygetwindow as gw, os, tempfile
            from tkinter import messagebox

            # ---- SCR CONTENT ----
            scr_content = r'''
            setlayer@ 0
            layer_alloff@ 1

            totalLayers% = numlayers!
            TotalLayerList% = totalLayers%
            TotalLayerListUp% = TotalLayerList% - 1

            FirstBrdLyrCount% = 18
            TotalBrds% = 3

            ToCheck% = FirstBrdLyrCount% * TotalBrds%


            If ToCheck% = TotalLayerList% Then 
                init_layer_reorder@ 
                For srcLayerUp% = 0 To FirstBrdLyrCount% - 1
                    setlayer@ srcLayerUp%
                    layer_alloff@ 1
                    change_layer_position@ srcLayerUp%,srcLayerUp%,srcLayerUp%
                Next
                For srcLayer% = 0 To FirstBrdLyrCount% - 1
                    setlayer@ srcLayer%
                    layer_alloff@ 1
                    CreateLyrName$ = Layername!
                    CreateLyrNameUp$ = ucase$(Layername!)
                    DotPos = instr(CreateLyrNameUp$, ".")
                    If DotPos > 0 Then
                        BaseName$ = left$(CreateLyrNameUp$, DotPos - 1)
                        Suffix$   = mid$(CreateLyrNameUp$, DotPos + 1)
                        ResFilt$ = "." + Suffix$

                        UpLayerList% = FirstBrdLyrCount%
                        For PanLayer% = 1 To TotalBrds% - 1
                            UpLayerList% = UpLayerList% + FirstBrdLyrCount%
                            LpLayer% = FirstBrdLyrCount% * PanLayer%
                            IncLayer% = FirstBrdLyrCount% * PanLayer%

                            For LpLayerInc% = LpLayer% To UpLayerList% - 1 
                                setlayer@ LpLayerInc%
                                layer_alloff@ 1
                                CreateLyrName1$ = Layername!
                                CreateLyrNameUp1$ = ucase$(Layername!)
                                DotPos1 = instr(CreateLyrNameUp1$, ".")
                                If DotPos1 > 0 Then
                                    BaseName1$ = left$(CreateLyrNameUp1$, DotPos1 - 1)
                                    Suffix1$   = mid$(CreateLyrNameUp1$, DotPos1 + 1)
                                    ResFilt1$ = "." + Suffix1$
                                    If ResFilt$ = ResFilt1$ Then
                                        UpdateLyr% = IncLayer% + srcLayer%  
                                        change_layer_position@ UpdateLyr%,LpLayerInc%,UpdateLyr%
                                    End If 
                                End If
                            Next

                        Next

                    End If
                Next
                apply_layer_order@ 

                view_toolbar@
                view_statbar@
                view_lyrbar@
                view_SuppressGrphcs@ 1
                setgrid@ 2,1.0000,1.0000

                setlayer@ 0
                layer_alloff@ 1

                topBoxX# = 0.00
                topBoxY# = 0.00
                bottomBoxX# = 0.00
                bottomBoxY# = 0.00
                CreateLayerCount% = 0
                DrillLayerCount% = 0
                PanelCount% = 0
                XSides% = 0
                Airgap# = 100

                OpenForm #1, 70, 10, "Bounding Box Coordinates", character, okcancel

                AddForm #1, 3, 0, "Top Box X Point                         -->", input, topBoxX#, -999999 to 999999
                AddForm #1, 3, 1, "Top Box Y Point                         -->", input, topBoxY#, -999999 to 999999
                AddForm #1, 3, 2, "Bot Box X Point                         -->", input, bottomBoxX#, -999999 to 999999
                AddForm #1, 3, 3, "Bot Box Y Point                         -->", input, bottomBoxY#, -999999 to 999999
                AddForm #1, 3, 4, "Total Num Of ALL Layers In Single Brd   -->", input, CreateLayerCount%, 0 to 999999
                AddForm #1, 3, 5, "Total Num Of Drill Layers In Single Brd -->", input, DrillLayerCount%, 0 to 999999
                AddForm #1, 3, 6, "Total Num Of Boards For Panelling       -->", input, PanelCount%, 0 to 999999
                AddForm #1, 3, 7, "Total Num Of Boards On X Sides          -->", input, XSides%, 0 to 999999
                AddForm #1, 3, 8, "Brd to Brd Airgap Distance              -->", input, Airgap#, 0 to 999999

                DisplayForm #1


                If formcancel(1) Then
                    DeleteForm #1
                    goto EndScript
                End If

                For Z% = 1 To 100
                    view_zoomin@
                    axy@ 0.00,0.00
                    end@
                Next


                AddValM# = -15
                AddValP# = 15
                If bottomBoxX# < topBoxX# Then
                    minX# = bottomBoxX# + AddValM#
                    maxX# = topBoxX# + AddValP
                Else
                    minX# = topBoxX# + AddValM#
                    maxX# = bottomBoxX# + AddValP#
                End If

                If bottomBoxY# < topBoxY# Then
                    minY# = bottomBoxY# + AddValM#
                    maxY# = topBoxY# + AddValP#
                Else
                    minY# = topBoxY# + AddValM#
                    maxY# = bottomBoxY# + AddValP#
                End If



                totalLayers% = numlayers!
                TotalLayerList% = totalLayers% 

                BoardWidth# = maxX# - minX#
                BoardHeight# = maxY# - minY#

                Airgap# = Airgap# - 30

                GraphicLayer% = CreateLayerCount% - DrillLayerCount%

                For srcLayer% = 0 To GraphicLayer% - 1
                    setlayer@ srcLayer%
                    layer_alloff@ 1
                    CreateLyrName$ = Layername!
                    PassMove# = BoardWidth#
                    newLayer% = TotalLayerList% + srcLayer%
                    edit_layer@ newLayer%, 0, 32, 4, 0, CreateLyrName$

                    subLayer% = srcLayer% 
                    For i% = 1 To PanelCount% 
                        setlayer@ subLayer%
                        layer_alloff@ 1
                        edit_copy@
                        setsnap@ 1
                        edit_group@
                        axy@ minX#, minY#
                        axy@ maxX#, maxY#
                        back@
                        setsnap@ 0
                        setcoplayer@ newLayer%, 1
                        copy_to_layer@
                        setlayer@ newLayer%
                        layer_alloff@ 1
                        edit_move@
                        setsnap@ 1
                        edit_group@
                        axy@ minX#, minY#
                        axy@ maxX#, maxY#
                        back@
                        axy@ 0.0000,0.0000
                        axy@ PassMove#,0.0000
                        PassMove# = PassMove# + BoardWidth#
                        subLayer% = CreateLayerCount% + subLayer%
                    Next
                Next

                XSidesInc% = XSides%


                For srcLayer% = 0 To GraphicLayer% - 1
                    setlayer@ srcLayer%
                    layer_alloff@ 1
                    CreateLyrName = Layername!
                    XSides% = XSidesInc%
                    AirgapY# = Airgap#
                    YLyrInc% = 1
                    newLayer% = TotalLayerList% + srcLayer%
                    edit_layer@ newLayer%, 0, 32, 4, 0, CreateLyrName
                    MoveObject# = BoardWidth#
                    YInc# = 0.0000
                    XInc# = BoardWidth#
                    For j% = 1 To PanelCount%
                        setlayer@ newLayer%
                        layer_alloff@ 1
                        edit_move@
                        setsnap@ 1
                        edit_group@
                        axy@ minX# + MoveObject#, minY#
                        axy@ maxX# + MoveObject#, maxY#
                        back@
                        axy@ 0.0000,0.0000
                        axy@ -XInc#,YInc#
                        XInc# = XInc# - Airgap#
                        MoveObject# = MoveObject# + BoardWidth#
                        If j% = XSides% Then
                            YIncPass# = YLyrInc% * BoardHeight#
                            YInc# = YIncPass# + AirgapY# 
                            YInc# = -YInc#
                            XSides% = XSides% + XSidesInc%
                            AirgapY# = AirgapY# + Airgap#
                            YLyrInc% = YLyrInc% + 1

                            XVal% = j% + 1
                            XInc# = XVal% * BoardWidth#  
                        End If
                    Next
                Next


                For DrillLyrlp% = 1 To DrillLayerCount%
                    DrillLyrMain% = CreateLayerCount% - DrillLyrlp%
                    setlayer@ DrillLyrMain%
                    layer_alloff@ 1
                    edit_move@
                    setsnap@ 1
                    edit_group@
                    axy@ minX#, minY#
                    axy@ maxX#, maxY#
                    back@
                    axy@ 0.0000,0.0000
                    axy@ BoardWidth#,0.0000

                    For k% = 2 To PanelCount%
                        DrillLyr% = CreateLayerCount% * k% 
                        DrillLyr% = DrillLyr% - DrillLyrlp%
                        setlayer@ DrillLyr%
                        layer_alloff@ 1
                        edit_move@
                        setsnap@ 1
                        edit_group@
                        axy@ minX#, minY#
                        axy@ maxX#, maxY#
                        back@
                        setsnap@ 0
                        setcoplayer@ DrillLyrMain%, 1
                        copy_to_layer@
                        setlayer@ DrillLyrMain%
                        layer_alloff@ 1
                        edit_move@
                        setsnap@ 1
                        edit_group@
                        axy@ minX#, minY#
                        axy@ maxX#, maxY#
                        back@
                        axy@ 0.0000,0.0000
                        axy@ BoardWidth# * k%,0.0000
                    Next
                Next

                For DrillLyrlp1% = 1 To DrillLayerCount%
                    DrillLyrMain1% = CreateLayerCount% - DrillLyrlp1%
                    XSides% = XSidesInc%
                    AirgapY# = Airgap#
                    YLyrInc% = 1
                    MoveObject# = BoardWidth#
                    YInc# = 0.0000
                    XInc# = BoardWidth#
                    For l% = 1 To PanelCount%
                        setlayer@ DrillLyrMain1%
                        layer_alloff@ 1
                        edit_move@
                        setsnap@ 1
                        edit_group@
                        axy@ minX# + MoveObject#, minY#
                        axy@ maxX# + MoveObject#, maxY#
                        back@
                        axy@ 0.0000,0.0000
                        axy@ -XInc#,YInc#
                        XInc# = XInc# - Airgap#
                        MoveObject# = MoveObject# + BoardWidth#
                        If l% = XSides% Then
                            YIncPass# = YLyrInc% * BoardHeight#
                            YInc# = YIncPass# + AirgapY#
                            YInc# = -YInc#
                            XSides% = XSides% + XSidesInc%
                            AirgapY# = AirgapY# + Airgap#
                            YLyrInc% = YLyrInc% + 1

                            XVal% = l% + 1
                            XInc# = XVal% * BoardWidth#  
                        End If
                    Next
                Next

                DelLyr% = CreateLayerCount% * PanelCount%

                For m% = CreateLayerCount% To DelLyr% - 1
                    edit_removelyr@ m%
                Next

                For n% = 0 To GraphicLayer% - 1
                    edit_removelyr@ n%
                Next

                RenameLyr% = DelLyr%

                init_layer_reorder@ 
                For P% = 0 To GraphicLayer% - 1
                    UpRenameLyr% = RenameLyr% + P%
                    change_layer_position@ P%,UpRenameLyr%,P%
                Next
                apply_layer_order@ 

                setlayer@ 0
                layer_alloff@ 1

                view_all@
                print "Multi Board Panelling Has Been Completed"

                DeleteForm #1

                EndScript:
            Else
                print "Error : Layer Count Missmatch"
            End If
            '''  # >> Keep your full SCR here <<

            # ---- Create temp .scr file ----
            temp_file = os.path.join(tempfile.gettempdir(), "AutoPanelling.scr")
            with open(temp_file, "w") as file:
                file.write(scr_content)

            def bring_cam_to_front():
                windows = gw.getWindowsWithTitle("CAM350")
                if windows:
                    cam_window = windows[0]
                    if cam_window.isMinimized:
                        cam_window.restore()
                        time.sleep(1)
                    cam_window.activate()
                    cam_window.maximize()
                    time.sleep(1)
                else:
                    messagebox.showinfo("CAM Panel", "Error : Please Open CAM350 and Try Again")

            def run_macro_with_scr():
                pyautogui.hotkey('alt', 'm') # open macro menu
                time.sleep(1)
                pyautogui.press('p')         # play macro
                time.sleep(1)

                pyautogui.write(temp_file)   # pass temp file path
                pyautogui.press('enter')
                time.sleep(2)

            bring_cam_to_front()
            run_macro_with_scr()

            # ---- Wait for macro to finish ----
            time.sleep(5)

            # ---- Delete temp .scr file ----
            try:
                os.remove(temp_file)
            except Exception as e:
                print("Warning: temp script not deleted:", e)

#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def PanelCopyFunc():
            UsedDt = "Design Menu - Used : Cam Panelling Copy"
            PurposeList.append(UsedDt)

            import subprocess, time, pyautogui, pygetwindow as gw, os, tempfile
            from tkinter import messagebox

            # ---- SCR CONTENT ----
            scr_content = r'''
            view_toolbar@
            view_statbar@
            view_lyrbar@
            view_SuppressGrphcs@ 1
            setgrid@ 2,0.1,0.1

            setlayer@ 0
            layer_alloff@ 1

            topBoxX# = 0
            topBoxY# = 0
            bottomBoxX# = 0
            bottomBoxY# = 0
            Copies% = 0
            Airgap% = 100
            XSides% = 0
            DrillLyrs% = 1

            OpenForm #1, 70, 10, "Bounding Box Coordinates", character, okcancel

            AddForm #1, 3, 0, "Top Box X Point                     -->", input, topBoxX#, -999999 to 999999
            AddForm #1, 3, 1, "Top Box Y Point                     -->", input, topBoxY#, -999999 to 999999
            AddForm #1, 3, 2, "Bot Box X Point                     -->", input, bottomBoxX#, -999999 to 999999
            AddForm #1, 3, 3, "Bot Box Y Point                     -->", input, bottomBoxY#, -999999 to 999999
            AddForm #1, 3, 4, "Total Num Of Copies                 -->", input, Copies%, 0 to 999999
            AddForm #1, 3, 5, "Total Num Of Drill Layers           -->", input, DrillLyrs%, 0 to 999999
            AddForm #1, 3, 6, "Brd to Brd Airgap Distance          -->", input, Airgap%, 0 to 999999
            AddForm #1, 3, 7, "Total Num Of Boards On X Sides      -->", input, XSides%, 0 to 999999

            DisplayForm #1

            If formcancel(1) Then
                DeleteForm #1
                goto EndScript
            End If

            For Z% = 1 To 100
                view_zoomin@
                axy@ 0.00,0.00
                end@
            Next

            AddValM# = -15
            AddValP# = 15
            If bottomBoxX# < topBoxX# Then
                minX# = bottomBoxX# + AddValM#
                maxX# = topBoxX# + AddValP
            Else
                minX# = topBoxX# + AddValM#
                maxX# = bottomBoxX# + AddValP#
            End If

            If bottomBoxY# < topBoxY# Then
                minY# = bottomBoxY# + AddValM#
                maxY# = topBoxY# + AddValP#
            Else
                minY# = topBoxY# + AddValM#
                maxY# = bottomBoxY# + AddValP#
            End If

            totalLayers% = numlayers!
            TotalLayerList% = totalLayers% 

            BoardWidth# = maxX# - minX#
            BoardHeight# = maxY# - minY#
            Airgap% = Airgap% - 30

            TotalGrpLyr% = TotalLayerList% - DrillLyrs%

            For srcLayer% = 0 To TotalGrpLyr% - 1
                setlayer@ srcLayer%
                layer_alloff@ 1
                CreateLyrName$ = Layername!
                PassCopies# = BoardWidth# 
                PassCopiesY# = BoardHeight# 
                XSidesInc% = XSides% 
                YPass# = 0.0000
                newLayer% = TotalLayerList% + srcLayer%
                edit_layer@ newLayer%, 0, 32, 4, 0, CreateLyrName$
                setlayer@ srcLayer%
                layer_alloff@ 1
                edit_copy@
                setsnap@ 1
                edit_group@
                axy@ minX#, minY#
                axy@ maxX#, maxY#
                back@
                setsnap@ 0
                setcoplayer@ newLayer%, 1
                copy_to_layer@
                setlayer@ newLayer%
                layer_alloff@ 1
                For CopiesLoop% = 1 To Copies%
                    edit_copy@
                    setsnap@ 1
                    edit_group@
                    axy@ minX#, minY#
                    axy@ maxX#, maxY#
                    back@

                    If CopiesLoop% = XSidesInc% Then
                        axy@ 0.0000,0.0000
                        axy@ 0.0000,-PassCopiesY# - Airgap%
                        PassCopies# = PassCopies# + BoardWidth# + Airgap%
                        XSidesInc% = XSidesInc% + XSides%

                        PassCopies# = BoardWidth# 
                        YPass# = -PassCopiesY# - Airgap%
                        PassCopiesY# = PassCopiesY# + BoardHeight# + Airgap%
                    Else 
                        axy@ 0.0000,0.0000
                        axy@ PassCopies# + Airgap%,YPass#
                        PassCopies# = PassCopies# + BoardWidth# + Airgap%
                    End If


                Next
            Next


            For DrillI% = 1 To DrillLyrs%
                PassCopies# = BoardWidth# 
                PassCopiesY# = BoardHeight# 
                XSidesInc% = XSides% 
                YPass# = 0.0000
                DrillLyrMain% = TotalLayerList% - DrillI%
                For CopiesLoop% = 1 To Copies%
                    setlayer@ DrillLyrMain%
                    layer_alloff@ 1
                    edit_copy@
                    setsnap@ 1
                    edit_group@
                    axy@ minX#, minY#
                    axy@ maxX#, maxY#
                    back@

                    If CopiesLoop% = XSidesInc% Then
                        axy@ 0.0000,0.0000
                        axy@ 0.0000,-PassCopiesY# - Airgap%
                        PassCopies# = PassCopies# + BoardWidth# + Airgap%
                        XSidesInc% = XSidesInc% + XSides%

                        PassCopies# = BoardWidth# 
                        YPass# = -PassCopiesY# - Airgap%
                        PassCopiesY# = PassCopiesY# + BoardHeight# + Airgap%
                    Else 
                        axy@ 0.0000,0.0000
                        axy@ PassCopies# + Airgap%,YPass#
                        PassCopies# = PassCopies# + BoardWidth# + Airgap%
                    End If

                Next
            Next



            For l% = 0 To TotalGrpLyr% - 1
                edit_removelyr@ l%
            Next

            init_layer_reorder@ 
            For RenameLayer% = 0 To TotalGrpLyr%
                UpRenameLayer% = TotalLayerList% + RenameLayer%
                change_layer_position@ RenameLayer%,UpRenameLayer%,RenameLayer%
            Next
            apply_layer_order@

            view_all@
            print "Copy Panelling Has Been Completed"

            DeleteForm #1

            EndScript:
            '''  # >> Keep your full SCR here <<

            # ---- Create temp .scr file ----
            temp_file = os.path.join(tempfile.gettempdir(), "AutoPanellingCopy.scr")
            with open(temp_file, "w") as file:
                file.write(scr_content)

            def bring_cam_to_front():
                windows = gw.getWindowsWithTitle("CAM350")
                if windows:
                    cam_window = windows[0]
                    if cam_window.isMinimized:
                        cam_window.restore()
                        time.sleep(1)
                    cam_window.activate()
                    cam_window.maximize()
                    time.sleep(1)
                else:
                    messagebox.showinfo("CAM Panel", "Error : Please Open CAM350 and Try Again")

            def run_macro_with_scr():
                pyautogui.hotkey('alt', 'm') # open macro menu
                time.sleep(1)
                pyautogui.press('p')         # play macro
                time.sleep(1)

                pyautogui.write(temp_file)   # pass temp file path
                pyautogui.press('enter')
                time.sleep(2)

            bring_cam_to_front()
            run_macro_with_scr()

            # ---- Wait for macro to finish ----
            time.sleep(5)

            # ---- Delete temp .scr file ----
            try:
                os.remove(temp_file)
            except Exception as e:
                print("Warning: temp script not deleted:", e)

#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def PanelRTCFunc():
            UsedDt = "Design Menu - Used : Cam Reorder, Type, Color Update"
            PurposeList.append(UsedDt)

            import subprocess, time, pyautogui, pygetwindow as gw, os, tempfile
            from tkinter import messagebox

            # ---- SCR CONTENT ----
            scr_content = r'''
            setlayer@ 0
            layer_alloff@ 1

            totalLayers% = numlayers!
            TotalLayerList% = totalLayers%
            TotalLayerListUp% = TotalLayerList% - 1

            For srcLayer% = 0 To TotalLayerListUp%

                setlayer@ srcLayer%
                layer_alloff@ 1
                CreateLyrName$ = Layername!
                CreateLyrNameUp$ = ucase$(Layername!)

                ' read the tool table for this layer (fallback to 1)
                ToolTable% = Tooltable!
                If ToolTable% <= 0 Then ToolTable% = 1

                SubTxt1$ = ".GT"
                SubTxt2$ = ".GB"
                SubTxt3$ = ".GM"
                SubTxt4$ = ".GD"
                SubTxt5$ = ".GG"
                MainSubTxt$ = ".G"

                pos1% = instr(CreateLyrNameUp$, SubTxt1$)
                pos2% = instr(CreateLyrNameUp$, SubTxt2$)
                pos3% = instr(CreateLyrNameUp$, SubTxt3$)
                pos4% = instr(CreateLyrNameUp$, SubTxt4$)
                pos5% = instr(CreateLyrNameUp$, SubTxt5$)

                MainPos% = instr(CreateLyrNameUp$, MainSubTxt$)

                If pos1% = 0 AND pos2% = 0 AND pos3% = 0 AND pos4% = 0 AND pos5% = 0 Then
                    If MainPos% > 0 Then
                        edit_layer@ srcLayer%,2,0,1,0,CreateLyrName$
                        nc_assign_tool_table_to_layer@ srcLayer%,ToolTable%
                        nc_set_layer_rank@ srcLayer%,0
                        view_LayerFlash@ srcLayer%,1
                        view_LayerDraw@ srcLayer%,1
                        edit_lmapdefaults@
                    End If
                End If

                If pos3% > 0 Then
                    edit_layer@ srcLayer%,6,6,4,0,CreateLyrName$
                    nc_assign_tool_table_to_layer@ srcLayer%,ToolTable%
                    nc_set_layer_rank@ srcLayer%,0
                    view_LayerFlash@ srcLayer%,1
                    view_LayerDraw@ srcLayer%,1
                    edit_lmapdefaults@
                End If

                If pos4% > 0 Then
                    edit_layer@ srcLayer%,6,6,4,0,CreateLyrName$
                    nc_assign_tool_table_to_layer@ srcLayer%,ToolTable%
                    nc_set_layer_rank@ srcLayer%,0
                    view_LayerFlash@ srcLayer%,1
                    view_LayerDraw@ srcLayer%,1
                    edit_lmapdefaults@
                End If

                If pos5% > 0 Then
                    edit_layer@ srcLayer%,6,6,4,0,CreateLyrName$
                    nc_assign_tool_table_to_layer@ srcLayer%,ToolTable%
                    nc_set_layer_rank@ srcLayer%,0
                    view_LayerFlash@ srcLayer%,1
                    view_LayerDraw@ srcLayer%,1
                    edit_lmapdefaults@
                End If

            Next

            For srcLayer% = 0 To TotalLayerListUp%

                setlayer@ srcLayer%
                layer_alloff@ 1
                CreateLyrName$ = Layername!
                CreateLyrNameUp$ = ucase$(Layername!)

                ' read the tool table for this layer (fallback to 1)
                ToolTable% = Tooltable!
                If ToolTable% <= 0 Then ToolTable% = 1

                MainSubTxt$ = ".GTO"
                MainPos% = instr(CreateLyrNameUp$, MainSubTxt$)

                If MainPos% > 0 Then
                    edit_layer@ srcLayer%,13,13,7,0,CreateLyrName$
                    nc_assign_tool_table_to_layer@ srcLayer%,ToolTable%
                    nc_set_layer_rank@ srcLayer%,0
                    view_LayerFlash@ srcLayer%,1
                    view_LayerDraw@ srcLayer%,1
                    edit_lmapdefaults@
                End If

                MainSubTxt1$ = ".GTP"
                MainPos1% = instr(CreateLyrNameUp$, MainSubTxt1$)
                If MainPos1% > 0 Then
                    edit_layer@ srcLayer%,8,8,19,0,CreateLyrName$
                    nc_assign_tool_table_to_layer@ srcLayer%,ToolTable%
                    nc_set_layer_rank@ srcLayer%,0
                    view_LayerFlash@ srcLayer%,1
                    view_LayerDraw@ srcLayer%,1
                    edit_lmapdefaults@
                End If

                MainSubTxt2$ = ".GTS"
                MainPos2% = instr(CreateLyrNameUp$, MainSubTxt2$)
                If MainPos2% > 0 Then
                    edit_layer@ srcLayer%,3,3,11,0,CreateLyrName$
                    nc_assign_tool_table_to_layer@ srcLayer%,ToolTable%
                    nc_set_layer_rank@ srcLayer%,0
                    view_LayerFlash@ srcLayer%,1
                    view_LayerDraw@ srcLayer%,1
                    edit_lmapdefaults@
                End If

                MainSubTxt3$ = ".GTL"
                MainPos3% = instr(CreateLyrNameUp$, MainSubTxt3$)
                If MainPos3% > 0 Then
                    edit_layer@ srcLayer%,2,0,0,0,CreateLyrName$
                    nc_assign_tool_table_to_layer@ srcLayer%,ToolTable%
                    nc_set_layer_rank@ srcLayer%,0
                    view_LayerFlash@ srcLayer%,1
                    view_LayerDraw@ srcLayer%,1
                    edit_lmapdefaults@
                End If

                MainSubTxt4$ = ".GBL"
                MainPos4% = instr(CreateLyrNameUp$, MainSubTxt4$)
                If MainPos4% > 0 Then
                    edit_layer@ srcLayer%,2,0,3,0,CreateLyrName$
                    nc_assign_tool_table_to_layer@ srcLayer%,ToolTable%
                    nc_set_layer_rank@ srcLayer%,0
                    view_LayerFlash@ srcLayer%,1
                    view_LayerDraw@ srcLayer%,1
                    edit_lmapdefaults@
                End If

                MainSubTxt5$ = ".GBS"
                MainPos5% = instr(CreateLyrNameUp$, MainSubTxt5$)
                If MainPos5% > 0 Then
                    edit_layer@ srcLayer%,3,3,12,0,CreateLyrName$
                    nc_assign_tool_table_to_layer@ srcLayer%,ToolTable%
                    nc_set_layer_rank@ srcLayer%,0
                    view_LayerFlash@ srcLayer%,1
                    view_LayerDraw@ srcLayer%,1
                    edit_lmapdefaults@
                End If

                MainSubTxt6$ = ".GBP"
                MainPos6% = instr(CreateLyrNameUp$, MainSubTxt6$)
                If MainPos6% > 0 Then
                    edit_layer@ srcLayer%,8,8,20,0,CreateLyrName$
                    nc_assign_tool_table_to_layer@ srcLayer%,ToolTable%
                    nc_set_layer_rank@ srcLayer%,0
                    view_LayerFlash@ srcLayer%,1
                    view_LayerDraw@ srcLayer%,1
                    edit_lmapdefaults@
                End If

                MainSubTxt7$ = ".GBO"
                MainPos7% = instr(CreateLyrNameUp$, MainSubTxt7$)
                If MainPos7% > 0 Then
                    edit_layer@ srcLayer%,13,13,8,0,CreateLyrName$
                    nc_assign_tool_table_to_layer@ srcLayer%,ToolTable%
                    nc_set_layer_rank@ srcLayer%,0
                    view_LayerFlash@ srcLayer%,1
                    view_LayerDraw@ srcLayer%,1
                    edit_lmapdefaults@
                End If

                MainSubTxt8$ = ".TX"
                MainPos8% = instr(CreateLyrNameUp$, MainSubTxt8$)
                If MainPos8% > 0 Then
                    edit_layer@ srcLayer%,14,14,21,0,CreateLyrName$
                    nc_assign_tool_table_to_layer@ srcLayer%,ToolTable%
                    nc_set_layer_rank@ srcLayer%,0
                    view_LayerFlash@ srcLayer%,1
                    view_LayerDraw@ srcLayer%,1
                    edit_lmapdefaults@
                End If
            Next

            totalLayers% = numlayers!
            TotalLayerList% = totalLayers%
            TotalLayerListUp% = TotalLayerList% - 1
            IncVal% = 0

            init_layer_reorder@

            ' Move .GTO first
            For srcLayer% = 0 To TotalLayerListUp%
                setlayer@ srcLayer%
                layer_alloff@ 1
                CreateLyrName$ = Layername!
                CreateLyrNameUp$ = ucase$(Layername!)
                MainSubTxt$ = ".GTO"
                MainPos% = instr(CreateLyrNameUp$, MainSubTxt$)
                If MainPos% > 0 Then
                    change_layer_position@ IncVal%,srcLayer%,IncVal%
                    IncVal% = IncVal% + 1
                End If
            Next

            ' Move .GTP next
            For srcLayer% = 0 To TotalLayerListUp%
                setlayer@ srcLayer%
                layer_alloff@ 1
                CreateLyrName$ = Layername!
                CreateLyrNameUp$ = ucase$(Layername!)
                MainSubTxt$ = ".GTP"
                MainPos% = instr(CreateLyrNameUp$, MainSubTxt$)
                If MainPos% > 0 Then
                    change_layer_position@ IncVal%,srcLayer%,IncVal%
                    IncVal% = IncVal% + 1
                End If
            Next

            ' Move .GTS next
            For srcLayer% = 0 To TotalLayerListUp%
                setlayer@ srcLayer%
                layer_alloff@ 1
                CreateLyrName$ = Layername!
                CreateLyrNameUp$ = ucase$(Layername!)
                MainSubTxt$ = ".GTS"
                MainPos% = instr(CreateLyrNameUp$, MainSubTxt$)
                If MainPos% > 0 Then
                    change_layer_position@ IncVal%,srcLayer%,IncVal%
                    IncVal% = IncVal% + 1
                End If
            Next

            ' Move .GTL next
            For srcLayer% = 0 To TotalLayerListUp%
                setlayer@ srcLayer%
                layer_alloff@ 1
                CreateLyrName$ = Layername!
                CreateLyrNameUp$ = ucase$(Layername!)
                MainSubTxt$ = ".GTL"
                MainPos% = instr(CreateLyrNameUp$, MainSubTxt$)
                If MainPos% > 0 Then
                    change_layer_position@ IncVal%,srcLayer%,IncVal%
                    IncVal% = IncVal% + 1
                End If
            Next

            ' Generic .G (excluding .GT, .GB, .GM, .GD, .GG)
            For srcLayer% = 0 To TotalLayerListUp%
                setlayer@ srcLayer%
                layer_alloff@ 1
                CreateLyrName$ = Layername!
                CreateLyrNameUp$ = ucase$(Layername!)

                SubTxt1$ = ".GT"
                SubTxt2$ = ".GB"
                SubTxt3$ = ".GM"
                SubTxt4$ = ".GD"
                SubTxt5$ = ".GG"
                MainSubTxt$ = ".G"

                pos1% = instr(CreateLyrNameUp$, SubTxt1$)
                pos2% = instr(CreateLyrNameUp$, SubTxt2$)
                pos3% = instr(CreateLyrNameUp$, SubTxt3$)
                pos4% = instr(CreateLyrNameUp$, SubTxt4$)
                pos5% = instr(CreateLyrNameUp$, SubTxt5$)

                MainPos% = instr(CreateLyrNameUp$, MainSubTxt$)

                If pos1% = 0 AND pos2% = 0 AND pos3% = 0 AND pos4% = 0 AND pos5% = 0 Then
                    If MainPos% > 0 Then
                        change_layer_position@ IncVal%,srcLayer%,IncVal%
                        IncVal% = IncVal% + 1
                    End If
                End If
            Next

            ' Move .GBL, .GBS, .GBP, .GBO in order
            For srcLayer% = 0 To TotalLayerListUp%
                setlayer@ srcLayer%
                layer_alloff@ 1
                CreateLyrName$ = Layername!
                CreateLyrNameUp$ = ucase$(Layername!)
                MainSubTxt$ = ".GBL"
                MainPos% = instr(CreateLyrNameUp$, MainSubTxt$)
                If MainPos% > 0 Then
                    change_layer_position@ IncVal%,srcLayer%,IncVal%
                    IncVal% = IncVal% + 1
                End If
            Next

            For srcLayer% = 0 To TotalLayerListUp%
                setlayer@ srcLayer%
                layer_alloff@ 1
                CreateLyrName$ = Layername!
                CreateLyrNameUp$ = ucase$(Layername!)
                MainSubTxt$ = ".GBS"
                MainPos% = instr(CreateLyrNameUp$, MainSubTxt$)
                If MainPos% > 0 Then
                    change_layer_position@ IncVal%,srcLayer%,IncVal%
                    IncVal% = IncVal% + 1
                End If
            Next

            For srcLayer% = 0 To TotalLayerListUp%
                setlayer@ srcLayer%
                layer_alloff@ 1
                CreateLyrName$ = Layername!
                CreateLyrNameUp$ = ucase$(Layername!)
                MainSubTxt$ = ".GBP"
                MainPos% = instr(CreateLyrNameUp$, MainSubTxt$)
                If MainPos% > 0 Then
                    change_layer_position@ IncVal%,srcLayer%,IncVal%
                    IncVal% = IncVal% + 1
                End If
            Next

            For srcLayer% = 0 To TotalLayerListUp%
                setlayer@ srcLayer%
                layer_alloff@ 1
                CreateLyrName$ = Layername!
                CreateLyrNameUp$ = ucase$(Layername!)
                MainSubTxt$ = ".GBO"
                MainPos% = instr(CreateLyrNameUp$, MainSubTxt$)
                If MainPos% > 0 Then
                    change_layer_position@ IncVal%,srcLayer%,IncVal%
                    IncVal% = IncVal% + 1
                End If
            Next

            ' Move .GM .GD .GG last among .G variants
            For srcLayer% = 0 To TotalLayerListUp%

                setlayer@ srcLayer%
                layer_alloff@ 1
                CreateLyrName$ = Layername!
                CreateLyrNameUp$ = ucase$(Layername!)

                SubTxt3$ = ".GM"
                SubTxt4$ = ".GD"
                SubTxt5$ = ".GG"
                MainSubTxt$ = ".G"

                pos3% = instr(CreateLyrNameUp$, SubTxt3$)
                pos4% = instr(CreateLyrNameUp$, SubTxt4$)
                pos5% = instr(CreateLyrNameUp$, SubTxt5$)

                If pos3% > 0 Then
                    change_layer_position@ IncVal%,srcLayer%,IncVal%
                    IncVal% = IncVal% + 1
                End If

                If pos4% > 0 Then
                    change_layer_position@ IncVal%,srcLayer%,IncVal%
                    IncVal% = IncVal% + 1
                End If

                If pos5% > 0 Then
                    change_layer_position@ IncVal%,srcLayer%,IncVal%
                    IncVal% = IncVal% + 1
                End If

            Next

            apply_layer_order@

            For srcLayer% = 0 To TotalLayerListUp%
                setlayer@ srcLayer%
            Next

            view_all@
            print "Reorder and Color Update Completed"

            '''  # >> Keep your full SCR here <<

            # ---- Create temp .scr file ----
            temp_file = os.path.join(tempfile.gettempdir(), "AutoReorderTyprClrUpdate.scr")
            with open(temp_file, "w") as file:
                file.write(scr_content)

            def bring_cam_to_front():
                windows = gw.getWindowsWithTitle("CAM350")
                if windows:
                    cam_window = windows[0]
                    if cam_window.isMinimized:
                        cam_window.restore()
                        time.sleep(1)
                    cam_window.activate()
                    cam_window.maximize()
                    time.sleep(1)
                else:
                    messagebox.showinfo("CAM Panel", "Error : Please Open CAM350 and Try Again")

            def run_macro_with_scr():
                pyautogui.hotkey('alt', 'm') # open macro menu
                time.sleep(1)
                pyautogui.press('p')         # play macro
                time.sleep(1)

                pyautogui.write(temp_file)   # pass temp file path
                pyautogui.press('enter')
                time.sleep(2)

            bring_cam_to_front()
            run_macro_with_scr()

            # ---- Wait for macro to finish ----
            time.sleep(5)

            # ---- Delete temp .scr file ----
            try:
                os.remove(temp_file)
            except Exception as e:
                print("Warning: temp script not deleted:", e)

#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def IPCSameBrdFunc():
            UsedDt = "Design Menu - Used : IPC Check : Same Board Multicopied"
            PurposeList.append(UsedDt)

            for widget in window.winfo_children():
                if isinstance(widget, tk.Label) and hasattr(widget, 'image'):
                    continue  
                widget.destroy()

            def RunIPCSingle():
                TopLen = entry1.get()
                BotLen = entry2.get()
                XSides = entry3.get()
                YSides = entry4.get()
                AirVal = entry5.get()

                import tkinter as tk
                from tkinter import filedialog
                from tkinter import messagebox
                import random
                import string
                import os
                import re
                        
                if AirVal == "" or TopLen == "" or BotLen == "" or XSides == "" or YSides == "":
                    messagebox.showerror("IPC Check : Single Brd", "-- Required Input Is Empty --")
                else:
                    try:
                        TopNum = int(TopLen)
                        BotNum = int(BotLen)
                        AirNum = int(AirVal)
                        XSidesNum = int(XSides)
                        YSidesNum = int(YSides)
                        TopLenVal = TopNum * 10
                        BotLenVal = BotNum * 10
                        AirValUp = AirNum * 10 

                        DEFAULT_INC_X = TopLenVal + AirValUp  
                        DEFAULT_INC_Y = BotLenVal + AirValUp     
                        XSIDES = XSidesNum             
                        YSIDES = YSidesNum            
                        messagebox.showinfo("IPC Check : Single Brd", "Please Select the IPC File..!")
                        
                        def make_random_string(length, allow_digits=True):
                            chars = string.ascii_uppercase + (string.digits if allow_digits else "")
                            return ''.join(random.choice(chars) for _ in range(length))

                        def unique_random_string(length, existing_set, generated_set, allow_digits=True, tries=20000):
                            for _ in range(tries):
                                s = make_random_string(length, allow_digits)
                                if s not in existing_set and s not in generated_set:
                                    generated_set.add(s)
                                    return s
                            raise RuntimeError("Unable to generate unique random string after many tries.")

                        XY_REGEX = re.compile(r'([ \+\-])(\d{6})Y([ \+\-])(\d{6})')

                        def parse_first_two_tokens(line):
                            i = 0
                            L = len(line)
                            while i < L and line[i] == ' ':
                                i += 1
                            if i >= L:
                                return None
                            t1_start = i
                            while i < L and line[i] != ' ':
                                i += 1
                            t1_end = i
                            while i < L and line[i] == ' ':
                                i += 1
                            if i >= L:
                                return (t1_start, t1_end, line[t1_start:t1_end], None, None, "")
                            t2_start = i
                            while i < L and line[i] != ' ':
                                i += 1
                            t2_end = i
                            t1 = line[t1_start:t1_end]
                            t2 = line[t2_start:t2_end]
                            return (t1_start, t1_end, t1, t2_start, t2_end, t2)

                        def normalize_token2(tok):
                            m = re.match(r'^(.*?)(-\d+)$', tok)
                            if m:
                                return m.group(1), m.group(2)
                            return tok, ""

                        def process_ipc_line(orig_line, existing_names_set, gen_set, token1_map, token2_map, inc_x=DEFAULT_INC_X, inc_y=DEFAULT_INC_Y, xsides=XSIDES, ysides=YSIDES):

                            line = orig_line.rstrip("\n")
                            output_lines = []
                            output_lines.append(line)  

                            if not (("X " in line or "X+" in line or "X-" in line) and ("Y " in line or "Y+" in line or "Y-" in line)):
                                return output_lines

                            m = XY_REGEX.search(line)
                            if not m:
                                return output_lines

                            sign_x = m.group(1)
                            digits_x = m.group(2)
                            sign_y = m.group(3)
                            digits_y = m.group(4)
                            match_start = m.start()
                            match_end = m.end()

                            X_raw_str = (sign_x if sign_x.strip() else "") + digits_x
                            Y_raw_str = (sign_y if sign_y.strip() else "") + digits_y
                            try:
                                X_num = int(X_raw_str)
                                Y_num = int(Y_raw_str)
                            except ValueError:
                                return output_lines

                            token_info = parse_first_two_tokens(line)
                            if token_info is None:
                                return output_lines
                            t1_start, t1_end, t1, t2_start, t2_end, t2 = token_info

                            t1_suffix = t1[3:] if len(t1) > 3 else ""
                            t1_key = t1_suffix.upper() if t1_suffix else t1.upper()

                            is_token1_n_c = (t1_suffix.upper() == "N/C")

                            base2, suffix2 = normalize_token2(t2)
                            t2_key = base2.upper()

                            copies_per_line = xsides * ysides - 1
                            if copies_per_line <= 0:
                                return output_lines

                            if not is_token1_n_c:
                                if t1_key not in token1_map:
                                    if len(t1) > 3:
                                        tail_len = len(t1) - 3
                                        tails = []
                                        for _ in range(copies_per_line):
                                            tail_rand = unique_random_string(tail_len, existing_names_set, gen_set, allow_digits=True)
                                            tails.append(tail_rand)
                                        token1_map[t1_key] = tails
                                    else:
                                        full_len = len(t1)
                                        tails = []
                                        for _ in range(copies_per_line):
                                            full_rand = unique_random_string(full_len, existing_names_set, gen_set, allow_digits=True)
                                            tails.append(full_rand)
                                        token1_map[t1_key] = tails
                            else:
                                token1_map[t1_key] = None

                            if t2_key not in token2_map:
                                base_len = len(base2)
                                repls = []
                                for _ in range(copies_per_line):
                                    repl_base = unique_random_string(base_len, existing_names_set, gen_set, allow_digits=True)
                                    repls.append(repl_base)
                                token2_map[t2_key] = repls

                            copy_index = 0
                            for y_idx in range(0, ysides):
                                offset_y = inc_y * (ysides - 1 - y_idx)
                                Y_new_val = Y_num + offset_y

                                if Y_new_val < 0:
                                    y_sign_out = '-'
                                    y_digits_out = str(abs(Y_new_val)).zfill(6)
                                else:
                                    y_digits_out = str(Y_new_val).zfill(6)
                                    y_sign_out = '+' if sign_y == '+' else ' '

                                for x_idx in range(0, xsides):
                                    offset_x = inc_x * x_idx
                                    X_new_val = X_num + offset_x
                                    if X_new_val < 0:
                                        x_sign_out = '-'
                                        x_digits_out = str(abs(X_new_val)).zfill(6)
                                    else:
                                        x_digits_out = str(X_new_val).zfill(6)
                                        x_sign_out = '+' if sign_x == '+' else ' '

                                    if x_idx == 0 and y_idx == 0:
                                        continue

                                    middle_new = f"{x_sign_out}{x_digits_out}Y{y_sign_out}{y_digits_out}"
                                    updated_line = line[:match_start] + middle_new + line[match_end:]

                                    parsed = parse_first_two_tokens(updated_line)
                                    if not parsed:
                                        output_lines.append(updated_line)
                                        copy_index += 1
                                        continue
                                    nt1_start, nt1_end, nt1_val, nt2_start, nt2_end, nt2_val = parsed

                                    if token1_map[t1_key] is None:
                                        new_first = nt1_val
                                    else:
                                        tails_list = token1_map[t1_key]
                                        tail = tails_list[copy_index]
                                        if len(nt1_val) > 3:
                                            new_first = nt1_val[:3] + tail
                                        else:
                                            new_first = tail

                                    repls2 = token2_map[t2_key]
                                    repl_base = repls2[copy_index]
                                    new_second = repl_base + suffix2

                                    final_line = (
                                        updated_line[:nt1_start]
                                        + new_first
                                        + updated_line[nt1_end:nt2_start]
                                        + new_second
                                        + updated_line[nt2_end:]
                                    )
                                    output_lines.append(final_line)
                                    copy_index += 1

                            return output_lines
                        

                        def extract_and_save_ipc():
                            root = tk.Tk()
                            root.withdraw()

                            input_path = filedialog.askopenfilename(
                                title="Select IPC File",
                                filetypes=[("IPC-356A files", "*.ipc *.net *.txt"), ("All files", "*.*")]
                            )
                            if not input_path:
                                messagebox.showerror("IPC Check : Single Brd", "-- No File Selected --")
                                return

                            existing_names_set = set()
                            with open(input_path, "r", encoding="utf-8", errors="ignore") as fh:
                                for ln in fh:
                                    toks = ln.strip().split()
                                    for t in toks[:6]:
                                        norm = re.sub(r'[^A-Z0-9]', '', t.upper())
                                        if norm:
                                            existing_names_set.add(norm)

                            token1_map = {}   
                            token2_map = {}   
                            generated_set = set()

                            base = os.path.splitext(input_path)[0]
                            output_path = base + "_UPDATED.ipc"

                            out_lines = []
                            with open(input_path, "r", encoding="utf-8", errors="ignore") as f:
                                for raw in f:
                                    line = raw.rstrip("\n")
                                    if not line.strip() or line.startswith("C") or line.startswith("P"):
                                        out_lines.append(line)
                                        continue

                                    processed = process_ipc_line(line, existing_names_set, generated_set, token1_map, token2_map, inc_x=DEFAULT_INC_X, inc_y=DEFAULT_INC_Y, xsides=XSIDES, ysides=YSIDES)
                                    if not processed:
                                        out_lines.append(line)
                                    else:
                                        out_lines.extend(processed)

                            with open(output_path, "w", encoding="utf-8", errors="ignore") as out_f:
                                for ln in out_lines:
                                    out_f.write(ln + "\n")

                            print("=== IPC UPDATED ===")
                            print("Input :", input_path)
                            print("Output:", output_path)
                            messagebox.showinfo("IPC Check : Single Brd", f"Result File Is Generated Successfully:\n{output_path}")
                            return output_path

                        if __name__ == "__main__":
                            extract_and_save_ipc()


                    except ValueError:
                        print("Invalid number entered!")

                

            def create_button(text, relx, rely, command):
                btn = tk.Label(window, text=text, font=("Trebuchet MS", 20), fg="#F6FAFC", bg="#020638", bd=0)
                btn.bind("<Enter>", lambda e: btn.config(fg="#1F5186"))
                btn.bind("<Leave>", lambda e: btn.config(fg="#F6FAFC"))
                btn.bind("<Button-1>", lambda e: command())
                btn.place(relx=relx, rely=rely, anchor="center")
                return btn


            def create_label_entry_lr(text, rely):
                """Left label, right entry field."""
                label = tk.Label(window, text=text, font=("Didot", 14), bg="#f0f0f0")
                label.place(relx=0.40, rely=rely, anchor="center")

                entry = tk.Entry(window, font=("Didot", 13), width=25)
                entry.place(relx=0.65, rely=rely, anchor="center")

                return entry

            form_start_y = 0.25
            spacing = 0.10

            entry1 = create_label_entry_lr("[ TOP X LENGTH ]", form_start_y)
            entry2 = create_label_entry_lr("[ BOT Y LENGTH ]", form_start_y + spacing)
            entry3 = create_label_entry_lr("[ TOTAL NUM OF X SIDE BRDS ]", form_start_y + spacing * 2)
            entry4 = create_label_entry_lr("[ TOTAL NUM OF Y SIDE BRDS ]", form_start_y + spacing * 3)
            entry5 = create_label_entry_lr("[ BRD-BRD AIRGAP ]", form_start_y + spacing * 4)

            create_button("[ CLICK HERE ]", 0.5, form_start_y + spacing * 6, RunIPCSingle)
            create_button("[<-- BACK ]", 0.1, 0.05, DesignFunc)

#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def IPCMultiBrdFunc():
            UsedDt = "Design Menu - Used : IPC Check : Multiple Board"
            PurposeList.append(UsedDt)

            for widget in window.winfo_children():
                if isinstance(widget, tk.Label) and hasattr(widget, 'image'):
                    continue  
                widget.destroy()
            
            
            def RunIPCMulti():
                TopLen = entry1.get()
                BotLen = entry2.get()
                XSides = entry3.get()
                TBrds = entry4.get()
                AirVal = entry5.get()
                
                import tkinter as tk
                from tkinter import filedialog
                import random
                import string
                import re
                import os
                
                if AirVal == "" or TopLen == "" or BotLen == "" or XSides == "" or TBrds == "":
                    messagebox.showerror("IPC Check : Multiple Brd", "-- Required Input Is Empty --")
                else:
                    TopNum = int(TopLen)
                    BotNum = int(BotLen)
                    AirNum = int(AirVal)
                    XSides = int(XSides)
                    num_files = int(TBrds)
                    TopNumVal = TopNum * 10
                    BotNumVal = BotNum * 10
                    AirValUp = AirNum * 10
                    
                    IncVal = TopNumVal + AirValUp 
                    IncValY = BotNumVal + AirValUp
                    first_width = 20
                    mid_width = 10


                    def random_alnum(n, used_set):
                        chars = string.ascii_uppercase + string.digits
                        while True:
                            val = ''.join(random.choice(chars) for _ in range(n))
                            if val not in used_set:
                                used_set.add(val)
                                return val


                    def transform_line(line, first_map, second_map, used_randoms,
                                    first_w=first_width, mid_w=mid_width):

                        if len(line) < first_w + mid_w:
                            line = line.ljust(first_w + mid_w)

                        first_field = line[:first_w]
                        mid_field   = line[first_w:first_w + mid_w]
                        rest        = line[first_w + mid_w:]

                        f_clean = first_field.strip()
                        if len(f_clean) >= 3:
                            prefix = f_clean[:3]
                            suffix = f_clean[3:]

                            if suffix == "N/C":
                                new_suffix = suffix
                            else:
                                if suffix not in first_map:
                                    first_map[suffix] = random_alnum(len(suffix), used_randoms)
                                new_suffix = first_map[suffix]

                            new_first_clean = prefix + new_suffix
                        else:
                            new_first_clean = f_clean

                        new_first = new_first_clean.ljust(first_w)

                        m = re.search(r'\S+', mid_field)
                        if not m:
                            return new_first + mid_field + rest

                        t_start, t_end = m.start(), m.end()
                        token = mid_field[t_start:t_end]

                        if "-" in token:
                            left, right = token.rsplit("-", 1)
                            if right.isdigit():
                                if left not in second_map:
                                    second_map[left] = random_alnum(len(left), used_randoms)
                                new_second = second_map[left] + "-" + right
                            else:
                                if token not in second_map:
                                    second_map[token] = random_alnum(len(token), used_randoms)
                                new_second = second_map[token]
                        else:
                            if token not in second_map:
                                second_map[token] = random_alnum(len(token), used_randoms)
                            new_second = second_map[token]

                        mid_new = mid_field[:t_start] + new_second + mid_field[t_end:]
                        mid_new = mid_new[:mid_w].ljust(mid_w)

                        return new_first + mid_new + rest


                    def update_xy_in_line(line, x_offset=0, y_offset=0):
                        m = re.search(r"X([ +-]\d{6})Y([ +-]\d{6})", line)
                        if not m:
                            return line, False

                        x_raw = m.group(1)
                        y_raw = m.group(2)
                        whole = m.group(0)

                        x_val = int(x_raw)
                        y_val = int(y_raw)

                        x_new = x_val + x_offset
                        y_new = y_val - y_offset   

                        def fmt(v):
                            if v < 0:
                                return "-" + str(abs(v)).zfill(6)
                            return " " + str(v).zfill(6)

                        new_full = "X" + fmt(x_new) + "Y" + fmt(y_new)
                        newline = line.replace(whole, new_full, 1)
                        return newline, True


                    def main():
                        root = tk.Tk()
                        root.withdraw()

                        file_paths = []

                        for i in range(1, num_files + 1):
                            messagebox.showinfo("IPC Check : Multiple Brd", f"Select IPC file - {i}")
                            path = filedialog.askopenfilename(
                                title=f"Select IPC file {i}",
                                filetypes=[("IPC files", "*.ipc"), ("All files", "*.*")]
                            )
                            if path:
                                file_paths.append(path)
                                print("Loaded:", path)
                            else:
                                messagebox.showerror("IPC Check : Multiple Brd", "-- No File Selected --")
                                file_paths.append(None)

                        root.destroy()

                        file_lists = []
                        for p in file_paths:
                            if p:
                                with open(p, "r", encoding="utf-8", errors="replace") as f:
                                    file_lists.append(f.read().splitlines())
                            else:
                                file_lists.append([])

                        MainList = []
                        for data in file_lists:
                            SubList = []
                            for ln in data:
                                s = ln.strip()
                                if len(s) >= 3 and "-" in s and s[:3].isdigit() and \
                                any(t in s for t in ("X ", "Y ", "X-", "X+", "Y-", "Y+")):
                                    SubList.append(ln)
                            MainList.append(SubList)

                        UpdatedMainList = []
                        for SubList in MainList:
                            first_map = {}
                            second_map = {}
                            used_randoms = set()
                            updated = []
                            for ln in SubList:
                                updated.append(transform_line(ln, first_map, second_map, used_randoms))
                            UpdatedMainList.append(updated)

                        FinalUpdatedLists = []
                        for idx, SubList in enumerate(UpdatedMainList):
                            col = idx % XSides
                            row = idx // XSides
                            x_offset = IncVal * col
                            y_offset = IncValY * row

                            processed = []
                            for ln in SubList:
                                new_ln, _ = update_xy_in_line(ln, x_offset, y_offset)
                                processed.append(new_ln)

                            FinalUpdatedLists.append(processed)


                        first_file_data = file_lists[0]

                        comp_start = None
                        comp_end = None

                        for i, ln in enumerate(first_file_data):
                            s = ln.strip()
                            is_comp = (
                                len(s) >= 3 and "-" in s and s[:3].isdigit()
                                and any(t in s for t in ("X ", "Y ", "X-", "X+", "Y-", "Y+"))
                            )

                            if comp_start is None:
                                if is_comp:
                                    comp_start = i
                            else:
                                if not is_comp:
                                    comp_end = i
                                    break

                        if comp_start is not None and comp_end is None:
                            comp_end = len(first_file_data)

                        CombinedList = []
                        for block in FinalUpdatedLists:
                            CombinedList.extend(block)

                        FinalIPC = []
                        FinalIPC.extend(first_file_data[:comp_start])
                        FinalIPC.extend(CombinedList)
                        FinalIPC.extend(first_file_data[comp_end:])

                        save_dir = os.path.dirname(file_paths[0])
                        save_path = os.path.join(save_dir, "MultiBrdPanelIPC.ipc")

                        with open(save_path, "w", encoding="utf-8") as f:
                            for ln in FinalIPC:
                                f.write(ln + "\n")

                        print("\n================================")
                        print("Merged IPC created successfully!")
                        print("Saved as:", save_path)
                        print("================================")
                        messagebox.showinfo("IPC Check : Multiple Brd", f"Result File Is Generated Successfully:\n{save_path}")
                        
                        return save_path


                    if __name__ == "__main__":
                        main()
                    
            
            def create_button(text, relx, rely, command):
                btn = tk.Label(window, text=text, font=("Trebuchet MS", 20), fg="#F6FAFC", bg="#020638", bd=0)
                btn.bind("<Enter>", lambda e: btn.config(fg="#1F5186"))
                btn.bind("<Leave>", lambda e: btn.config(fg="#F6FAFC"))
                btn.bind("<Button-1>", lambda e: command())
                btn.place(relx=relx, rely=rely, anchor="center")
                return btn


            def create_label_entry_lr(text, rely):
                """Left label, right entry field."""
                label = tk.Label(window, text=text, font=("Didot", 14), bg="#f0f0f0")
                label.place(relx=0.40, rely=rely, anchor="center")

                entry = tk.Entry(window, font=("Didot", 13), width=25)
                entry.place(relx=0.65, rely=rely, anchor="center")

                return entry

            form_start_y = 0.25
            spacing = 0.10

            entry1 = create_label_entry_lr("[ TOP X LENGTH ]", form_start_y)
            entry2 = create_label_entry_lr("[ BOT Y LENGTH ]", form_start_y + spacing)
            entry3 = create_label_entry_lr("[ MAXIMUM NUM OF BRDS IN X-AXIS ]", form_start_y + spacing * 2)
            entry4 = create_label_entry_lr("[ TOTAL NUM OF BRDS PANELLED ]", form_start_y + spacing * 3)
            entry5 = create_label_entry_lr("[ BRD-BRD AIRGAP ]", form_start_y + spacing * 4)

            create_button("[ CLICK HERE ]", 0.5, form_start_y + spacing * 6, RunIPCMulti)
            create_button("[<-- BACK ]", 0.1, 0.05, DesignFunc)

#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def GerberCompFunc():
            UsedDt = "Design Menu - Used : Cam Gerber Compare"
            PurposeList.append(UsedDt)

            import subprocess, time, pyautogui, pygetwindow as gw, os, tempfile
            from tkinter import messagebox

            # ---- SCR CONTENT ----
            scr_content = r'''
            setlayer@ 0
            layer_alloff@ 1

            totalLayers% = numlayers!
            TotalLayerList% = totalLayers% 
            TotalLayerListUp% = TotalLayerList% - 1

            SingleLayerList% = TotalLayerList% / 2
            IncLayerList% = TotalLayerList

            init_layer_reorder@ 
            For srcLayer% = 0 To SingleLayerList% - 1
                setlayer@ srcLayer%
                layer_alloff@ 1
                CreateLyrName$ = Layername!
                MainSubTxt8$ = ".TX"
                MainPos8% = instr(CreateLyrName$, MainSubTxt8$)
                If MainPos8% > 0 Then
                    edit_layer@ srcLayer%, 2, 2, 21, 0, CreateLyrName$
                    change_layer_position@ IncLayerList%,srcLayer%,IncLayerList%
                    IncLayerList% = IncLayerList% + 1

                    UpsrcLayer% = srcLayer% + SingleLayerList%
                    setlayer@ UpsrcLayer%
                    layer_alloff@ 1
                    CreateLyrNameUp$ = Layername!
                    edit_layer@ UpsrcLayer%, 0, 0, 21, 0, CreateLyrNameUp$
                    change_layer_position@ IncLayerList%,UpsrcLayer%,IncLayerList%
                    IncLayerList% = IncLayerList% + 1

                Else
                    edit_layer@ srcLayer%, 2, 2, 4, 0, CreateLyrName$
                    change_layer_position@ IncLayerList%,srcLayer%,IncLayerList%
                    IncLayerList% = IncLayerList% + 1

                    UpsrcLayer% = srcLayer% + SingleLayerList%
                    setlayer@ UpsrcLayer%
                    layer_alloff@ 1
                    CreateLyrNameUp$ = Layername!
                    edit_layer@ UpsrcLayer%, 0, 0, 4, 0, CreateLyrNameUp$
                    change_layer_position@ IncLayerList%,UpsrcLayer%,IncLayerList%
                    IncLayerList% = IncLayerList% + 1

                End If

            Next

            apply_layer_order@ 


            init_layer_reorder@ 
            For P% = 0 To TotalLayerList% - 1
                UpRenameLyr% = TotalLayerList% + P%
                change_layer_position@ P%,UpRenameLyr%,P%
            Next
            apply_layer_order@ 


            For srcLayer% = 0 To TotalLayerListUp%
                setlayer@ srcLayer%
            Next
            view_all@
            print "Gerber Compare has been Completed"
            '''  # >> Keep your full SCR here <<

            # ---- Create temp .scr file ----
            temp_file = os.path.join(tempfile.gettempdir(), "AutoGerberCompare.scr")
            with open(temp_file, "w") as file:
                file.write(scr_content)

            def bring_cam_to_front():
                windows = gw.getWindowsWithTitle("CAM350")
                if windows:
                    cam_window = windows[0]
                    if cam_window.isMinimized:
                        cam_window.restore()
                        time.sleep(1)
                    cam_window.activate()
                    cam_window.maximize()
                    time.sleep(1)
                else:
                    messagebox.showinfo("CAM", "Error : Please Open CAM350 and Try Again")

            def run_macro_with_scr():
                pyautogui.hotkey('alt', 'm') # open macro menu
                time.sleep(1)
                pyautogui.press('p')         # play macro
                time.sleep(1)

                pyautogui.write(temp_file)   # pass temp file path
                pyautogui.press('enter')
                time.sleep(2)

            bring_cam_to_front()
            run_macro_with_scr()

            # ---- Wait for macro to finish ----
            time.sleep(5)

            # ---- Delete temp .scr file ----
            try:
                os.remove(temp_file)
            except Exception as e:
                print("Warning: temp script not deleted:", e)

#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def AutoDFMChkFunc():
            UsedDt = "Design Menu - Used : Auto DFM Check"
            PurposeList.append(UsedDt)

            for widget in window.winfo_children():
                if isinstance(widget, tk.Label) and hasattr(widget, 'image'):
                    continue
                widget.destroy()

            def RunDFMChkFunc():
                TrackToTrack = entry1.get()
                TrackToPad = entry2.get()
                PadToPad = entry3.get()
                MinTrackWid = entry4.get()
                MinPad = entry5.get()
                StartLayer = entry6.get()
                EndLayer = entry7.get()

                import os, time, tempfile, subprocess
                import tkinter as tk
                from tkinter import filedialog, messagebox
                import pyautogui, pygetwindow as gw
                from openpyxl import Workbook
                from datetime import datetime

                RULE_MAP = {
                    "Trk-Trk Clr": "TRACK TO TRACK",
                    "Trk-Pad Clr": "TRACK TO PAD",
                    "Pad-Pad Clr": "PAD TO PAD",
                    "Min Trk": "MIN TRACK WIDTH",
                    "Min Pad": "MIN PAD",
                }

                # ✅ Runtime rule inputs
                RULE_INPUTS = {}

                if TrackToTrack:
                    RULE_INPUTS["TRACK TO TRACK"] = TrackToTrack
                if TrackToPad:
                    RULE_INPUTS["TRACK TO PAD"] = TrackToPad
                if PadToPad:
                    RULE_INPUTS["PAD TO PAD"] = PadToPad
                if MinTrackWid:
                    RULE_INPUTS["MIN TRACK WIDTH"] = MinTrackWid
                if MinPad:
                    RULE_INPUTS["MIN PAD"] = MinPad


                def choose_location_and_delete():
                    messagebox.showinfo("DFM Check", "Please Select the CAM File located Folder..")
                    folder_path = filedialog.askdirectory(title="Select a Folder")
                    if not folder_path:
                        return

                    folder_path = folder_path.lower()
                    drc_file = os.path.join(folder_path, "errors.drc")
                    if os.path.isfile(drc_file):
                        os.remove(drc_file)

                    UpdateList = []

                    UpdateList.append("util_drc@")
                    UpdateList.append(f'drc_seterrorfile@ "{folder_path}\\errors.drc"')
                    UpdateList.append(f'drc_deletefile@ "{folder_path}\\errors.drc"')

                    if TrackToTrack:
                        UpdateList += ["drc_Set_Trk2Trk_Test@ 0,1", f"drc_Set_Trk2Trk_Val@ 0,{TrackToTrack}"]
                    if TrackToPad:
                        UpdateList += ["drc_Set_Trk2Pad_Test@ 0,1", f"drc_Set_Trk2Pad_Val@ 0,{TrackToPad}"]
                    if PadToPad:
                        UpdateList += ["drc_Set_Pad2Pad_Test@ 0,1", f"drc_Set_Pad2Pad_Val@ 0,{PadToPad}"]
                    if MinTrackWid:
                        UpdateList += ["drc_Set_MinTrk_Test@ 0,1", f"drc_Set_MinTrk_Val@ 0,{MinTrackWid}"]
                    if MinPad:
                        UpdateList += ["drc_Set_MinPad_Test@ 0,1", f"drc_Set_MinPad_Val@ 0,{MinPad}"]

                    UpdateList += [
                        "totalLayers% = numlayers!",
                        "TotalLayerList% = totalLayers%",
                        "TotalLayerListUp% = TotalLayerList% - 1",
                        "For srcLayer% = 0 To TotalLayerListUp%",
                        "drc_Delete_Drc_Lyr@ 0,srcLayer%,,-1,0",
                        "Next"
                    ]

                    for i in range(int(StartLayer), int(EndLayer) + 1):
                        UpdateList.append(f"drc_Add_Drc_Lyr_To_Pass@ 0,{i-1},-1,0")

                    UpdateList += [
                        "axy@ 56300,56300",
                        "axy@ -56300,-56300",
                        "drc_deleteall@",
                        "drc_NetlistExtract@",
                        "drc_Run_Pass@ 0",
                        "explode_allnets@",
                        "end@",
                        "report_init_sorting@",
                        f'report_drc_save@ 0,0,0,"{folder_path}\\DRCReport.rpt"'
                    ]

                    temp_file = os.path.join(tempfile.gettempdir(), "AutoDFMCheck.scr")
                    with open(temp_file, "w", encoding="utf-8") as f:
                        f.write("\n".join(UpdateList))

                    def bring_cam_to_front():
                        wins = gw.getWindowsWithTitle("CAM350")
                        if not wins:
                            messagebox.showerror("Error", "CAM350 not running")
                            return False
                        w = wins[0]
                        if w.isMinimized:
                            w.restore()
                        w.activate()
                        w.maximize()
                        return True

                    def run_macro():
                        pyautogui.hotkey('alt', 'm')
                        time.sleep(1)
                        pyautogui.press('p')
                        time.sleep(1)
                        pyautogui.write(temp_file)
                        pyautogui.press('enter')

                    def wait_file(path):
                        last = -1
                        stable = 0
                        while True:
                            if os.path.exists(path):
                                size = os.path.getsize(path)
                                if size == last:
                                    stable += 1
                                    if stable >= 3:
                                        return
                                else:
                                    stable = 0
                                    last = size
                            time.sleep(1)

                    def parse_rpt(path):
                        data = {}
                        with open(path, "r", errors="ignore") as f:
                            for line in f:
                                p = line.split()
                                if len(p) < 6:
                                    continue
                                key = f"{p[0]} {p[1]}"
                                if key not in RULE_MAP:
                                    continue
                                try:
                                    dist = float(p[2]) * 1000
                                except:
                                    continue
                                layer = next((x for x in p if x.startswith("L")), None)
                                if not layer:
                                    continue
                                rule = RULE_MAP[key]
                                data.setdefault(layer, {}).setdefault(rule, []).append(dist)
                        return data

                    def gen_excel(rpt, folder):
                        raw = parse_rpt(rpt)

                        layers = sorted(raw.keys())
                        rules = list(RULE_INPUTS.keys())

                        wb = Workbook()
                        ws = wb.active
                        ws.title = "DFM Summary"

                        # ---------- HEADER ROW (RULE + LIMIT) ----------
                        ws.cell(row=1, column=1, value="LAYER")

                        col = 2
                        for rule in rules:
                            limit = RULE_INPUTS.get(rule, "NIL")
                            ws.cell(row=1, column=col, value=f"{rule} ({limit})")
                            col += 1

                        # ---------- DATA ROWS (LAYER WISE) ----------
                        row = 2
                        for layer in layers:
                            ws.cell(row=row, column=1, value=layer)

                            col = 2
                            for rule in rules:
                                value = raw.get(layer, {}).get(rule, [])
                                ws.cell(row=row, column=col, value=min(value) if value else "NIL")
                                col += 1

                            row += 1

                        # ---------- SHEET 2 : OVERALL MIN ----------
                        ws2 = wb.create_sheet("Overall Min")
                        ws2.append(["RULE", "RULE LIMIT", "MIN DISTANCE"])

                        for rule in rules:
                            vals = []
                            for layer in layers:
                                if rule in raw.get(layer, {}):
                                    vals.append(min(raw[layer][rule]))

                            ws2.append([
                                rule,
                                RULE_INPUTS.get(rule, "NIL"),
                                min(vals) if vals else "NIL"
                            ])

                        out = os.path.join(
                            folder,
                            f"AutoDFM_Report_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
                        )
                        wb.save(out)
                        os.startfile(out)
                        return out

                    if bring_cam_to_front():
                        run_macro()
                        wait_file(os.path.join(folder_path, "errors.drc"))
                        wait_file(os.path.join(folder_path, "DRCReport.rpt"))
                        excel = gen_excel(os.path.join(folder_path, "DRCReport.rpt"), folder_path)
                        messagebox.showinfo("DFM Check", f"DFM Excel Created:\n{excel}")

                root = tk.Tk()
                root.withdraw()
                choose_location_and_delete()

            def create_button(text, relx, rely, command):
                btn = tk.Label(window, text=text, font=("Trebuchet MS", 20),
                            fg="#F6FAFC", bg="#020638", bd=0)
                btn.bind("<Enter>", lambda e: btn.config(fg="#1F5186"))
                btn.bind("<Leave>", lambda e: btn.config(fg="#F6FAFC"))
                btn.bind("<Button-1>", lambda e: command())
                btn.place(relx=relx, rely=rely, anchor="center")

            def create_label_entry_lr(text, rely):
                tk.Label(window, text=text, font=("Didot", 14), bg="#f0f0f0").place(relx=0.40, rely=rely, anchor="center")
                e = tk.Entry(window, font=("Didot", 13), width=25)
                e.place(relx=0.65, rely=rely, anchor="center")
                return e

            def create_label_entry_range(text, rely):
                tk.Label(window, text=text, font=("Didot", 14), bg="#f0f0f0").place(relx=0.40, rely=rely, anchor="center")
                e1 = tk.Entry(window, font=("Didot", 8), width=10)
                e1.place(relx=0.60, rely=rely, anchor="center")
                tk.Label(window, text="To", font=("Didot", 13), bg="#f0f0f0").place(relx=0.65, rely=rely, anchor="center")
                e2 = tk.Entry(window, font=("Didot", 8), width=10)
                e2.place(relx=0.70, rely=rely, anchor="center")
                return e1, e2

            form_y = 0.25
            sp = 0.10

            entry1 = create_label_entry_lr("[ TRACK TO TRACK ]", form_y)
            entry2 = create_label_entry_lr("[ TRACK TO PAD ]", form_y + sp)
            entry3 = create_label_entry_lr("[ PAD TO PAD ]", form_y + sp * 2)
            entry4 = create_label_entry_lr("[ MIN TRACK WIDTH ]", form_y + sp * 3)
            entry5 = create_label_entry_lr("[ MIN PAD ]", form_y + sp * 4)
            entry6, entry7 = create_label_entry_range("[ START LAYER TO END LAYER ]", form_y + sp * 5)

            create_button("[ CLICK HERE ]", 0.5, form_y + sp * 6, RunDFMChkFunc)
            create_button("[<-- BACK ]", 0.1, 0.05, DesignFunc)

#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def AutoDFMChkDVFunc():
            UsedDt = "Design Menu - Used : Auto DFM Check (Default Values)"
            PurposeList.append(UsedDt) 
            import os
            import tempfile
            import tkinter as tk
            from tkinter import filedialog
            import subprocess, time, pyautogui, pygetwindow as gw, os, tempfile
            from openpyxl import Workbook
            from datetime import datetime
            
            RULE_MAP = {
                "Trk-Trk Clr": "TRACK TO TRACK",
                "Trk-Pad Clr": "TRACK TO PAD",
                "Pad-Pad Clr": "PAD TO PAD",
                "Min Trk": "MIN TRACK WIDTH",
                "Min Pad": "MIN PAD",
            }



            def choose_location_and_delete():
                messagebox.showinfo("DFM Check", "Please Select the CAM File located Folder..")
                folder_path = filedialog.askdirectory(title="Select a Folder")

                if not folder_path:
                    return  # user cancelled

                folder_path = folder_path.lower()
                print("Selected folder:", folder_path)
                drc_file = os.path.join(folder_path, "errors.drc")

                if os.path.isfile(drc_file):
                    try:
                        os.remove(drc_file)
                        print("Deleted file:", drc_file)
                    except Exception as e:
                        print("Error deleting file:", e)
                else:
                    print("errors.drc file not found")
                
                UpdateList = []
                
                #LineT1 = "setlayer@ 0"
                #LineT2 = "layer_alloff@ 1"
                #UpdateList.append(LineT1)
                #UpdateList.append(LineT2)
                
                #for k in range(int(StartLayer), int(EndLayer) + 1):
                    #l = k - 1
                    #LineT3 = f"setlayer@ {l}%"
                    #UpdateList.append(LineT3)
                
                LineInt1 = "setlayer@ 0"
                UpdateList.append(LineInt1)
                LineInt2 = "layer_alloff@ 1"
                UpdateList.append(LineInt2)

                LineInt3 = "totalLayers% = numlayers!"
                UpdateList.append(LineInt3)
                LineInt4 = "TotalLayerList% = totalLayers%"
                UpdateList.append(LineInt4)
                LineInt5 = "TotalLayerListUp% = TotalLayerList% - 1"
                UpdateList.append(LineInt5)

                LineInt6 = "StartLyr% = 100"
                UpdateList.append(LineInt6)
                LineInt7 = "EndLyr% = 100"
                UpdateList.append(LineInt7)

                LineInt8 = "For srcLayer% = 0 To TotalLayerListUp%"
                UpdateList.append(LineInt8)
                LineInt9 = "setlayer@ srcLayer%"
                UpdateList.append(LineInt9)
                LineInt10 = "layer_alloff@ 1"
                UpdateList.append(LineInt10)
                LineInt11 = "LyrType% = Layertype!"
                UpdateList.append(LineInt11)
                LineInt12 = "If LyrType% = 0 Then"
                UpdateList.append(LineInt12)
                LineInt13 = "StartLyr% = srcLayer%" 
                UpdateList.append(LineInt13)
                LineInt14 = "End If"
                UpdateList.append(LineInt14)

                LineInt15 = "If LyrType% = 3 Then"
                UpdateList.append(LineInt15)
                LineInt16 = "EndLyr% = srcLayer%"
                UpdateList.append(LineInt16)
                LineInt17 = "End If"
                UpdateList.append(LineInt17)
                LineInt18 = "Next"
                UpdateList.append(LineInt18)
                LineInt19 = "If StartLyr% <> 100 AND EndLyr% <> 100 Then"  
                UpdateList.append(LineInt19)

                Line1 = "util_drc@"
                UpdateList.append(Line1)
                
                Line2 = f'drc_seterrorfile@ "{folder_path}/errors.drc"'
                Line2 = Line2.replace("/", "\\")
                UpdateList.append(Line2)
                
                Line2u = f'drc_deletefile@ "{folder_path}/errors.drc"'
                Line2u = Line2u.replace("/", "\\")
                UpdateList.append(Line2u)
            
                Line3 = "drc_Set_Trk2Trk_Test@ 0,1"
                UpdateList.append(Line3)
            
                Line4 = "drc_Set_Trk2Pad_Test@ 0,1"
                UpdateList.append(Line4)
                
                Line5 = "drc_Set_Pad2Pad_Test@ 0,1"
                UpdateList.append(Line5)

                Line6 = "drc_Set_MinTrk_Test@ 0,1"
                UpdateList.append(Line6)
            
                Line7 = "drc_Set_MinPad_Test@ 0,1"
                UpdateList.append(Line7)
            
                Line8 = f"drc_Set_Trk2Trk_Val@ 0,10"
                UpdateList.append(Line8)
            
                Line9 = f"drc_Set_Trk2Pad_Val@ 0,10"
                UpdateList.append(Line9)
                
                Line10 = f"drc_Set_Pad2Pad_Val@ 0,10"
                UpdateList.append(Line10)

                Line11 = f"drc_Set_MinTrk_Val@ 0,10"
                UpdateList.append(Line11)
            
                Line12 = f"drc_Set_MinPad_Val@ 0,10"
                UpdateList.append(Line12)      
                
                LineTot1 = "totalLayers% = numlayers!"
                LineTot2 = "TotalLayerList% = totalLayers%"
                LineTot3 = "TotalLayerListUp% = TotalLayerList% - 1"
                
                LineTot4 = "For srcLayer% = 0 To TotalLayerListUp%"
                LineTot5 = "drc_Delete_Drc_Lyr@ 0,srcLayer%,,-1,0"
                LineTot6 = "Next"
                
                UpdateList.append(LineTot1)
                UpdateList.append(LineTot2)
                UpdateList.append(LineTot3)
                UpdateList.append(LineTot4)
                UpdateList.append(LineTot5)
                UpdateList.append(LineTot6)
                
                ForLine = "For IncLyr% = StartLyr% To EndLyr%"
                UpdateList.append(ForLine)
                LineLyr = "drc_Add_Drc_Lyr_To_Pass@ 0,IncLyr%,-1,0"
                UpdateList.append(LineLyr)
                ForEnd = "Next"
                UpdateList.append(ForEnd)
                
                Line19 = "axy@ 56300.0000,56300.0000"
                UpdateList.append(Line19)
                
                Line20 = "axy@ -56300.0000,-56300.0000"
                UpdateList.append(Line20)
                
                Line21a = "drc_deleteall@ "
                UpdateList.append(Line21a)
                
                Line21 = "drc_NetlistExtract@"
                UpdateList.append(Line21)
                
                Line22 = "drc_Run_Pass@ 0"
                UpdateList.append(Line22)
                
                Line22a = "explode_allnets@"
                UpdateList.append(Line22a)
                
                Line22b = "end@"
                UpdateList.append(Line22b)                   
                
                Line23 = "report_init_sorting@"
                UpdateList.append(Line23)
                
                Line24 = f'report_drc_save@ 0,0,0,"{folder_path}/DRCReport.rpt"'
                Line24 = Line24.replace("/", "\\")
                UpdateList.append(Line24)
                
                LineInt20 = "Else"
                UpdateList.append(LineInt20)
                LineInt21 = '"print Error : Top Layer or Bottom Layer Not Found"'
                UpdateList.append(LineInt21)
                LineInt22 = "End If"
                UpdateList.append(LineInt22)
                
                temp_file = os.path.join(tempfile.gettempdir(), "AutoDFMCheck.scr")

                with open(temp_file, "w", encoding="utf-8") as file:
                    for line in UpdateList:
                        file.write(line + "\n")

                print("Auto DFM script saved at:", temp_file)
                
                def bring_cam_to_front():
                    windows = gw.getWindowsWithTitle("CAM350")
                    if windows:
                        cam_window = windows[0]
                        if cam_window.isMinimized:
                            cam_window.restore()
                            time.sleep(1)
                        cam_window.activate()
                        cam_window.maximize()
                        time.sleep(1)
                    else:
                        messagebox.showinfo("CAM Panel", "Error : Please Open CAM350 and Try Again")

                def run_macro_with_scr():
                    pyautogui.hotkey('alt', 'm') # open macro menu
                    time.sleep(1)
                    pyautogui.press('p')         # play macro
                    time.sleep(1)

                    pyautogui.write(temp_file)   # pass temp file path
                    pyautogui.press('enter')
                    time.sleep(2)
                
                def wait_for_drc_file_complete(drc_path, stable_seconds=3):
                    last_size = -1
                    stable_count = 0

                    while True:
                        if os.path.exists(drc_path):
                            size = os.path.getsize(drc_path)
                            if size == last_size:
                                stable_count += 1
                                if stable_count >= stable_seconds:
                                    return True
                            else:
                                stable_count = 0
                                last_size = size
                        time.sleep(1)
                
                def parse_drc_report_distances(rpt_path):
                    data = {}

                    with open(rpt_path, "r", encoding="utf-8", errors="ignore") as f:
                        for line in f:
                            parts = line.split()
                            if len(parts) < 6:
                                continue

                            rule_key = f"{parts[0]} {parts[1]}".strip()
                            if rule_key not in RULE_MAP:
                                continue

                            try:
                                distance = float(parts[2]) * 1000  # ✅ inches → mils
                            except ValueError:
                                continue

                            layer = next((p for p in parts if ":" in p), None)
                            if not layer:
                                continue

                            rule_name = RULE_MAP[rule_key]
                            data.setdefault(layer, {}).setdefault(rule_name, []).append(distance)

                    return data


                
                def get_min_distances(data):
                    result = {}

                    for layer, rules in data.items():
                        result[layer] = {}
                        for rule, values in rules.items():
                            result[layer][rule] = min(values)

                    return result
                
                import re

                def extract_all_layers_from_rpt(rpt_path):
                    layers = set()
                    layer_pattern = re.compile(r"^L\d+:.+")

                    with open(rpt_path, "r", encoding="utf-8", errors="ignore") as f:
                        for line in f:
                            for part in line.split():
                                if layer_pattern.match(part):
                                    layers.add(part)

                    return sorted(layers)


                
                
                from datetime import datetime

                def generate_excel_from_rpt(rpt_path, output_folder):
                    raw_data = parse_drc_report_distances(rpt_path)
                    min_data = get_min_distances(raw_data)

                    layers = extract_all_layers_from_rpt(rpt_path)
                    rules = [
                        "TRACK TO TRACK",
                        "TRACK TO PAD",
                        "PAD TO PAD",
                        "MIN TRACK WIDTH",
                        "MIN PAD"
                    ]

                    wb = Workbook()
                    ws_overall = wb.active
                    ws_overall.title = "DFM OVERALL"
                    
                    ws_layer = wb.create_sheet(title="DFM LAYER WISE")

                    ws_overall.cell(row=1, column=1, value="RULE")
                    ws_overall.cell(row=1, column=2, value="TOP & BOT MIN (mils)")
                    ws_overall.cell(row=1, column=3, value="OTHER LAYERS MIN (mils)")

                    row_idx = 2

                    for rule in rules:
                        top_bot_vals = []
                        other_vals = []

                        for layer, rule_data in min_data.items():
                            if rule not in rule_data:
                                continue

                            if layer.endswith(".GTL") or layer.endswith(".GBL"):
                                top_bot_vals.append(rule_data[rule])
                            else:
                                other_vals.append(rule_data[rule])

                        ws_overall.cell(row=row_idx, column=1, value=rule)
                        ws_overall.cell(row=row_idx, column=2, value=min(top_bot_vals) if top_bot_vals else "PASS")
                        ws_overall.cell(row=row_idx, column=3, value=min(other_vals) if other_vals else "PASS")

                        row_idx += 1
                    
                    ws_layer.cell(row=1, column=1, value="RULE")

                    for col, layer in enumerate(layers, start=2):
                        ws_layer.cell(row=1, column=col, value=layer)

                    for row, rule in enumerate(rules, start=2):
                        ws_layer.cell(row=row, column=1, value=rule)

                        for col, layer in enumerate(layers, start=2):
                            value = min_data.get(layer, {}).get(rule, "PASS")
                            ws_layer.cell(row=row, column=col, value=value)




                    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                    excel_path = os.path.join(
                        output_folder,
                        f"AutoDFM_Report_{timestamp}.xlsx"
                    )

                    wb.save(excel_path)
                    return excel_path


                bring_cam_to_front()
                run_macro_with_scr()
                
                # -------- WAIT FOR DRC COMPLETION --------
                drc_path = os.path.join(folder_path, "errors.drc")
                wait_for_drc_file_complete(drc_path)
                
                rpt_path = os.path.join(folder_path, "DRCReport.rpt")
                wait_for_drc_file_complete(rpt_path)

                excel_file = generate_excel_from_rpt(rpt_path, folder_path)
                print("DFM Excel created:", excel_file)
                os.startfile(excel_file)
                messagebox.showinfo("DFM Check", f"DFM Excel created: {excel_file}")
                
            root = tk.Tk()
            root.withdraw()

            choose_location_and_delete()
    
#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def AutoLayerOrderFunc():
            UsedDt = "Design Menu - Used : Drill to Copper Layer Order Update"
            PurposeList.append(UsedDt)

            for widget in window.winfo_children():
                if isinstance(widget, tk.Label) and hasattr(widget, 'image'):
                    continue
                widget.destroy()

            def RunDrillToCopperFunc():
                import subprocess, time, pyautogui, pygetwindow as gw, os, tempfile
                from tkinter import messagebox  
                
                DrillLayer = entry1.get()
                StartLayer = entry6.get()
                EndLayer = entry7.get()
                
                UpdateList = []

                UpdateList.append("setlayer@ 0")
                UpdateList.append("layer_alloff@ 1")
                UpdateList.append(f"Startlayer% = {StartLayer}")
                UpdateList.append(f"Endlayer% = {EndLayer}")
                UpdateList.append(f"DrillLayer% = {DrillLayer}")
                UpdateList.append("For srcLayer% = Startlayer% - 1 To Endlayer% - 1")
                UpdateList.append("setlayer@ srcLayer%")
                UpdateList.append("layer_alloff@ 1")
                UpdateList.append("edit_delete@")
                UpdateList.append("setsnap@ 1")
                UpdateList.append("seteditdraw@ 0")
                UpdateList.append("seteditvia@ 0")
                UpdateList.append("seteditflash@ 1")
                UpdateList.append("seteditdrill@ 0")
                UpdateList.append("seteditmill@ 0")
                UpdateList.append("edit_multiple@")
                UpdateList.append("edit_selectall2@")
                UpdateList.append("delete_edits@ 4,1")
                UpdateList.append("setlayer@ DrillLayer% - 1")
                UpdateList.append("edit_copy@")
                UpdateList.append("setsnap@ 1")
                UpdateList.append("edit_selectall2@")
                UpdateList.append("setcoplayer@ srcLayer%,1")
                UpdateList.append("copy_to_layer@")
                UpdateList.append("Next")
                UpdateList.append("totalLayers% = numlayers!")
                UpdateList.append("TotalLayerList% = totalLayers%")
                UpdateList.append("TotalLayerListUp% = TotalLayerList% - 1")
                UpdateList.append("For srcLayer% = 0 To TotalLayerListUp%")
                UpdateList.append("setlayer@ srcLayer%")
                UpdateList.append("Next")
                UpdateList.append("view_all@")
                UpdateList.append('print "Drill Layer Update Completed"')
                
                # ---- Create temp .scr file ----
                temp_file = os.path.join(tempfile.gettempdir(), "AutoDrillLayerUpdate.scr")
                with open(temp_file, "w") as file:
                    for line in UpdateList:
                        file.write(line + "\n")

                def bring_cam_to_front():
                    windows = gw.getWindowsWithTitle("CAM350")
                    if windows:
                        cam_window = windows[0]
                        if cam_window.isMinimized:
                            cam_window.restore()
                            time.sleep(1)
                        cam_window.activate()
                        cam_window.maximize()
                        time.sleep(1)
                    else:
                        messagebox.showinfo("CAM", "Error : Please Open CAM350 and Try Again")

                def run_macro_with_scr():
                    pyautogui.hotkey('alt', 'm') # open macro menu
                    time.sleep(1)
                    pyautogui.press('p')         # play macro
                    time.sleep(1)

                    pyautogui.write(temp_file)   # pass temp file path
                    pyautogui.press('enter')
                    time.sleep(2)

                bring_cam_to_front()
                run_macro_with_scr()

                # ---- Wait for macro to finish ----
                time.sleep(5)

                # ---- Delete temp .scr file ----
                try:
                    os.remove(temp_file)
                except Exception as e:
                    print("Warning: temp script not deleted:", e)
                
            def create_button(text, relx, rely, command):
                btn = tk.Label(window, text=text, font=("Trebuchet MS", 20),
                            fg="#F6FAFC", bg="#020638", bd=0)
                btn.bind("<Enter>", lambda e: btn.config(fg="#1F5186"))
                btn.bind("<Leave>", lambda e: btn.config(fg="#F6FAFC"))
                btn.bind("<Button-1>", lambda e: command())
                btn.place(relx=relx, rely=rely, anchor="center")

            def create_label_entry_lr(text, rely):
                tk.Label(window, text=text, font=("Didot", 14), bg="#f0f0f0").place(relx=0.40, rely=rely, anchor="center")
                e = tk.Entry(window, font=("Didot", 13), width=25)
                e.place(relx=0.75, rely=rely, anchor="center")
                return e

            def create_label_entry_range(text, rely):
                tk.Label(window, text=text, font=("Didot", 14), bg="#f0f0f0").place(relx=0.40, rely=rely, anchor="center")
                e1 = tk.Entry(window, font=("Didot", 8), width=10)
                e1.place(relx=0.70, rely=rely, anchor="center")
                tk.Label(window, text="To", font=("Didot", 13), bg="#f0f0f0").place(relx=0.75, rely=rely, anchor="center")
                e2 = tk.Entry(window, font=("Didot", 8), width=10)
                e2.place(relx=0.80, rely=rely, anchor="center")
                return e1, e2

            form_y = 0.25
            sp = 0.10

            entry1 = create_label_entry_lr("[ DRILL LAYER NUMBER ]", form_y)
            entry6, entry7 = create_label_entry_range("[ DRILL TO COPY START LAYER TO END LAYER ]", form_y + sp)

            create_button("[ CLICK HERE ]", 0.5, form_y + sp * 3, RunDrillToCopperFunc)
            create_button("[<-- BACK ]", 0.1, 0.05, DesignFunc)

#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def AutoLayerRTCByInpFunc():
            UsedDt = "Design Menu - Used : Layer Reorder Type and Color By Input File"
            PurposeList.append(UsedDt)

            import subprocess, time, pyautogui, pygetwindow as gw, os, tempfile
            from tkinter import messagebox

            # ---- SCR CONTENT ----
            scr_content = r'''
           
            True%   = 1
            False%  = 0

            TxtFile$        = ""
            TxtLine$        = ""
            LineCount%      = 0
            Counter%        = 0


            gosub GETTXTFILE
            gosub COUNTLINES
            gosub READTXT
            gosub SHOWINFO

            MACROEND:
            print_msg "Macro finished."
            delay 500
            end


            GETTXTFILE:

            OpenForm #1, 80, 6, "Layer Reorder", character, okcancel
            AddForm #1, 0, 1, "Select the Input .TXT file for layer reordering", print
            AddForm #1, 0, 3, "Layer File:", GETOPENFILENAME, TxtFile$
            DisplayForm #1

            Canceled% = formCancel(1)
            DeleteForm #1

            if Canceled% then goto MACROEND

            if TxtFile$ = "" then
                print "Please select a text file before continuing."
                goto GETTXTFILE
            end if

            return


            COUNTLINES:

            Counter% = 0
            open TxtFile$ for input as #1

            do
                input #1, TxtLine$
                Counter% = Counter% + 1
            loop until eof(1)

            close #1
            return


            READTXT:

            print_msg "Reading text file..."

            open TxtFile$ for input as #1

            IncVal% = 0

            init_layer_reorder@

            for LineCount% = 1 to Counter%
                line input #1, TxtLine$
                'print TxtLine$

                setlayer@ 0
                layer_alloff@ 1

                totalLayers% = numlayers!
                TotalLayerList% = totalLayers%
                TotalLayerListUp% = TotalLayerList% - 1

                For srcLayer% = 0 To TotalLayerListUp%
                    setlayer@ srcLayer%
                    layer_alloff@ 1
                    CreateLyrName$ = Layername!
                    ToolTable% = Tooltable!
                    If ucase$(CreateLyrName$) = ucase$(TxtLine$) Then

                        If LineCount% = 1 Then
                            edit_layer@ srcLayer%,2,0,0,0,CreateLyrName$
                            nc_assign_tool_table_to_layer@ srcLayer%,ToolTable%
                            nc_set_layer_rank@ srcLayer%,0
                            view_LayerFlash@ srcLayer%,1
                            view_LayerDraw@ srcLayer%,1
                            edit_lmapdefaults@
                        End If

                        If LineCount% = Counter% Then
                            edit_layer@ srcLayer%,2,0,3,0,CreateLyrName$
                            nc_assign_tool_table_to_layer@ srcLayer%,ToolTable%
                            nc_set_layer_rank@ srcLayer%,0
                            view_LayerFlash@ srcLayer%,1
                            view_LayerDraw@ srcLayer%,1
                            edit_lmapdefaults@
                        End If

                        If LineCount% <> Counter% And LineCount% <> 1 Then
                            edit_layer@ srcLayer%,2,0,1,0,CreateLyrName$
                            nc_assign_tool_table_to_layer@ srcLayer%,ToolTable%
                            nc_set_layer_rank@ srcLayer%,0
                            view_LayerFlash@ srcLayer%,1
                            view_LayerDraw@ srcLayer%,1
                            edit_lmapdefaults@
                        End If

                        change_layer_position@ IncVal%,srcLayer%,IncVal%
                        IncVal% = IncVal% + 1

                    End If
                Next

            next

            apply_layer_order@

            close #1
            close_msg
            return


            SHOWINFO:

            OpenForm #2, 80, 5, "TXT Read Complete", character, ok
            AddForm #2, 0, 1, "File read successfully:", print
            AddForm #2, 0, 2, TxtFile$, print
            AddForm #2, 0, 3, "Total Layers Ordered: " + str$(IncVal%), print
            DisplayForm #2

            Dummy% = formOk(2)
            DeleteForm #2
            return

            '''  # >> Keep your full SCR here <<

            # ---- Create temp .scr file ----
            temp_file = os.path.join(tempfile.gettempdir(), "AutoRTCByInp.scr")
            with open(temp_file, "w") as file:
                file.write(scr_content)

            def bring_cam_to_front():
                windows = gw.getWindowsWithTitle("CAM350")
                if windows:
                    cam_window = windows[0]
                    if cam_window.isMinimized:
                        cam_window.restore()
                        time.sleep(1)
                    cam_window.activate()
                    cam_window.maximize()
                    time.sleep(1)
                else:
                    messagebox.showinfo("CAM", "Error : Please Open CAM350 and Try Again")

            def run_macro_with_scr():
                pyautogui.hotkey('alt', 'm') # open macro menu
                time.sleep(1)
                pyautogui.press('p')         # play macro
                time.sleep(1)

                pyautogui.write(temp_file)   # pass temp file path
                pyautogui.press('enter')
                time.sleep(2)

            bring_cam_to_front()
            run_macro_with_scr()

            # ---- Wait for macro to finish ----
            time.sleep(5)

            # ---- Delete temp .scr file ----
            try:
                os.remove(temp_file)
            except Exception as e:
                print("Warning: temp script not deleted:", e)
            
#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def BrdPadStackDefinitionReport():
            UsedDt = "QC : Padstack Definition Report"
            PurposeList.append(UsedDt)

            import os
            import re
            import subprocess
            import time
            from tkinter.filedialog import askopenfilename
            from tkinter import messagebox

            # -------------------------------------------------------------------------
            # Select BRD file
            # -------------------------------------------------------------------------
            messagebox.showinfo("QC_Padstack Definition Report",
                                "Please Select The Current Design BRD File")

            SelectedFile = askopenfilename(
                title="Select BRD File",
                filetypes=[("Board Files", "*.brd")]
            )
            if not SelectedFile:
                messagebox.showerror("QC_Padstack Definition Report",
                                    "File selection cancelled.")
                return

            output_folder = os.path.dirname(SelectedFile)
            brd_name      = os.path.splitext(os.path.basename(SelectedFile))[0]

            # -------------------------------------------------------------------------
            # Find allegro.exe
            # -------------------------------------------------------------------------
            cadence_path = r"C:\Cadence"
            if not os.path.exists(cadence_path):
                messagebox.showerror("QC_Padstack Definition Report",
                                    f"Cadence path not found:\n{cadence_path}")
                return

            version_map  = {}
            tool_folders = []
            for folder in os.listdir(cadence_path):
                if os.path.isdir(os.path.join(cadence_path, folder)):
                    m = re.match(r"SPB_(\d+)\.(\d+)", folder)
                    if m:
                        version_map[folder] = tuple(map(int, m.groups()))
                        tool_folders.append(folder)

            found_folders = []
            allegro_dir   = ""
            for folder in tool_folders:
                bin_path = os.path.join(cadence_path, folder, "tools", "bin")
                if os.path.exists(os.path.join(bin_path, "allegro.exe")):
                    found_folders.append(folder)

            if not found_folders:
                messagebox.showerror("QC_Padstack Definition Report",
                                    "allegro.exe not found.")
                return

            highest_folder = max(found_folders, key=lambda f: version_map[f])
            allegro_exe    = os.path.join(
                cadence_path, highest_folder, "tools", "bin", "allegro.exe"
            )
            allegro_bin    = os.path.join(cadence_path, highest_folder, "tools", "bin")

            # -------------------------------------------------------------------------
            # Output report path  (forward slashes for SKILL strings)
            # -------------------------------------------------------------------------
            report_file     = os.path.join(output_folder,
                                        f"{brd_name}_PadStack_Definition_Report.txt")
            report_file_sk  = report_file.replace("\\", "/")
            brd_file_sk     = SelectedFile.replace("\\", "/")

            # -------------------------------------------------------------------------
            # Write SKILL script
            # Allegro's built-in padstack report is triggered by:
            #   axlPadstackReport() — SPB 16.x/17.x
            # which is the same function called by Reports > Padstack > Definition.
            # We open the design, run the report, save, exit.
            # -------------------------------------------------------------------------
            temp_ts     = int(time.time())
            skill_path  = os.path.join(output_folder, f"ps_report_{temp_ts}.il")
            skill_sk    = skill_path.replace("\\", "/")

            skill_script = f'''; Padstack Definition Report generator
        ; Auto-generated — do not edit

        let((result)
        ; Open the board
        axlOpenDesign("{brd_file_sk}")

        ; Generate the padstack definition report
        ; axlPadstackReport writes the exact .txt that Allegro produces
        ; via Reports > Padstack > Definition Report
        result = axlPadstackReport(
            "{report_file_sk}"
            "DEFINITION"
            "MILS"
            2
        )

        ; If axlPadstackReport not available, fall back to shell command
        when(result == nil
            axlShellCommand(
            strcat("allegro -nograph -padstack_report \\"" "{report_file_sk}" "\\"")
            )
        )

        axlSaveDesign()
        exit()
        )
        '''

            with open(skill_path, "w") as f:
                f.write(skill_script)

            # -------------------------------------------------------------------------
            # Run Allegro in batch / no-graph mode
            # Multiple invocation styles tried in order (SPB version differences)
            # -------------------------------------------------------------------------
            log_path = os.path.join(output_folder, f"ps_report_{temp_ts}.log")

            # Build environment so allegro can find its libraries
            env = os.environ.copy()
            env["PATH"] = allegro_bin + os.pathsep + env.get("PATH", "")

            # Try invocation styles from most to least common for SPB 17.2 batch
            invocations = [
                # Style 1: standard batch with -s skill script
                f'"{allegro_exe}" -nograph -batch -log "{log_path}" -s "{skill_path}" "{SelectedFile}"',
                # Style 2: no -batch flag (some SPB versions)
                f'"{allegro_exe}" -nograph -log "{log_path}" -s "{skill_path}" "{SelectedFile}"',
                # Style 3: script before design file
                f'"{allegro_exe}" -nograph -batch -log "{log_path}" "{SelectedFile}" -s "{skill_path}"',
            ]

            run_ok   = False
            last_err = ""

            for cmd in invocations:
                try:
                    proc = subprocess.run(
                        cmd, shell=True,
                        stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                        text=True, timeout=180, env=env,
                        cwd=output_folder
                    )
                    # Check if report was actually created
                    if os.path.exists(report_file) and os.path.getsize(report_file) > 0:
                        run_ok = True
                        break
                    last_err = (proc.stderr or proc.stdout or "").strip()[:500]
                except subprocess.TimeoutExpired:
                    last_err = "Allegro timed out after 180 seconds."
                    break
                except Exception as e:
                    last_err = str(e)

            # -------------------------------------------------------------------------
            # Clean up temp files
            # -------------------------------------------------------------------------
            for f in [skill_path, log_path]:
                try:
                    if os.path.exists(f):
                        os.remove(f)
                except:
                    pass

            # -------------------------------------------------------------------------
            # Result
            # -------------------------------------------------------------------------
            if run_ok:
                # Count padstacks in generated report
                ps_count = 0
                try:
                    with open(report_file) as f:
                        for line in f:
                            if line.startswith("Padstack:"):
                                ps_count += 1
                except:
                    pass

                messagebox.showinfo(
                    "QC_Padstack Definition Report",
                    f"Report generated successfully!\n\n"
                    f"File     : {os.path.basename(report_file)}\n"
                    f"Padstacks: {ps_count}\n\n"
                    f"Location:\n{output_folder}"
                )
            else:
                # ----------------------------------------------------------------
                # Fallback: allegro.exe -padstack_report command-line flag
                # Some SPB versions support this direct flag
                # ----------------------------------------------------------------
                fallback_cmd = (
                    f'"{allegro_exe}" -nograph '
                    f'-padstack_report "{report_file}" '
                    f'-log "{log_path}" '
                    f'"{SelectedFile}"'
                )
                try:
                    subprocess.run(
                        fallback_cmd, shell=True,
                        stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                        text=True, timeout=120, env=env,
                        cwd=output_folder
                    )
                except:
                    pass

                if os.path.exists(report_file) and os.path.getsize(report_file) > 0:
                    messagebox.showinfo(
                        "QC_Padstack Definition Report",
                        f"Report generated (via fallback)!\n\n"
                        f"File     : {os.path.basename(report_file)}\n\n"
                        f"Location:\n{output_folder}"
                    )
                else:
                    # ----------------------------------------------------------------
                    # Save debug log for diagnosis
                    # ----------------------------------------------------------------
                    debug_out = os.path.join(output_folder,
                                            f"{brd_name}_allegro_debug.txt")
                    try:
                        with open(debug_out, "w") as f:
                            f.write(f"Allegro path : {allegro_exe}\n")
                            f.write(f"BRD file     : {SelectedFile}\n")
                            f.write(f"Last error   : {last_err}\n\n")
                            f.write("Invocations tried:\n")
                            for i, cmd in enumerate(invocations, 1):
                                f.write(f"  [{i}] {cmd}\n")
                            f.write(f"\nSKILL script:\n{skill_script}\n")
                    except:
                        pass

                    messagebox.showerror(
                        "QC_Padstack Definition Report",
                        f"Could not generate report.\n\n"
                        f"Root cause: This BRD is in binary format. "
                        f"extracta cannot provide per-layer padstack data, "
                        f"and Allegro batch mode did not produce the report.\n\n"
                        f"Debug info saved to:\n{debug_out}\n\n"
                        f"Please share that file."
                    )

#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def create_button(text, y_pixel_pos, command):
            btn = tk.Label(window, text=text, font=("Trebuchet MS", 15), fg="#F6FAFC",
                    bg="#020638", bd=0)
            btn.bind("<Enter>", lambda e: btn.config(fg="#1F5186"))
            btn.bind("<Leave>", lambda e: btn.config(fg="#F6FAFC"))
            btn.bind("<Button-1>", lambda e: command())
            btn.place(x=20, y=y_pixel_pos)  

        create_button("[<-- BACK ]", 20, show_main_menu)
        create_button("[ -- Excel Compare -- ]", 100, ExcelComp)
        create_button("[ -- SCH Mapping ATE-1 -- ]", 140, CHMapATE1)
        create_button("[ -- PIN Report MLO -- ]", 180, CompPinRepFilt)
        create_button("[ -- PAD STACK REPORT & CHECK -- ]", 220, BrdPadStackDefinitionReport)
        
        
        def create_button1(text, y_pixel_pos1, command):
            btn1 = tk.Label(window, text=text, font=("Trebuchet MS", 15), fg="#F6FAFC",
                    bg="#020638", bd=0)
            btn1.bind("<Enter>", lambda e: btn1.config(fg="#1F5186"))
            btn1.bind("<Leave>", lambda e: btn1.config(fg="#F6FAFC"))
            btn1.bind("<Button-1>", lambda e: command())
            btn1.place(x=500, y=y_pixel_pos1)  

        create_button1("[|-- DESIGN SKILL MENU --|]", 20, MenuPass)
        create_button1("[ -- EXCEL -- ]", 100, ExcelFunc)
        create_button1("[ -- ALIGN -- ]", 140, AlignFunc)
        create_button1("[ -- ATE - 1 -- ]", 180, ATE1Func)
        create_button1("[ -- ATE - 2 -- ]", 220, ATE2Func)
        create_button1("[ -- ATE - 3 -- ]", 260, ATE3Func)
        create_button1("[ -- ATE - 4 -- ]", 300, ATE4Func)
        create_button1("[ -- EXP - 1 -- ]", 340, ExpFunc)
        create_button1("[ -- APD -- ]", 380, ApdFunc)
        create_button1("[ -- LIBRARY -- ]", 420, LibFunc)
        
        
        def create_button2(text, y_pixel_pos1, command):
            btn1 = tk.Label(window, text=text, font=("Trebuchet MS", 15), fg="#F6FAFC",
                    bg="#020638", bd=0)
            btn1.bind("<Enter>", lambda e: btn1.config(fg="#1F5186"))
            btn1.bind("<Leave>", lambda e: btn1.config(fg="#F6FAFC"))
            btn1.bind("<Button-1>", lambda e: command())
            btn1.place(x=800, y=y_pixel_pos1)  

        create_button2("[|-- CAM --|]", 20, CamMen)
        create_button2("[ -- Panelling : Multi Board -- ]", 100, PanelFunc)
        create_button2("[ -- Panelling : Multi Board (Altium Only)-- ]", 140, PanelAltFunc)
        create_button2("[ -- Panelling : Same Board Multi Copying -- ]", 180, PanelCopyFunc)
        create_button2("[ -- Layer Reorder,Type,Color(Altium Only)-- ]", 220, PanelRTCFunc)
        create_button2("[ -- IPC Check : Same Board MultiCopied -- ]", 260, IPCSameBrdFunc)
        create_button2("[ -- IPC Check : Multiple Board -- ]", 300, IPCMultiBrdFunc)
        create_button2("[ -- Gerber Compare (Altium Only) -- ]", 340, GerberCompFunc)
        create_button2("[ -- Auto DFM Check -- ]", 380, AutoDFMChkFunc)
        create_button2("[ -- Auto DFM Check(MIN VAL-10) (Altium Only) ]", 420, AutoDFMChkDVFunc)
        create_button2("[ -- Auto DFM Drill to Copper Layer Order ]", 460, AutoLayerOrderFunc)
        create_button2("[ -- Electrical Layer Reorder,Type,Color (By .Txt Inp) ]", 500, AutoLayerRTCByInpFunc)


#------------------------------------------------------------------------------------------------------------------------------------------------------------------


    def SimFunc():
        global  cap, after_id


        UsedDt = "Opened - Simulation Menu"
        PurposeList.append(UsedDt)
        
        if after_id:
            window.after_cancel(after_id)
            after_id = None
        if cap:
            cap.release()
        for widget in window.winfo_children():
            widget.destroy()

        cap_path = resource_path("VidMain1.mp4")
        cap = cv2.VideoCapture(cap_path)
        bg_label = tk.Label(window)
        bg_label.place(relwidth=1, relheight=1)
        update_video_frame(bg_label, cap_path)
        
#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def StackupUpdateExport():
            UsedDt = "Simulation Menu - Used : StackUp Export"
            PurposeList.append(UsedDt)
            import tkinter as tk
            from tkinter import filedialog
            from tkinter import messagebox
            messagebox.showinfo("StackUp Export", "Please Select The StackUp File To Export Excel..!")

            def STK_file():
                root = tk.Tk()
                root.withdraw()  
                STKfile_path = filedialog.askopenfilename(title="Select a STK File", filetypes=[("STK Files", "*.stk")])
                return STKfile_path

            STKfile_path = STK_file()
            if STKfile_path:
                try:
                    with open(STKfile_path, 'r') as file:
                        StackUpList = [line.strip() for line in file.readlines()]
                except Exception as e:
                    messagebox.showerror("Error", f"Error reading the file: {e}")
            else:
                messagebox.showwarning("No File Selected", "You did not select any file.")
            
            DummyVar = "Break"
            PassVar = "Nil"
            ExpList = []
            ExpHead = "LayerName,Thickness,Dielectric Material,Dielectric Fill"
            ExpList.append(ExpHead)
            
            for StackUpListlp in StackUpList:
                if DummyVar == "Pass":
                    if StackUpListlp != "$end 'Layers'": 
                        LinePrse = StackUpListlp.split(",")
                        for LinePrselp in LinePrse:
                            if "LayerName=" in LinePrselp:
                                LyrPrse = LinePrselp.split("=")
                                if LyrPrse[1] == "''":
                                    LyrPrseUp = "Nil"
                                else:
                                    LyrPrseUp = LyrPrse[1].replace("'", "")
                            if "Material=" in LinePrselp:
                                MatPrse = LinePrselp.split("=")
                                if MatPrse[1] == "''":
                                    MatPrseUp = "Nil"
                                else:
                                    MatPrseUp = MatPrse[1].replace("'", "")
                            if "Thickness=" in LinePrselp:
                                ThiPrse = LinePrselp.split("=")
                                if ThiPrse[1] == "''":
                                    ThiPrseUp = "Nil"
                                else:
                                    ThiPrseUp = ThiPrse[1].replace("'", "")
                            if "DefinedDielectricFill=" in LinePrselp:
                                DefPrse = LinePrselp.split("=")
                                if DefPrse[1] == "''":
                                    DefPrseUp = "Nil"
                                else:
                                    DefPrseUp = DefPrse[1].replace("'", "")
                        ExpVal = f"{LyrPrseUp},{ThiPrseUp},{MatPrseUp},{DefPrseUp}"
                        ExpList.append(ExpVal)
                if StackUpListlp == "$begin 'Layers'":
                    DummyVar = "Pass"
                elif StackUpListlp == "$end 'Layers'":
                    DummyVar = "Break"
            
            import pandas as pd
            import tkinter as tk
            from tkinter import filedialog

            data = [item.split(",") for item in ExpList]

            # Create a DataFrame
            df = pd.DataFrame(data)

            # Ask user for the file location and name to save the Excel file
            root = tk.Tk()
            root.withdraw()  # Hide the root window
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])

            if file_path:
                # Save the DataFrame to Excel
                df.to_excel(file_path, index=False, header=False)  # Save without index and headers
                print(f"File saved at: {file_path}")
                messagebox.showinfo("Stack Up Export", f"Result successfully saved to {file_path}")
            else:
                print("No file selected.")

#------------------------------------------------------------------------------------------------------------------------------------------------------------------
        
        def StackupUpdateImport():
            UsedDt = "Simulation Menu - Used : StackUp Import"
            PurposeList.append(UsedDt)
            import pandas as pd
            import tkinter as tk
            from tkinter import messagebox
            from tkinter import filedialog

            messagebox.showinfo("StackUp Import", "Please Select The Input Excel File To Update Values In Stackup File..!")
            def get_file_path():
                root = tk.Tk()
                root.withdraw()  
                file_path = filedialog.askopenfilename(
                    title="Select Excel file", 
                    filetypes=[("Excel files", "*.xlsx;*.xls")]
                )
                return file_path
            file_path = get_file_path()
            df = pd.read_excel(file_path, header=None)
            data = df.iloc[1:, [0, 1, 2, 3]].values.tolist()
            formatted_data = [f"{row[0]},{row[1]},{row[2]},{row[3]}" for row in data]
            DiMatList = [f"{row[2]}" for row in data]
            DiFillList = [f"{row[3]}" for row in data]
            DiFillList = [item for item in DiFillList if item != "Nil"]
            DiFillList = [item for item in DiFillList if item != "nan"]

            PassVar = "GO"
            import math
            for RowLp in data:
                DiMatVal = RowLp[2]
                DiMatVal = DiMatVal.upper()
                if "COPPER" in DiMatVal:
                    ValDt = RowLp[3]
                    if ValDt == "Nil":
                        PassVar = "BLOCK"
                    if ValDt == "nan":
                        PassVar = "BLOCK"
                    if isinstance(ValDt, float) and math.isnan(ValDt):
                        PassVar = "BLOCK"
                
            non_duplicates = [item for item in DiFillList if item not in DiMatList]
            if len(non_duplicates) == 0 and PassVar == "GO":
                import tkinter as tk
                from tkinter import filedialog
                from tkinter import messagebox
                messagebox.showinfo("StackUp Export", "Please Select The StackUp File To Update..!")

                def STK_file():
                    root = tk.Tk()
                    root.withdraw()  
                    STKfile_path = filedialog.askopenfilename(title="Select a STK File", filetypes=[("STK Files", "*.stk")])
                    return STKfile_path

                STKfile_path = STK_file()
                if STKfile_path:
                    try:
                        with open(STKfile_path, 'r') as file:
                            StackUpList = [line.strip() for line in file.readlines()]
                    except Exception as e:
                        messagebox.showerror("Error", f"Error reading the file: {e}")
                else:
                    messagebox.showwarning("No File Selected", "You did not select any file.") 
                
                ElevationList = []
                for formatted_datalp in formatted_data:
                    ElevationPass = formatted_datalp.split(",")
                    ElevationDt = ElevationPass[1] 
                    if ElevationDt == "nan":
                        ElevationDt = 0
                    else:
                        ElevationDt = float(ElevationDt)
                    ElevationList.append(ElevationDt)
                
                ElevationSumList = []
                for i in range(len(ElevationList)):
                    EleVal = sum(ElevationList[i + 1:])
                    ElevationSumList.append(EleVal)
                
                DummyVar = "Break"
                UpVal = 0
                FinalList = []
                NewMatList = []
                for StackUpListlp in StackUpList:
                    if DummyVar == "Pass":
                        if StackUpListlp != "$end 'Layers'": 
                            LinePrse = StackUpListlp.split(",")
                            UpdataPrse = formatted_data[UpVal].split(",")
                            UpdatedTxt = StackUpListlp
                            for LinePrselp in LinePrse:
                                if "LayerName=" in LinePrselp:
                                    LyrNameChk = UpdataPrse[0]
                                    if LyrNameChk == "Nil" or LyrNameChk == "nan":
                                        LyrStr = f"LayerName=''"
                                    else:
                                        LyrStr = f"LayerName='{UpdataPrse[0]}'"
                                    UpdatedTxt = UpdatedTxt.replace(LinePrselp, LyrStr)
                                if "Material=" in LinePrselp:
                                    MatNameChk = UpdataPrse[2]
                                    NewMatList.append(MatNameChk)
                                    if MatNameChk == "Nil" or MatNameChk == "nan":
                                        MatStr = f"Material=''"
                                    else:
                                        MatStr = f"Material='{UpdataPrse[2]}'"
                                    UpdatedTxt = UpdatedTxt.replace(LinePrselp, MatStr)
                                if "Thickness=" in LinePrselp:
                                    ThiNameChk = UpdataPrse[1]
                                    if ThiNameChk == "Nil" or ThiNameChk == "nan":
                                        Zero = 0
                                        ThiStr = f"Thickness={Zero}"
                                    else:
                                        UpFlt = float(UpdataPrse[1])
                                        ThiStr = f"Thickness={UpFlt}"
                                    UpdatedTxt = UpdatedTxt.replace(LinePrselp, ThiStr)
                                if "Elevation=" in LinePrselp:
                                    Elepass  = ElevationSumList[UpVal]
                                    EleStr = f"Elevation={Elepass}"
                                    UpdatedTxt = UpdatedTxt.replace(LinePrselp, EleStr)
                                if "DefinedDielectricFill=" in LinePrselp:
                                    DefNameChk = UpdataPrse[3]
                                    if DefNameChk == "Nil" or DefNameChk == "nan":
                                        DefStr = f"DefinedDielectricFill=''"
                                    else:
                                        DefStr = f"DefinedDielectricFill='{UpdataPrse[3]}'"
                                    UpdatedTxt = UpdatedTxt.replace(LinePrselp, DefStr)
                                    
                            FinalList.append(UpdatedTxt)
                            UpVal = UpVal + 1 
                            
                    if StackUpListlp == "$begin 'Layers'":
                        FinalList.append(StackUpListlp)
                        DummyVar = "Pass"
                    elif StackUpListlp == "$end 'Layers'":
                        FinalList.append(StackUpListlp)
                        DummyVar = "Break"
                        
                    elif DummyVar == "Break":
                        if StackUpListlp != "$end 'Layers'":
                            FinalList.append(StackUpListlp)
                    
                    
                    
                NewMatList = list(set(NewMatList)) 
                
                MatList = []
                ExtMatNameList = []
                DummyVarMat = "NO"
                for FinalListlp in FinalList:
                    if DummyVarMat == "PASS":
                        if FinalListlp != "$end 'Materials'":
                            MatList.append(FinalListlp)
                            if "Name=" in FinalListlp:
                                MatName = FinalListlp.replace("'", "")
                                MatNameup = MatName.split("=")
                                ExtMatNameList.append(MatNameup[1])
                        
                    if FinalListlp == "$begin 'Materials'":
                        DummyVarMat = "PASS"
                    elif FinalListlp == "$end 'Materials'":
                        DummyVarMat = "NO"
                
                
                MissMatList = [item for item in NewMatList if item not in ExtMatNameList]
                

                for MissMatListlp in MissMatList:
                    Val1 = "$begin 'Insulator'"
                    MatList.append(Val1)
                    
                    Val2 = f"Name='{MissMatListlp}'"
                    MatList.append(Val2)
                    
                    Val3 = "Permittivity=3.1"
                    MatList.append(Val3)
                    
                    Val4 = "LossTangent=0.035"
                    MatList.append(Val4)
                    
                    Val5 = "DerivedFrom='vacuum'"
                    MatList.append(Val5)
                    
                    Val6 = "AnsoftID=0"
                    MatList.append(Val6)
                    
                    Val7 = "$begin 'Djordjevic-Sarkar'"
                    MatList.append(Val7)
                    
                    Val8 = "MeasurementFreq=1000000000"
                    MatList.append(Val8)
                    
                    Val9 = "SigmaDC=1e-12"
                    MatList.append(Val9)
                    
                    Val10 = "EpsDC=0"
                    MatList.append(Val10)
                    
                    Val11 = "$end 'Djordjevic-Sarkar'"
                    MatList.append(Val11)
                    
                    Val12 = "$end 'Insulator'"
                    MatList.append(Val12)


                FirstSetList = []
                DummypassSet1 = "GO"
                for FinalListlp1 in FinalList:
                    if DummypassSet1 == "GO":
                        FirstSetList.append(FinalListlp1)
                    if FinalListlp1 == "$begin 'Materials'":
                        DummypassSet1 = "Break"
                
                SecSetList = []
                DummypassSet2 = "Break"
                for FinalListlp2 in FinalList:
                    if DummypassSet2 == "GO":
                        SecSetList.append(FinalListlp2)
                    if FinalListlp2 == "$end 'Materials'":
                        SecSetList.append(FinalListlp2)
                        DummypassSet2 = "GO"
                
                FinalListUP = FirstSetList + MatList + SecSetList  
                    

                import tkinter as tk
                from tkinter import filedialog

                root = tk.Tk()
                root.withdraw()  
                file_path = filedialog.asksaveasfilename(defaultextension=".stk", filetypes=[("STK Files", "*.stk")])

                if file_path:
                    with open(file_path, 'w') as file:
                        for item in FinalListUP:
                            file.write(item + '\n')  
                    print(f"File saved at: {file_path}")
                    messagebox.showinfo("Stack Up Import", f"Result successfully saved to {file_path}")
                else:
                    print("No file selected.")
            else:
                messagebox.showerror("Error", f"Error In Input Values Please Check..")
        
#------------------------------------------------------------------------------------------------------------------------------------------------------------------
        
        def create_button(text, y_pixel_pos, command):
            btn = tk.Label(window, text=text, font=("Trebuchet MS", 15), fg="#F6FAFC",
                    bg="#020638", bd=0)
            btn.bind("<Enter>", lambda e: btn.config(fg="#1F5186"))
            btn.bind("<Leave>", lambda e: btn.config(fg="#F6FAFC"))
            btn.bind("<Button-1>", lambda e: command())
            btn.place(x=20, y=y_pixel_pos)  

        create_button("[<-- BACK ]", 20, show_main_menu)
        create_button("[ -- Sim Stackup Export -- ]", 100, StackupUpdateExport)
        create_button("[ -- Sim Stackup Import -- ]", 140, StackupUpdateImport)


#------------------------------------------------------------------------------------------------------------------------------------------------------------------


    def SchFunc():
        global  cap, after_id


        UsedDt = "Opened - Application Menu"
        PurposeList.append(UsedDt)
        
        if after_id:
            window.after_cancel(after_id)
            after_id = None
        if cap:
            cap.release()
        for widget in window.winfo_children():
            widget.destroy()

        cap_path = resource_path("VidMain1.mp4")
        cap = cv2.VideoCapture(cap_path)
        bg_label = tk.Label(window)
        bg_label.place(relwidth=1, relheight=1)
        update_video_frame(bg_label, cap_path)
        
#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def NetLisDt():
            UsedDt = "Application Menu - Used : CH Mapping Netlist"
            PurposeList.append(UsedDt)
            
            for widget in window.winfo_children():
                if isinstance(widget, tk.Label) and hasattr(widget, 'image'):
                    continue  # Keep background video label
                widget.destroy()
                
            from tkinter import filedialog
            import numpy as np
            import csv
            from tkinter.filedialog import asksaveasfilename
            
            def NetLisFunc():
                input1 = entry1.get()
                DutLetter = input1.upper()
                if ',' in DutLetter:
                    DutList = DutLetter.split(",")
                else:
                    DutList = []
                    DutList.append(DutLetter)
                
                input2 = entry2.get()
                Pogo = input2.upper()
                if ',' in Pogo:
                    PogoList = Pogo.split(",")
                else:
                    PogoList = []
                    PogoList.append(Pogo)
            
                input3 = entry3.get()
                GndSkip = input3.upper()
                if ',' in GndSkip:
                    GndSkipList = GndSkip.split(",")
                else:
                    GndSkipList = []
                    GndSkipList.append(GndSkip)
                
                input4 = entry4.get()
                CompSkip = input4.upper()
                if ',' in CompSkip:
                    CompSkipList = CompSkip.split(",")
                else:
                    CompSkipList = []
                    CompSkipList.append(CompSkip)
                
                def open_file():
                    file_path = filedialog.askopenfilename(title="Select a DAT File", filetypes=[("DAT Files", "*.dat")])
                    if file_path:  # If a file was selected
                        try:
                            with open(file_path, 'r') as file:
                                content_list = file.readlines()  
                                content_list = [line.strip() for line in content_list]
                                
                                my_array = np.array(content_list)
                                LenVal = len(content_list)
                                FilteredList = []
                                for i in range(LenVal):
                                    if my_array[i] == "NET_NAME":
                                        j = i + 1
                                        PassNetName = my_array[i + 1][1:-1]
                                        StrNetName = "NET_NAME" + "--" + PassNetName
                                        FilteredList.append(StrNetName)
                                        while j <= LenVal - 1:
                                            parsed_list = my_array[j].split()
                                            FirstEle = parsed_list[0]
                                            if FirstEle == "NODE_NAME":
                                                StrPassVal = "NODE_NAME" + "--" + parsed_list[1] + "--" + parsed_list[2] 
                                                FilteredList.append(StrPassVal)
                                            if my_array[j] == "NET_NAME":
                                                break
                                            j += 1
                                            
                                #print(len(FilteredList)) 
                            
                                RemoveList = []
                                for mRem in range(len(FilteredList)):
                                    NetFilt = FilteredList[mRem].split("--") 
                                    NetNameFilt = NetFilt[1]
                                    if any(gnd1 in NetNameFilt for gnd1 in GndSkipList):
                                        RemVal = mRem + 1
                                        while RemVal <= len(FilteredList) - 1:
                                            BrkNN = FilteredList[RemVal].split("--") 
                                            BrkNNVal = BrkNN[0] 
                                            if BrkNNVal == "NET_NAME":
                                                break
                                            else:
                                                RemoveList.append(FilteredList[RemVal])  
                                            
                                            RemVal += 1

                                set_b = set(RemoveList)

                                result = [item for item in FilteredList if item not in set_b]
        
                                NetListArr = np.array(result)
                                NetListLen = len(result)
                                #print(NetListLen)
                                RefList = []
                                for k in range(NetListLen - 1):
                                    RefPrse = NetListArr[k].split("--")
                                    RefFilt = RefPrse[1]
                                    for DutListlp in DutList:
                                        InpDutValLen = len(DutListlp)
                                        ListDutFilt = (RefFilt[:InpDutValLen])
                                        if ListDutFilt == DutListlp:
                                            RefList.append(RefFilt)
                                    
                                Unq_RefList = list(set(RefList))
                                
                                FinalList = []
                                l = 0
                                while l <= NetListLen - 1:
                                    NamePrse = NetListArr[l].split("--")
                                    NameFilt = NamePrse[0]
                                    if NameFilt == "NET_NAME":
                                        StartPoint = l + 1
                                        NetName = NamePrse[1]
                                    else:
                                        RefVal = NamePrse[1] 
                                        PinVal = NamePrse[2] 
                                        for Reflp in Unq_RefList:
                                            if Reflp == RefVal:
                                                PassList = []
                                                while StartPoint <= NetListLen - 1:
                                                    PogoVal = NetListArr[StartPoint].split("--") 
                                                    BreakPoint = PogoVal[0]
                                                    if BreakPoint == "NET_NAME":
                                                        break
                                                    else:
                                                        DummyPassVar = "GO"
                                                        PogoRefVal = PogoVal[1]
                                                        PogoPinVal = PogoVal[2]
                                                        PassNodeVal = PogoRefVal + "-" + PogoPinVal
                                                        for PogoListlp in PogoList:
                                                            InpPogoValLen = len(PogoListlp)
                                                            ListPogoFilt = (PogoRefVal[:InpPogoValLen])
                                                            if ListPogoFilt == PogoListlp:
                                                                DummyPassVar = "NO"
                                                                PassPrintVal = "DUT PIN - " + RefVal + "." + PinVal + "," + "DUT SIGNAL - " + NetName + "," +  "POGO PIN - " + PogoRefVal + "." + PogoPinVal + "," + "POGO SIGNAL - " + NetName
                                                                if PassPrintVal not in FinalList:
                                                                    FinalList.append(PassPrintVal)
                                                        
                                                        if PogoRefVal == RefVal:
                                                            DummyPassVar = "NO"
                                                        
                                                        for CompSkipListlp in CompSkipList:
                                                            if CompSkipListlp in PogoRefVal:
                                                                DummyPassVar = "NO"
                                                        
                                                        for GndSkipListlp in GndSkipList:
                                                            if GndSkipListlp in NetName:
                                                                DummyPassVar = "NO"
                                                        
                                                        if DummyPassVar == "GO":
                                                            PassList.append(PassNodeVal)       
                                                    StartPoint += 1
                                                    
                                                n = 0
                                                PassListLen = len(PassList)
                                                if PassListLen != 0:
                                                    while n <= PassListLen - 1:
                                                        DummyPass = PassList[n].split("-")
                                                        DummyPassRef = DummyPass[0]
                                                        DummyPassPin = DummyPass[1]
                                                        m = 0
                                                        for m in range(len(NetListArr)):
                                                            PassArrVal = NetListArr[m].split("--") 
                                                            PassArrValStr = PassArrVal[0]
                                                            if PassArrValStr == "NET_NAME":
                                                                PogoNetName = PassArrVal[1]
                                                                StoreNetLoc = m + 1
                                                            else:
                                                                PassRefName = PassArrVal[1]
                                                                PassPinName = PassArrVal[2]
                                                                if PassRefName == DummyPassRef and PassPinName != DummyPassPin:
                                                                    while StoreNetLoc <= NetListLen - 1:
                                                                        PassArrVal1 = NetListArr[StoreNetLoc].split("--") 
                                                                        PassArrValStr1 = PassArrVal1[0]
                                                                        if PassArrValStr1 == "NET_NAME":
                                                                            break
                                                                        else:
                                                                            GetRefName = PassArrVal1[1]
                                                                            GetPinName = PassArrVal1[2]
                                                                            SPassNodeVal = f"{GetRefName}-{GetPinName}"
                                                                            if SPassNodeVal not in PassList:
                                                                                DummyGndPass = "YES"
                                                                                if any(gnd in PogoNetName for gnd in GndSkipList):
                                                                                        DummyGndPass = "NO"
                                                                                        
                                                                                if DummyGndPass == "YES":
                                                                                    PassVar = "ADD"
                                                                                    if any(GetRefName[:len(pogo)] == pogo for pogo in PogoList):
                                                                                        PassVar = "NO"
                                                                                        PassPrintVal1 = f"DUT PIN - {RefVal}.{PinVal}, DUT SIGNAL - {NetName}, POGO PIN - {GetRefName}.{GetPinName}, POGO SIGNAL - {PogoNetName}"
                                                                                        if PassPrintVal1 not in FinalList:
                                                                                            FinalList.append(PassPrintVal1)
                                                                                    
                                                                                    if any(GetRefName[:len(dut)] == dut for dut in DutList):
                                                                                            PassVar = "NO"
                                                                                            
                                                                                    if PassVar == "ADD":
                                                                                        PassList.append(SPassNodeVal)                                                                                           
                                                                                        PassListLen = len(PassList)
                                                                                
                                                                                
                                                                        StoreNetLoc += 1                                                          
                                                        
                                                        n += 1                                                                  
                                    l += 1
                                            
                                input_list = list(set(FinalList))
                                
                                from collections import defaultdict

                                grouped_data = defaultdict(lambda: {'POGO_PINS': [], 'POGO_SIGNALS': []})

                                for item in input_list:
                                    parts = item.split(',')
                                    dut_pin = parts[0]  
                                    dut_signal = parts[1]  
                                    pogo_pin = parts[2].replace("POGO PIN - ", "")  
                                    pogo_signal = parts[3].replace("POGO SIGNAL - ", "")  

                                    grouped_data[(dut_pin, dut_signal)]['POGO_PINS'].append(pogo_pin)
                                    grouped_data[(dut_pin, dut_signal)]['POGO_SIGNALS'].append(pogo_signal)

                                output_list = []

                                for (dut_pin, dut_signal), data in grouped_data.items():
                                    pogo_pins = '/'.join(data['POGO_PINS'])
                                    pogo_signals = '/'.join(data['POGO_SIGNALS'])
                                    output_item = f"{dut_pin},{dut_signal},POGO PIN - {pogo_pins},POGO SIGNAL - {pogo_signals}"
                                    output_list.append(output_item)



                                import csv
                                from tkinter import Tk
                                from tkinter.filedialog import asksaveasfilename
                                Tk().withdraw()  
                                SaveFile = asksaveasfilename(defaultextension='.csv', filetypes=[('CSV Files', '*.csv'), ('All Files', '*.*')],
                                                            title="Save CSV File")

                                if SaveFile:
                                    with open(SaveFile, mode='w', newline='') as file:
                                        writer = csv.writer(file)
                                        for line in output_list:
                                            writer.writerow(line.split(','))  
                                        
                                    print(f"File saved to {SaveFile}")
                                else:
                                    print("No file selected.")           
                                        
                        except Exception as e:
                            print(f"Error opening the file: {e}")
                    else:
                        print("No file was selected.")

                NetL = tk.Tk()
                NetL.withdraw()  
                open_file()
            
            def create_button(text, relx, rely, command):
                btn = tk.Label(window, text=text, font=("Trebuchet MS", 20), fg="#F6FAFC", bg="#020638", bd=0)
                btn.bind("<Enter>", lambda e: btn.config(fg="#1F5186"))
                btn.bind("<Leave>", lambda e: btn.config(fg="#F6FAFC"))
                btn.bind("<Button-1>", lambda e: command())
                btn.place(relx=relx, rely=rely, anchor="center")
                return btn



            # ------------------ Centered Input Fields ----------------------
            form_start_y = 0.25
            spacing = 0.12
            entry_width = 30

            def create_label_entry(text, rel_y):
                label = tk.Label(window, text=text, font=("Didot", 12), bg="#f0f0f0")
                label.place(relx=0.5, rely=rel_y, anchor="center")
                entry = tk.Entry(window, font=("Didot", 12), width=entry_width)
                entry.place(relx=0.5, rely=rel_y + 0.05, anchor="center")
                return entry

            entry1 = create_label_entry("[ ENTER DUT COMMON NAME ]", form_start_y)
            entry2 = create_label_entry("[ ENTER POGO COMMON NAME ]", form_start_y + spacing)
            entry3 = create_label_entry("[ ENTER NET NAMES TO SKIP ]", form_start_y + spacing * 2)
            entry4 = create_label_entry("[ ENTER COMP NAME TO SKIP ]", form_start_y + spacing * 3)

            create_button("[ CLICK HERE ]", 0.5, form_start_y + spacing * 4.5, NetLisFunc)
            create_button("[<-- BACK ]", 0.1, 0.05, SchFunc)

#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def BGAPinListFunc():
            UsedDt = "Application Menu - Used : BGA Pin List Check"
            PurposeList.append(UsedDt)
            
            from tkinter import messagebox
            messagebox.showinfo("BGA Pin List Check", "Please Select The Input Excel File..!")
            
            import pandas as pd
            from tkinter import Tk
            from tkinter.filedialog import askopenfilename

            Tk().withdraw()
            file_path = askopenfilename(title="Select Excel file", filetypes=[("Excel files", "*.xlsx;*.xls")])

            if file_path:
                excel_data = pd.ExcelFile(file_path)
                PinMapList = []
                PinList = []
                for sheet_name in excel_data.sheet_names:
                    df = pd.read_excel(excel_data, sheet_name=sheet_name, header=None)

                    rows, columns = df.shape
                    #print(f"Sheet: {sheet_name} | Rows: {rows} | Columns: {columns}")

                    if columns > 10:
                        for i in range(1, rows):
                            valueRow = df.iloc[i, 0]  
                            for j in range(1, columns):
                                valueCol = df.iloc[i, j]  
                                valueHead = df.iloc[0, j]
                                if pd.notna(valueCol) and pd.notna(valueRow):  
                                    if pd.notna(valueHead):
                                        DataPass = f"{valueRow}{valueHead},{valueCol}"
                                        PinMapList.append(DataPass)
                    else:
                        for k in range(1, rows):
                            ValPinName = df.iloc[k, 1] 
                            ValNetName = df.iloc[k, 2] 
                            DataPass1 = f"{ValPinName},{ValNetName}"
                            PinList.append(DataPass1)
                
                ErrorList = []
                PinMapErrList = []
                PinErrList = []
                PinMapList = sorted(PinMapList, key=lambda x: x.split(',')[1])
                PinList = sorted(PinList, key=lambda y: y.split(',')[1])
                if PinList == PinMapList:
                    count_dict = {}
                    output_list = []
                    for item in PinMapList:
                        first_part, second_part = item.split(',')
                        
                        if second_part in count_dict:
                            count_dict[second_part] += 1
                            new_item = f"{first_part},{second_part}_{count_dict[second_part]}"
                        else:
                            count_dict[second_part] = 1
                            new_item = f"{first_part},{second_part}"
                        
                        output_list.append(new_item)
                else:
                    PinMapMissList = list(set(PinMapList) - set(PinList))
                    PinDtMissList = list(set(PinList) - set(PinMapList))
                    
                    for PinMapMissListlp in PinMapMissList:
                        ErrDtPM = f"{PinMapMissListlp},This Found in PinMap Not Found in PinList.."
                        PinMapErrList.append(ErrDtPM)
                        
                        
                    for PinDtMissListlp in PinDtMissList:
                        ErrDtPL = f"{PinDtMissListlp},This Found in PinList Not Found in PinMap.."
                        PinErrList.append(ErrDtPL)
                    
                    ErrorList = PinMapErrList + PinErrList
                    
                    count_dict = {}
                    output_list = []
                    for item in PinMapList:
                        first_part, second_part = item.split(',')
                        
                        if second_part in count_dict:
                            count_dict[second_part] += 1
                            new_item = f"{first_part},{second_part}_{count_dict[second_part]}"
                        else:
                            count_dict[second_part] = 1
                            new_item = f"{first_part},{second_part}"
                        
                        output_list.append(new_item)


                    import numpy as np
                    PinListArr = np.array(output_list)
                    ListLen = len(PinListArr)
                    
                    FinalList = []
                    for i in range(ListLen - 1):
                        PinListNxtVal = PinListArr[i + 1]
                        last_underscore_index = PinListNxtVal.rfind('_')
                        PinListPrse = PinListNxtVal[last_underscore_index + 1:]
                        if PinListPrse == "2":
                            RestVal = PinListNxtVal.rfind('_')
                            ChkVal = PinListNxtVal[:RestVal]
                            Prse1 = PinListArr[i].split(",")
                            Prse2 = ChkVal.split(",")
                            if Prse1[1] == Prse2[1]:
                                PassVal = f"{PinListArr[i]}_1"
                                FinalList.append(PassVal)
                        else:
                            FinalList.append(PinListArr[i])
                    FinalList.append(PinListArr[ListLen - 1])
                
                
                import pandas as pd
                import tkinter as tk
                from tkinter import filedialog

                df1 = pd.DataFrame([item.split(',') for item in FinalList], columns=['PIN_NUMBER', 'PIN_NAME'])
                if ErrorList != []:
                    df2 = pd.DataFrame([item.split(',') for item in ErrorList], columns=['PIN_NUMBER', 'PIN_NAME', 'ERROR'])

                root = tk.Tk()
                root.withdraw()  

                file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

                if file_path:
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        df1.to_excel(writer, sheet_name='PIN-MAP', index=False)
                        if ErrorList != []:
                            df2.to_excel(writer, sheet_name='ERROR-LIST', index=False)
                    from tkinter import messagebox
                    messagebox.showinfo("BGA Pin List Check", f"Result successfully saved to {file_path}")
                else:
                    print("No file selected, data not saved.")

            else:
                print("No file selected.")

#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def PinListFunc():
            UsedDt = "Application Menu - Used : Pin List Rename"
            PurposeList.append(UsedDt)
            
            from tkinter import messagebox
            messagebox.showinfo("Pin List Rename", "Please Select The Input Excel File..!")
            
            import pandas as pd
            from tkinter import Tk
            from tkinter.filedialog import askopenfilename

            Tk().withdraw()
            file_path = askopenfilename(title="Select Excel file", filetypes=[("Excel files", "*.xlsx;*.xls")])

            if file_path:
                excel_data = pd.ExcelFile(file_path)
                PinList = []
                for sheet_name in excel_data.sheet_names:
                    df = pd.read_excel(excel_data, sheet_name=sheet_name, header=None)

                    rows, columns = df.shape
                    for k in range(1, rows):
                        ValPinName = df.iloc[k, 1] 
                        ValNetName = df.iloc[k, 2] 
                        DataPass1 = f"{ValPinName},{ValNetName}"
                        PinList.append(DataPass1)

                PinList = sorted(PinList, key=lambda x: x.split(',')[1])
                count_dict = {}
                output_list = []

                for item in PinList:
                    first_part, second_part = item.split(',')
                    
                    if second_part in count_dict:
                        count_dict[second_part] += 1
                        new_item = f"{first_part},{second_part}_{count_dict[second_part]}"
                    else:
                        count_dict[second_part] = 1
                        new_item = f"{first_part},{second_part}"
                    
                    output_list.append(new_item)
                    
                
                import numpy as np
                PinListArr = np.array(output_list)
                ListLen = len(PinListArr)
                
                FinalList = []
                for i in range(ListLen - 1):
                    PinListNxtVal = PinListArr[i + 1]
                    last_underscore_index = PinListNxtVal.rfind('_')
                    PinListPrse = PinListNxtVal[last_underscore_index + 1:]
                    if PinListPrse == "2":
                        RestVal = PinListNxtVal.rfind('_')
                        ChkVal = PinListNxtVal[:RestVal]
                        Prse1 = PinListArr[i].split(",")
                        Prse2 = ChkVal.split(",")
                        if Prse1[1] == Prse2[1]:
                            PassVal = f"{PinListArr[i]}_1"
                            FinalList.append(PassVal)
                    else:
                        FinalList.append(PinListArr[i])
                FinalList.append(PinListArr[ListLen - 1])



                import pandas as pd
                import tkinter as tk
                from tkinter import filedialog

                df1 = pd.DataFrame([item.split(',') for item in FinalList], columns=['PIN_NUMBER', 'PIN_NAME'])

                root = tk.Tk()
                root.withdraw()  

                file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

                if file_path:
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        df1.to_excel(writer, sheet_name='PINLIST-RENAME', index=False)

                    from tkinter import messagebox
                    messagebox.showinfo("Pin List Rename", f"Result successfully saved to {file_path}")
                else:
                    print("No file selected, data not saved.")

                

            else:
                print("No file selected.")

#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def TempNetListComp():

            UsedDt = "Application Menu - Used : NetList Compare"
            PurposeList.append(UsedDt)
            
            import tkinter as tk
            import numpy as np
            from tkinter import filedialog
            from tkinter import messagebox
            messagebox.showinfo("NetList Compare", "Please Select The Template Netlist File..!")
                
            def DAT_file():
                root = tk.Tk()
                root.withdraw()  
                DATfile_path = filedialog.askopenfilename(title="Select a DAT File", filetypes=[("DAT Files", "*.dat")])
                return DATfile_path

            DATfile_path = DAT_file()
            if DATfile_path:
                try:
                    with open(DATfile_path, 'r') as file:
                        TempList = [line.strip() for line in file.readlines()]
                except Exception as e:
                    messagebox.showerror("Error", f"Error reading the file: {e}")
            else:
                messagebox.showwarning("No File Selected", "You did not select any file.") 
                
            
            my_array = np.array(TempList)
            LenVal = len(TempList)
            FilteredList = []
            for i in range(LenVal):
                if my_array[i] == "NET_NAME":
                    j = i + 1
                    PassNetName = my_array[i + 1][1:-1]
                    StrNetName = "NET_NAME" + "--" + PassNetName
                    FilteredList.append(StrNetName)
                    while j <= LenVal - 1:
                        parsed_list = my_array[j].split()
                        FirstEle = parsed_list[0]
                        if FirstEle == "NODE_NAME":
                            StrPassVal = "NODE_NAME" + "--" + parsed_list[1] + "--" + parsed_list[2] 
                            FilteredList.append(StrPassVal)
                        if my_array[j] == "NET_NAME":
                            break
                        j += 1
                        
            
            NetListArr = np.array(FilteredList)
            NetListLen = len(FilteredList)
            
            
            TempList = []
            l = 0
            while l <= NetListLen - 1:
                NamePrse = NetListArr[l].split("--")
                NameFilt = NamePrse[0]
                if NameFilt == "NET_NAME":
                    NetNameTemp = NamePrse[1]
                else:
                    RefValTemp = NamePrse[1] 
                    PinValTemp = NamePrse[2]  
                    TempStr = f"{NetNameTemp},{RefValTemp},{PinValTemp}"
                    TempList.append(TempStr)
                    
                l += 1
            
            
            import tkinter as tk
            from tkinter import filedialog
            from tkinter import messagebox
            messagebox.showinfo("NetList Compare", "Please Select The Current Netlist File..!")
                
            def CDAT_file():
                root = tk.Tk()
                root.withdraw()  
                CDATfile_path = filedialog.askopenfilename(title="Select a DAT File", filetypes=[("DAT Files", "*.dat")])
                return CDATfile_path

            CDATfile_path = CDAT_file()
            if CDATfile_path:
                try:
                    with open(CDATfile_path, 'r') as file:
                        CurrentList = [line.strip() for line in file.readlines()]
                except Exception as e:
                    messagebox.showerror("Error", f"Error reading the file: {e}")
            else:
                messagebox.showwarning("No File Selected", "You did not select any file.") 
            
            
            my_arrayC = np.array(CurrentList)
            LenValC = len(CurrentList)
            FilteredListC = []
            for iC in range(LenValC):
                if my_arrayC[iC] == "NET_NAME":
                    jC = iC + 1
                    PassNetNameC = my_arrayC[iC + 1][1:-1]
                    StrNetNameC = "NET_NAME" + "--" + PassNetNameC
                    FilteredListC.append(StrNetNameC)
                    while jC <= LenValC - 1:
                        parsed_listC = my_arrayC[jC].split()
                        FirstEleC = parsed_listC[0]
                        if FirstEleC == "NODE_NAME":
                            StrPassValC = "NODE_NAME" + "--" + parsed_listC[1] + "--" + parsed_listC[2] 
                            FilteredListC.append(StrPassValC)
                        if my_arrayC[jC] == "NET_NAME":
                            break
                        jC += 1
                        
            
            NetListArrC = np.array(FilteredListC)
            NetListLenC = len(FilteredListC)
            
            
            CurrList = []
            lC = 0
            while lC <= NetListLenC - 1:
                NamePrseC = NetListArrC[lC].split("--")
                NameFiltC = NamePrseC[0]
                if NameFiltC == "NET_NAME":
                    NetNameTempC = NamePrseC[1]
                else:
                    RefValTempC = NamePrseC[1] 
                    PinValTempC = NamePrseC[2]  
                    TempStrC = f"{NetNameTempC},{RefValTempC},{PinValTempC}"
                    CurrList.append(TempStrC)
                    
                lC += 1
            
            
            #print(len(TempList))
            #print(len(CurrList))
            
            NonMemCurr = list(set(CurrList) - set(TempList))
            
            #NonMemRemp = list(set(TempList) - set(CurrList))
                    
            #print(NonMemCurr)
            
            
            if len(NonMemCurr) > 0:
                import pandas as pd
                import tkinter as tk
                from tkinter import filedialog

                df1 = pd.DataFrame([item.split(',') for item in NonMemCurr], columns=['NET NAME', 'REFDES', 'PIN_NAME'])

                root = tk.Tk()
                root.withdraw()  

                file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

                if file_path:
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        df1.to_excel(writer, sheet_name='ERROR LIST', index=False)

                    from tkinter import messagebox
                    messagebox.showinfo("NetList Compare", f"Result successfully saved to {file_path}")
                else:
                    print("No file selected, data not saved.")
            else:
                from tkinter import messagebox
                messagebox.showinfo("NetList Compare", "No Error Found...!")

#------------------------------------------------------------------------------------------------------------------------------------------------------------------   

        def NetUpFunc():
            UsedDt = "Application Menu - Used : DRC Report"
            PurposeList.append(UsedDt)
            
            import tkinter as tk
            from tkinter import filedialog, messagebox
            from tkinter.filedialog import askopenfilename
            import csv
            import os
            
            messagebox.showinfo("DRC REPORT", "Please Select The Generated CSV File..!")

            tk.Tk().withdraw()
            file_path = askopenfilename(filetypes=[("CSV files", "*.csv")])

            results = []

            if file_path:
                with open(file_path, newline='', encoding='utf-8') as csvfile:
                    reader = csv.reader(csvfile)
                    for row in reader:
                        results.append(row)

                    
                if len(results) > 0:
                    messagebox.showinfo("DRC REPORT", f"Please Select The NET File..! ")
                    def CNET_file():
                        root = tk.Tk()
                        root.withdraw()  
                        CNETfile_path = filedialog.askopenfilename(title="Select a NET File", filetypes=[("NET Files", "*.net")])
                        return CNETfile_path

                    CNETfile_path = CNET_file()
                    if CNETfile_path:
                        try:
                            with open(CNETfile_path, 'r') as file:
                                CurrentList = [line.strip() for line in file.readlines()]
                        except Exception as e:
                            messagebox.showerror("Error", f"Error reading the file: {e}")
                    else:
                        messagebox.showwarning("No File Selected", "You did not select any file.")
                    
                    if len(CurrentList) > 0:
                        PackageList = []
                        NetList = []
                        
                        DummyVar = "Package"
                        for CurrentListlp in CurrentList:
                            if "$NETS" in CurrentListlp:
                                DummyVar = "Net"
                            
                            if DummyVar == "Package":
                                PackageList.append(CurrentListlp)
                            
                            elif DummyVar == "Net":
                                NetList.append(CurrentListlp)
                        
                        #------------------------------------------------------------------
                        if len(NetList) > 0:   
                            UpdatedNetList = []  
                            UpdatedNetReportList = []
                            for NetListlp in NetList:
                                if ";" in NetListlp:
                                    SplitNetName = NetListlp.split(";")
                                    NetListNetName = SplitNetName[0].strip()
                                    
                                    DummyVarChk = "NO"
                                    for resultslp in results:
                                        ResSubLen = len(resultslp)
                                        for i in range(0, ResSubLen):
                                            DrcNetName = resultslp[i].strip()
                                            if DrcNetName == NetListNetName:
                                                DummyVarChk = "YES"
                                                UnderScoreList = []                                                 
                                                for resultslp_US in resultslp:
                                                    if "__" in resultslp_US:
                                                        UnderScoreList.append(resultslp_US)
                                                
                                                if len(UnderScoreList) > 1:
                                                    UnderScoreList = sorted(UnderScoreList)
                                                    StrLength = len(UnderScoreList[0])
                                                    
                                                    if all(len(s) == StrLength for s in UnderScoreList):
                                                        PassSplitNetName = f"{SplitNetName[0]};"
                                                        PassRepNetName = f"{UnderScoreList[0]};"
                                                        NetListlp = NetListlp.replace(PassSplitNetName, PassRepNetName)
                                                        UpdatedNetList.append(NetListlp)
                                                        
                                                        if PassSplitNetName != PassRepNetName:
                                                            UnderScoreList_Str = " AND ".join(UnderScoreList)
                                                            ReplacedDataForReport = f"This Netname : {PassSplitNetName}, Replaced by : {PassRepNetName}, OverallList : {UnderScoreList_Str}"
                                                            UpdatedNetReportList.append(ReplacedDataForReport)
                                                        
                                                    else:
                                                        UnderScoreList_Sort = sorted(UnderScoreList, key=len, reverse=True)
                                                        PassSplitNetName = f"{SplitNetName[0]};"
                                                        PassRepNetName = f"{UnderScoreList_Sort[0]};"
                                                        NetListlp = NetListlp.replace(PassSplitNetName, PassRepNetName)
                                                        UpdatedNetList.append(NetListlp)
                                                        
                                                        if PassSplitNetName != PassRepNetName:
                                                            UnderScoreList_Str = " AND ".join(UnderScoreList)
                                                            ReplacedDataForReport = f"This Netname : {PassSplitNetName}, Replaced by : {PassRepNetName}, OverallList : {UnderScoreList_Str}"
                                                            UpdatedNetReportList.append(ReplacedDataForReport)
                                                else:
                                                    if len(UnderScoreList) == 1:
                                                        PassSplitNetName = f"{SplitNetName[0]};"
                                                        PassRepNetName = f"{UnderScoreList[0]};"
                                                        NetListlp = NetListlp.replace(PassSplitNetName, PassRepNetName)
                                                        UpdatedNetList.append(NetListlp)
                                                        
                                                        if PassSplitNetName != PassRepNetName:
                                                            UnderScoreList_Str = " AND ".join(UnderScoreList)
                                                            ReplacedDataForReport = f"This Netname : {PassSplitNetName}, Replaced by : {PassRepNetName}, OverallList : {UnderScoreList_Str}"
                                                            UpdatedNetReportList.append(ReplacedDataForReport)
                                                        
                                                    else:
                                                        UpdatedNetList.append(NetListlp)
                                                        
                                    if DummyVarChk == "NO":
                                        UpdatedNetList.append(NetListlp)    
                                else:
                                    
                                    if "$NETS" in NetListlp or "$END" in NetListlp:
                                        NonChkPass = f"{NetListlp}"
                                        UpdatedNetList.append(NonChkPass)
                                    else:
                                        NonChkPass = f"    {NetListlp}"
                                        UpdatedNetList.append(NonChkPass)
                                    
                            
                            FinalUpdatedNetList = PackageList + UpdatedNetList
                            
                            def save_net_file(data_list):
                                root = tk.Tk()
                                root.withdraw()

                                file_path = filedialog.asksaveasfilename(
                                    defaultextension=".net",
                                    filetypes=[("NET Files", "*.net")],
                                    title="Save as .net File"
                                )

                                if file_path:
                                    try:
                                        with open(file_path, 'w') as f:
                                            for line in data_list:
                                                f.write(line + '\n')
                                        messagebox.showinfo("Success", f"File saved successfully:\n{file_path}")
                                    except Exception as e:
                                        messagebox.showerror("Error", f"Could not save the file:\n{e}")
                                else:
                                    messagebox.showwarning("Cancelled", "Save operation was cancelled.")

                            save_net_file(FinalUpdatedNetList)
                            
                            UpdatedNetReportList = list(set(UpdatedNetReportList))
                            
                            if len(UpdatedNetReportList) > 0:
                                import os
                                import csv

                                output_dir1 = os.path.dirname(file_path)
                                output_path1 = os.path.join(output_dir1, "DRCModifiedNetListReport.CSV")

                                try:
                                    with open(output_path1, 'w', newline='') as csvfile:
                                        writer = csv.writer(csvfile)
                                        
                                        #writer.writerow(["Netname", "Replaced by", "Label", "Final Net List"])
                                        
                                        for row in UpdatedNetReportList:
                                            columns = [col.strip() for col in row.split(",")]
                                            writer.writerow(columns)
                                            
                                    messagebox.showinfo("Success", f"Report saved to:\n{output_path1}")
                                except Exception as e:
                                    messagebox.showerror("Error", f"Could not write to CSV:\n{e}")                   
                                            
                        else:
                            messagebox.showerror("Error", "Error : Netlist is Empty")
                        #------------------------------------------------------------------
                else:
                    messagebox.showerror("Error", "Error : DRC List is Empty")
            else:
                messagebox.showwarning("No File Selected", "You did not select any CSV file.")

#------------------------------------------------------------------------------------------------------------------------------------------------------------------ 

        def ExcelFunc():
            UsedDt = "Application Menu - Added : Excel Menu"
            PurposeList.append(UsedDt)
            
            import os
            import shutil
            import getpass
            import time
            import win32com.client
            from tkinter import messagebox

            SOURCE_PATH = r"\\sng-psrvr06\Cadence_Skills\Menu-enc\MyCustomMenu.xlam"

            USERNAME = getpass.getuser()

            DEST_FOLDER = rf"C:\Users\{USERNAME}\AppData\Roaming\Microsoft\AddIns\Menu"
            DEST_PATH = os.path.join(DEST_FOLDER, os.path.basename(SOURCE_PATH))

            if not os.path.exists(SOURCE_PATH):
                raise FileNotFoundError(f"Source add-in not found: {SOURCE_PATH}")
            print(f"✅ Source file found: {SOURCE_PATH}")

            os.makedirs(DEST_FOLDER, exist_ok=True)
            print(f"✅ Destination folder ready: {DEST_FOLDER}")

            if os.path.exists(DEST_PATH):
                print(f"⚠️ Existing add-in found, replacing: {DEST_PATH}")
                os.remove(DEST_PATH)  
            shutil.copy2(SOURCE_PATH, DEST_PATH)
            print(f"✅ Copied new add-in to: {DEST_PATH}")

            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False 

            excel.Workbooks.Add()
            time.sleep(1)

            addin_name = os.path.basename(DEST_PATH).lower()
            existing_addin = None

            for a in excel.AddIns:
                if os.path.basename(a.FullName).lower() == addin_name:
                    existing_addin = a
                    break

            if existing_addin:
                try:
                    print(f"🔄 Disabling old version of add-in: {existing_addin.FullName}")
                    existing_addin.Installed = False
                    del existing_addin
                    time.sleep(0.5)
                except Exception as e:
                    print(f"⚠️ Warning: could not disable old version: {e}")


            try:
                print(f"➕ Registering new add-in: {DEST_PATH}")
                new_addin = excel.AddIns.Add(DEST_PATH)
                new_addin.Installed = True
                print(f"✅ '{new_addin.Name}' successfully installed and enabled in Excel!")

            except Exception as e:
                print(f"❌ Failed to add or enable add-in: {e}")
                print("➡️ Check if the add-in is in a Trusted Location or not blocked (Properties → Unblock).")

            finally:
                excel.Quit()

            messagebox.showinfo("Excel Update", "✅ EDA Excel Menu.\n -- Updated Successfully --")

#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def Apps4Func():
            UsedDt = "Application Menu - Added : APPS - 4 Menu"
            PurposeList.append(UsedDt)
            
            from pathlib import Path
            from datetime import datetime
            import shutil
            import getpass
            import os
            import tkinter as tk
            from tkinter import messagebox

            # --- Initialize Tkinter safely (to avoid 'no default root' errors) ---
            root = tk.Tk()
            root.withdraw()  # Hide the main window

            base_path = Path(r"C:\Cadence")

            insert_text1 = """POPUP "&APPS-4"
        BEGIN
            MENUITEM "apps_menu",           "apps_menu"
        END"""

            # --- Update Allegro menu definitions (both allegro.men & allegro_viewer_plus.men) ---
            for folder in base_path.iterdir():
                if folder.is_dir() and "SPB_" in folder.name:
                    link_val = folder / "share" / "pcb" / "text" / "cuimenus"

                    # Target both files
                    menu_files = [
                        link_val / "allegro.men",
                        link_val / "allegro_viewer_plus.men",
                    ]

                    for allegro_file in menu_files:
                        if not allegro_file.exists():
                            continue

                        try:
                            content = allegro_file.read_text(encoding="utf-8")
                        except Exception as e:
                            messagebox.showerror("Menu Update", f"Error reading {allegro_file}: {e}")
                            continue

                        if 'POPUP "&APPS-4"' in content:
                            continue  # Skip already updated files

                        # Backup original
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        backup_file = allegro_file.with_name(f"{allegro_file.stem}_Backup_{timestamp}{allegro_file.suffix}")
                        shutil.copy2(allegro_file, backup_file)

                        # Find insertion point
                        lines = content.splitlines()
                        insert_index = next((i for i in range(len(lines) - 1, -1, -1) if lines[i].strip() == "END"), len(lines))

                        updated_lines = lines[:insert_index] + insert_text1.splitlines() + ["END"]

                        # Write updated file
                        try:
                            allegro_file.write_text("\n".join(updated_lines) + "\n", encoding="utf-8")
                        except Exception as e:
                            messagebox.showerror("Menu Update", f"Error writing to {allegro_file}: {e}")

            # --- Source and destination paths ---
            source_path = Path(r"\\sng-psrvr06\Cadence_Skills\Menu-enc")

            allegro_env = os.environ.get("ALLEGRO_PCBENV")
            home_env = os.environ.get("HOME") or os.environ.get("USERPROFILE")

            # Ensure we have a valid destination path
            if allegro_env:
                dest_path = Path(allegro_env)
            elif home_env:
                dest_path = Path(home_env) / "pcbenv"
            else:
                messagebox.showerror("Menu Update", "Cannot determine user environment path.")
                return

            files_to_copy = ["allegro.ilinit", "env"]

            allowed_lines = [
                'load("//sng-psrvr06/Cadence_Skills/Menu-enc/apps_menu.il")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align.ile" "EncFileKar")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align_do.ile" "EncFileKar")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align_utils.ile" "EncFileKar")',
            ]

            # --- Copy and clean up files ---
            if not source_path.exists():
                messagebox.showerror("Menu Update", "Error: Source path not found. Contact Automation Team.")
                return

            dest_path.mkdir(parents=True, exist_ok=True)

            for file_name in files_to_copy:
                src_file = source_path / file_name
                dest_file = dest_path / file_name

                if src_file.exists():
                    # Backup existing file
                    if dest_file.exists():
                        backup_file = dest_file.with_suffix(dest_file.suffix + ".bak")
                        shutil.copy2(dest_file, backup_file)

                    # Copy new file
                    shutil.copy2(src_file, dest_file)

                    # Clean up allegro.ilinit after copy
                    if file_name.lower() == "allegro.ilinit":
                        try:
                            lines = dest_file.read_text(encoding="utf-8").splitlines()
                            filtered_lines = [line for line in lines if line.strip() in allowed_lines]
                            dest_file.write_text("\n".join(filtered_lines) + "\n", encoding="utf-8")
                        except Exception as e:
                            messagebox.showerror("Menu Update", f"Error updating allegro.ilinit: {e}")
                else:
                    messagebox.showerror("Menu Update", f"Missing file: {src_file}. Contact Automation Team.")

            messagebox.showinfo("Menu Update", "✅ Please restart Allegro.\nIf SKILL was not updated, contact Automation Team.")
            root.destroy()

#------------------------------------------------------------------------------------------------------------------------------------------------------------------       

        def Apps2Func():
            UsedDt = "Application Menu - Added : APPS - 2 Menu"
            PurposeList.append(UsedDt)
            
            from pathlib import Path
            from datetime import datetime
            import shutil
            import getpass
            import os
            import tkinter as tk
            from tkinter import messagebox

            # --- Initialize Tkinter safely (to avoid 'no default root' errors) ---
            root = tk.Tk()
            root.withdraw()  # Hide the main window

            base_path = Path(r"C:\Cadence")

            insert_text1 = """POPUP "&APPS-2"
        BEGIN
            MENUITEM "apps2_menu",           "apps2_menu"
        END"""

            # --- Update Allegro menu definitions (both allegro.men & allegro_viewer_plus.men) ---
            for folder in base_path.iterdir():
                if folder.is_dir() and "SPB_" in folder.name:
                    link_val = folder / "share" / "pcb" / "text" / "cuimenus"

                    # Target both files
                    menu_files = [
                        link_val / "allegro.men",
                        link_val / "allegro_viewer_plus.men",
                    ]

                    for allegro_file in menu_files:
                        if not allegro_file.exists():
                            continue

                        try:
                            content = allegro_file.read_text(encoding="utf-8")
                        except Exception as e:
                            messagebox.showerror("Menu Update", f"Error reading {allegro_file}: {e}")
                            continue

                        if 'POPUP "&APPS-2"' in content:
                            continue  # Skip already updated files

                        # Backup original
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        backup_file = allegro_file.with_name(f"{allegro_file.stem}_Backup_{timestamp}{allegro_file.suffix}")
                        shutil.copy2(allegro_file, backup_file)

                        # Find insertion point
                        lines = content.splitlines()
                        insert_index = next((i for i in range(len(lines) - 1, -1, -1) if lines[i].strip() == "END"), len(lines))

                        updated_lines = lines[:insert_index] + insert_text1.splitlines() + ["END"]

                        # Write updated file
                        try:
                            allegro_file.write_text("\n".join(updated_lines) + "\n", encoding="utf-8")
                        except Exception as e:
                            messagebox.showerror("Menu Update", f"Error writing to {allegro_file}: {e}")

            # --- Source and destination paths ---
            source_path = Path(r"\\sng-psrvr06\Cadence_Skills\Menu-enc")

            allegro_env = os.environ.get("ALLEGRO_PCBENV")
            home_env = os.environ.get("HOME") or os.environ.get("USERPROFILE")

            # Ensure we have a valid destination path
            if allegro_env:
                dest_path = Path(allegro_env)
            elif home_env:
                dest_path = Path(home_env) / "pcbenv"
            else:
                messagebox.showerror("Menu Update", "Cannot determine user environment path.")
                return

            files_to_copy = ["allegro.ilinit", "env"]

            allowed_lines = [
                'load("//sng-psrvr06/Cadence_Skills/Menu-enc/apps2_menu.il")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align.ile" "EncFileKar")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align_do.ile" "EncFileKar")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align_utils.ile" "EncFileKar")',
            ]

            # --- Copy and clean up files ---
            if not source_path.exists():
                messagebox.showerror("Menu Update", "Error: Source path not found. Contact Automation Team.")
                return

            dest_path.mkdir(parents=True, exist_ok=True)

            for file_name in files_to_copy:
                src_file = source_path / file_name
                dest_file = dest_path / file_name

                if src_file.exists():
                    # Backup existing file
                    if dest_file.exists():
                        backup_file = dest_file.with_suffix(dest_file.suffix + ".bak")
                        shutil.copy2(dest_file, backup_file)

                    # Copy new file
                    shutil.copy2(src_file, dest_file)

                    # Clean up allegro.ilinit after copy
                    if file_name.lower() == "allegro.ilinit":
                        try:
                            lines = dest_file.read_text(encoding="utf-8").splitlines()
                            filtered_lines = [line for line in lines if line.strip() in allowed_lines]
                            dest_file.write_text("\n".join(filtered_lines) + "\n", encoding="utf-8")
                        except Exception as e:
                            messagebox.showerror("Menu Update", f"Error updating allegro.ilinit: {e}")
                else:
                    messagebox.showerror("Menu Update", f"Missing file: {src_file}. Contact Automation Team.")

            messagebox.showinfo("Menu Update", "✅ Please restart Allegro.\nIf SKILL was not updated, contact Automation Team.")
            root.destroy()

#------------------------------------------------------------------------------------------------------------------------------------------------------------------     

        def MenuPass():
            from tkinter import messagebox
            messagebox.showinfo("Allegro Menu", "Please Choose Allegro Skill Menu According To Your Team..")

#------------------------------------------------------------------------------------------------------------------------------------------------------------------
        
        def TCLSch2Func():
            UsedDt = "Application Menu - Added : TCL - 2 Menu"
            PurposeList.append(UsedDt)
            
            from pathlib import Path
            import shutil
            import getpass
            from tkinter import messagebox

            base_path = Path(r"C:\Cadence")
            source_path = Path(r"\\sng-psrvr06\Cadence_Skills\Menu-enc")

            if not source_path.exists():
                messagebox.showerror("TCL Menu Update", "Error: Network path not found. Contact Automation Team..!")
                return

            TCL_file = "PacMenuAutoSCH.tcl"

            for folder in base_path.iterdir():
                if folder.is_dir() and "SPB_" in folder.name:
                    forms_path = folder / "tools" / "capture" / "tclscripts" / "capAutoLoad"

                    if forms_path.exists():
                        src_form = source_path / TCL_file
                        dest_form = forms_path / TCL_file
                        try:
                            if src_form.exists():
                                shutil.copy2(src_form, dest_form)
                        except Exception as e:
                            messagebox.showerror("Menu Update", f"Error copying {TCL_file}: {e}")
                                
            messagebox.showinfo("TCL Menu Update", "Please Restart the Tool and Check -- If TCL Menu was not updated Contact Automation Team..!")
            
#------------------------------------------------------------------------------------------------------------------------------------------------------------------
        
        def TCLSch4Func():
            UsedDt = "Application Menu - Added : TCL - 4 Menu"
            PurposeList.append(UsedDt)
            
            from pathlib import Path
            import shutil
            import getpass
            from tkinter import messagebox

            base_path = Path(r"C:\Cadence")
            source_path = Path(r"\\sng-psrvr06\Cadence_Skills\Menu-enc")

            if not source_path.exists():
                messagebox.showerror("TCL Menu Update", "Error: Network path not found. Contact Automation Team..!")
                return

            TCL_file = "PacMenuAuto.tcl"

            for folder in base_path.iterdir():
                if folder.is_dir() and "SPB_" in folder.name:
                    forms_path = folder / "tools" / "capture" / "tclscripts" / "capAutoLoad"

                    if forms_path.exists():
                        src_form = source_path / TCL_file
                        dest_form = forms_path / TCL_file
                        try:
                            if src_form.exists():
                                shutil.copy2(src_form, dest_form)
                        except Exception as e:
                            messagebox.showerror("Menu Update", f"Error copying {TCL_file}: {e}")
                                
            messagebox.showinfo("TCL Menu Update", "Please Restart the Tool and Check -- If TCL Menu was not updated Contact Automation Team..!")
            
#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def SiteChkByNLFunc():
            
            UsedDt = "Application Menu - Used : Site Check By Netlist"
            PurposeList.append(UsedDt)
            
            import tkinter as tk
            import numpy as np
            from tkinter import filedialog, messagebox, simpledialog
            import pandas as pd
            from collections import OrderedDict
            import re

            # --- Prompt user to choose DAT file ---
            messagebox.showinfo("Site Check By Netlist", "Please Select The Netlist File..!")
            def DAT_file():
                root = tk.Tk()
                root.withdraw()
                path = filedialog.askopenfilename(title="Select a DAT File", filetypes=[("DAT Files", "*.dat")])
                return path

            DATfile_path = DAT_file()
            if not DATfile_path:
                messagebox.showwarning("No File Selected", "You did not select any file.")
                return

            try:
                with open(DATfile_path, 'r') as f:
                    TempListRaw = [line.strip() for line in f.readlines()]
            except Exception as e:
                messagebox.showerror("Error", f"Error reading the file: {e}")
                return

            # ------------------ Ask runtime for patterns (converted to UPPER) -----------------------
            root = tk.Tk()
            root.withdraw()
            prompt = ("Enter site patterns separated by commas.\n"
                    "Use '*' where digits should appear. Examples:\n"
                    "  _S*  ,  S*_  ,  SITE*\n\n"
                    "Default: S*")
            user_input = simpledialog.askstring("Site Patterns", prompt)
            if not user_input:
                user_input = "S*"

            # Convert user input to uppercase so small letters become upper
            user_input = user_input.upper()

            raw_patterns = [p.strip() for p in user_input.split(",") if p.strip()]
            pattern_fragments = []
            for pat in raw_patterns:
                parts = pat.split("*")
                # parts already uppercased; escape fixed text parts
                escaped_parts = [re.escape(part) for part in parts]
                frag = r"\d+".join(escaped_parts)  # '*' -> \d+
                pattern_fragments.append(frag)

            # ------------------ PARSE DAT file into intermediate list ----------
            my_array = np.array(TempListRaw)
            LenVal = len(TempListRaw)
            FilteredList = []

            for i in range(LenVal):
                if my_array[i] == "NET_NAME":
                    j = i + 1
                    if j <= LenVal - 1:
                        PassNetName = my_array[i + 1]
                        # strip double quotes if present
                        if PassNetName.startswith('"') and PassNetName.endswith('"'):
                            PassNetName = PassNetName[1:-1]
                    else:
                        PassNetName = ""
                    FilteredList.append("NET_NAME--" + PassNetName)
                    while j <= LenVal - 1:
                        parsed_list = my_array[j].split()
                        if not parsed_list:
                            j += 1
                            continue
                        FirstEle = parsed_list[0]
                        if FirstEle == "NODE_NAME" and len(parsed_list) >= 3:
                            FilteredList.append("NODE_NAME--" + parsed_list[1] + "--" + parsed_list[2])
                        if my_array[j] == "NET_NAME":
                            break
                        j += 1

            NetListArr = np.array(FilteredList)
            NetListLen = len(FilteredList)

            # ------------------ Build TempList extracting SITE token using patterns -----
            TempList = []
            l = 0

            # Precompile prefix/suffix regexes for each fragment (no (?i) needed since we uppercase everything)
            compiled_pairs = []
            for frag in pattern_fragments:
                re_pref = re.compile(rf'^{frag}')   # prefix (start)
                re_suf = re.compile(rf'{frag}$')   # suffix (end)
                compiled_pairs.append((re_pref, re_suf))

            # Token extractor for alpha+digits (uppercase)
            token_extract_re = re.compile(r'([A-Z]+[0-9]+)')

            while l <= NetListLen - 1:
                parts = NetListArr[l].split("--")
                if not parts:
                    l += 1
                    continue
                tag = parts[0]
                if tag == "NET_NAME":
                    NetNameTemp = parts[1]
                else:
                    RefValTemp = parts[1] if len(parts) > 1 else ""
                    # uppercase the ref value for matching
                    RefValTemp_up = RefValTemp.upper()
                    matched_site = None
                    for (re_pref, re_suf) in compiled_pairs:
                        m = re_pref.search(RefValTemp_up) or re_suf.search(RefValTemp_up)
                        if m:
                            sub = m.group(0)
                            tok = token_extract_re.search(sub)
                            if tok:
                                matched_site = tok.group(1).upper()
                                break
                    if matched_site:
                        TempList.append(f"{NetNameTemp},{matched_site}")
                l += 1

            # remove duplicates while preserving order
            TempList = list(dict.fromkeys(TempList))

            if not TempList:
                messagebox.showwarning("No matches", "No components matched the provided site patterns.")
                return

            # ------------------ GROUP INTO NET -> SITES and dedupe per net ----------
            groups = OrderedDict()
            for item in TempList:
                if "," not in item:
                    continue
                net, site = item.split(",", 1)
                net = net.strip()
                site = site.strip()
                groups.setdefault(net, []).append(site)

            for net in groups:
                groups[net] = list(dict.fromkeys(groups[net]))  # unique sites preserving order

            # Build rows: NETNAME | SITE_VALUES(comma separated) | COUNT
            rows = []
            all_sites = []  # collect all site tokens for range analysis
            for net, sites in groups.items():
                combined = ",".join(sites)
                count = len(sites)
                rows.append([net, combined, count])
                all_sites.extend(sites)

            df = pd.DataFrame(rows, columns=["NETNAME", "SITE_VALUES", "COUNT"])

            # ------------------ Compute per-prefix min..max ranges --------------------
            # Map: prefix -> list of integers
            prefix_map = {}
            site_token_re = re.compile(r'^([A-Z]+)(\d+)$')
            for s in all_sites:
                m = site_token_re.match(s)
                if not m:
                    continue
                pref = m.group(1)
                num = int(m.group(2))
                prefix_map.setdefault(pref, []).append(num)

            if not prefix_map:
                # fallback: show highest count if tokens couldn't be parsed numerically
                highest_count = df["COUNT"].max() if not df.empty else 0
                confirm_msg = (f"Could not determine numeric site ranges from tokens.\n"
                            f"Highest site count across all nets is: {highest_count}\n\n"
                            "Click OK to generate the report.\n"
                            "Click Cancel to abort and contact the automation team.")
            else:
                # Build range strings like "S: S0 to S55; SITE: SITE1 to SITE12"
                range_parts = []
                for pref, nums in prefix_map.items():
                    mn = min(nums)
                    mx = max(nums)
                    range_parts.append(f"{pref}: {pref}{mn} to {pref}{mx}")
                range_text = "; ".join(range_parts)
                confirm_msg = (f"Detected site ranges:\n{range_text}\n\n")

            # ------------------ CONFIRMATION POPUP ------------------------------------
            root = tk.Tk()
            root.withdraw()
            proceed = messagebox.askokcancel("Confirm Report Generation", confirm_msg)

            if not proceed:
                messagebox.showinfo("Aborted", "Contact automation team")
                return

            # ------------------ Ask user where to save Excel and write file -------------
            save_path = filedialog.asksaveasfilename(
                title="Save Excel File As",
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")],
                initialfile="Netlist_Sites_Output.xlsx"
            )

            if save_path:
                try:
                    df.to_excel(save_path, index=False)
                    messagebox.showinfo("Success", f"Excel saved to:\n{save_path}")
                except Exception as e:
                    messagebox.showerror("Save Error", f"Could not save Excel file: {e}")
            else:
                messagebox.showwarning("Cancelled", "Save operation cancelled.")

            print("Done. Nets found:", len(groups))

#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def SocSiteChkByNLFunc():
            
            UsedDt = "Application Menu - Used : Socket Site Check By Netlist"
            PurposeList.append(UsedDt)
            
            import tkinter as tk
            import numpy as np
            from tkinter import filedialog, messagebox, simpledialog
            import pandas as pd
            from collections import OrderedDict
            import re

            # --- Prompt user to choose DAT file ---
            messagebox.showinfo("Socket Site Check By Netlist", "Please Select The Netlist File..!")
            def DAT_file():
                root = tk.Tk()
                root.withdraw()
                path = filedialog.askopenfilename(title="Select a DAT File", filetypes=[("DAT Files", "*.dat")])
                return path

            DATfile_path = DAT_file()
            if not DATfile_path:
                messagebox.showwarning("No File Selected", "You did not select any file.")
                return

            try:
                with open(DATfile_path, 'r') as f:
                    TempListRaw = [line.strip() for line in f.readlines()]
            except Exception as e:
                messagebox.showerror("Error", f"Error reading the file: {e}")
                return

            # ------------------ Ask runtime for patterns (converted to UPPER) -----------------------
            root = tk.Tk()
            root.withdraw()
            prompt = ("Enter site patterns separated by commas.\n"
                    "Use '*' where digits should appear. Examples:\n"
                    "  _S*  ,  S*_  ,  SITE*\n\n"
                    "Default: S*")
            user_input = simpledialog.askstring("Site Patterns", prompt)
            if not user_input:
                user_input = "S*"

            # Convert user input to uppercase so small letters become upper
            user_input = user_input.upper()

            raw_patterns = [p.strip() for p in user_input.split(",") if p.strip()]
            pattern_fragments = []
            for pat in raw_patterns:
                parts = pat.split("*")
                # parts already uppercased; escape fixed text parts
                escaped_parts = [re.escape(part) for part in parts]
                frag = r"\d+".join(escaped_parts)  # '*' -> \d+
                pattern_fragments.append(frag)

            # ------------------ PARSE DAT file into intermediate list ----------
            my_array = np.array(TempListRaw)
            LenVal = len(TempListRaw)
            FilteredList = []

            for i in range(LenVal):
                if my_array[i] == "NET_NAME":
                    j = i + 1
                    if j <= LenVal - 1:
                        PassNetName = my_array[i + 1]
                        # strip double quotes if present
                        if PassNetName.startswith('"') and PassNetName.endswith('"'):
                            PassNetName = PassNetName[1:-1]
                    else:
                        PassNetName = ""
                    FilteredList.append("NET_NAME--" + PassNetName)
                    while j <= LenVal - 1:
                        parsed_list = my_array[j].split()
                        if not parsed_list:
                            j += 1
                            continue
                        FirstEle = parsed_list[0]
                        if FirstEle == "NODE_NAME" and len(parsed_list) >= 3:
                            FilteredList.append("NODE_NAME--" + parsed_list[1] + "--" + parsed_list[2])
                        if my_array[j] == "NET_NAME":
                            break
                        j += 1

            NetListArr = np.array(FilteredList)
            NetListLen = len(FilteredList)

            # ------------------ Build TempList extracting SITE token using patterns -----
            TempList = []
            l = 0

            # Precompile prefix/suffix regexes for each fragment (no (?i) needed since we uppercase everything)
            compiled_pairs = []
            for frag in pattern_fragments:
                re_pref = re.compile(rf'^{frag}')   # prefix (start)
                re_suf = re.compile(rf'{frag}$')   # suffix (end)
                compiled_pairs.append((re_pref, re_suf))

            # Token extractor for alpha+digits (uppercase)
            token_extract_re = re.compile(r'([A-Z]+[0-9]+)')

            while l <= NetListLen - 1:
                parts = NetListArr[l].split("--")
                if not parts:
                    l += 1
                    continue
                tag = parts[0]
                if tag == "NET_NAME":
                    NetNameTemp = parts[1]
                else:
                    RefValTemp = parts[2] if len(parts) > 1 else ""
                    # uppercase the ref value for matching
                    RefValTemp_up = RefValTemp.upper()
                    matched_site = None
                    for (re_pref, re_suf) in compiled_pairs:
                        m = re_pref.search(RefValTemp_up) or re_suf.search(RefValTemp_up)
                        if m:
                            sub = m.group(0)
                            tok = token_extract_re.search(sub)
                            if tok:
                                matched_site = tok.group(1).upper()
                                break
                    if matched_site:
                        TempList.append(f"{NetNameTemp},{matched_site}")
                l += 1

            # remove duplicates while preserving order
            TempList = list(dict.fromkeys(TempList))

            if not TempList:
                messagebox.showwarning("No matches", "No components matched the provided site patterns.")
                return

            # ------------------ GROUP INTO NET -> SITES and dedupe per net ----------
            groups = OrderedDict()
            for item in TempList:
                if "," not in item:
                    continue
                net, site = item.split(",", 1)
                net = net.strip()
                site = site.strip()
                groups.setdefault(net, []).append(site)

            for net in groups:
                groups[net] = list(dict.fromkeys(groups[net]))  # unique sites preserving order

            # Build rows: NETNAME | SITE_VALUES(comma separated) | COUNT
            rows = []
            all_sites = []  # collect all site tokens for range analysis
            for net, sites in groups.items():
                combined = ",".join(sites)
                count = len(sites)
                rows.append([net, combined, count])
                all_sites.extend(sites)

            df = pd.DataFrame(rows, columns=["NETNAME", "SITE_VALUES", "COUNT"])

            # ------------------ Compute per-prefix min..max ranges --------------------
            # Map: prefix -> list of integers
            prefix_map = {}
            site_token_re = re.compile(r'^([A-Z]+)(\d+)$')
            for s in all_sites:
                m = site_token_re.match(s)
                if not m:
                    continue
                pref = m.group(1)
                num = int(m.group(2))
                prefix_map.setdefault(pref, []).append(num)

            if not prefix_map:
                # fallback: show highest count if tokens couldn't be parsed numerically
                highest_count = df["COUNT"].max() if not df.empty else 0
                confirm_msg = (f"Could not determine numeric site ranges from tokens.\n"
                            f"Highest site count across all nets is: {highest_count}\n\n"
                            "Click OK to generate the report.\n"
                            "Click Cancel to abort and contact the automation team.")
            else:
                # Build range strings like "S: S0 to S55; SITE: SITE1 to SITE12"
                range_parts = []
                for pref, nums in prefix_map.items():
                    mn = min(nums)
                    mx = max(nums)
                    range_parts.append(f"{pref}: {pref}{mn} to {pref}{mx}")
                range_text = "; ".join(range_parts)
                confirm_msg = (f"Detected site ranges:\n{range_text}\n\n")

            # ------------------ CONFIRMATION POPUP ------------------------------------
            root = tk.Tk()
            root.withdraw()
            proceed = messagebox.askokcancel("Confirm Report Generation", confirm_msg)

            if not proceed:
                messagebox.showinfo("Aborted", "Contact automation team")
                return

            # ------------------ Ask user where to save Excel and write file -------------
            save_path = filedialog.asksaveasfilename(
                title="Save Excel File As",
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")],
                initialfile="Netlist_Sites_Output.xlsx"
            )

            if save_path:
                try:
                    df.to_excel(save_path, index=False)
                    messagebox.showinfo("Success", f"Excel saved to:\n{save_path}")
                except Exception as e:
                    messagebox.showerror("Save Error", f"Could not save Excel file: {e}")
            else:
                messagebox.showwarning("Cancelled", "Save operation cancelled.")

            print("Done. Nets found:", len(groups))

#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def AutoConRepFunc():
            UsedDt = "Schematic Menu - Used : Auto Connection Report By Netlist"
            PurposeList.append(UsedDt)

            for widget in window.winfo_children():
                if isinstance(widget, tk.Label) and hasattr(widget, 'image'):
                    continue
                widget.destroy()

            def ConnectionReport():           
                PwrNetNames = entry1.get()
                PwrNetNamesUp = PwrNetNames.upper()
                if ',' in PwrNetNamesUp:
                    PwrNetNameList = PwrNetNamesUp.split(",")
                else:
                    PwrNetNameList = []
                    PwrNetNameList.append(PwrNetNamesUp)
            
                
                PwrConStartLetter = entry2.get()
                PwrConStartLetterUp = PwrConStartLetter.upper()
                PwrConStartLetterLen = len(PwrConStartLetter)
                
                
                StartPinNo = entry6.get()
                EndPinNo = entry7.get()
                
                if PwrNetNames != "" and PwrConStartLetter != "" and StartPinNo != "" and EndPinNo != "":
                    from tkinter import filedialog
                    import numpy as np
                    import csv
                    from tkinter.filedialog import asksaveasfilename
                    from openpyxl import Workbook
                    from openpyxl.styles import PatternFill
                    from datetime import datetime
                    import os
                    import sys
                    import subprocess
                    
                    def open_file():
                        messagebox.showinfo("Connection", "Please Select The Netlist File..!")
                        file_path = filedialog.askopenfilename(title="Select a DAT File", filetypes=[("DAT Files", "*.dat")])
                        folder_path = os.path.dirname(file_path)
                        if file_path:  # If a file was selected
                            try:
                                with open(file_path, 'r') as file:
                                    content_list = file.readlines()  
                                    content_list = [line.strip() for line in content_list]
                                    
                                    my_array = np.array(content_list)
                                    LenVal = len(content_list)
                                    FilteredListInp = []
                                    FilteredListChk = []
                                    for i in range(LenVal):
                                        if my_array[i] == "NET_NAME":
                                            j = i + 1
                                            PassNetName = my_array[i + 1][1:-1]
                                            PassNetNameUp = PassNetName.upper()
                                            for PwrNetNameListlp in PwrNetNameList:
                                                if PassNetNameUp == PwrNetNameListlp:
                                                    StrNetName = "NET_NAME" + "--" + PassNetName
                                                    # FilteredListInp.append(StrNetName)
                                                    while j <= LenVal - 1:
                                                        parsed_list = my_array[j].split()
                                                        FirstEle = parsed_list[0]
                                                        if FirstEle == "NODE_NAME":
                                                            StrPassVal = StrNetName + "--" + "NODE_NAME" + "--" + parsed_list[1] + "--" + parsed_list[2] 
                                                            first_letter = parsed_list[1][0]
                                                            first_letterUp = first_letter.upper()
                                                            if first_letterUp == PwrConStartLetterUp:
                                                                FilteredListInp.append(StrPassVal)
                                                        if my_array[j] == "NET_NAME":
                                                            break
                                                        j += 1
                                                else:
                                                    if "+" not in PassNetName:
                                                        StrNetName = "NET_NAME" + "--" + PassNetName
                                                        # FilteredListChk.append(StrNetName)
                                                        while j <= LenVal - 1:
                                                            parsed_list = my_array[j].split()
                                                            FirstEle = parsed_list[0]
                                                            if FirstEle == "NODE_NAME":
                                                                StrPassVal = StrNetName + "--" + "NODE_NAME" + "--" + parsed_list[1] + "--" + parsed_list[2] 
                                                                if parsed_list[2] == StartPinNo or parsed_list[2] == EndPinNo:
                                                                    FilteredListChk.append(StrPassVal)
                                                            if my_array[j] == "NET_NAME":
                                                                break
                                                            j += 1                                                                            
                                    
                                    FinalList = []
                                    for FilteredListInplp in FilteredListInp:
                                        OrderList = []
                                        FiltInpPass = FilteredListInplp.split("--")
                                        Filt_Chk_Inp = FiltInpPass[3][PwrConStartLetterLen:]
                                        OrderList.append(FiltInpPass[1])
                                        OrderList.append(FiltInpPass[3])
                                        for FilteredListChklp in FilteredListChk:
                                            FiltGetPass = FilteredListChklp.split("--")
                                            Filt_Chk_Get = FiltGetPass[3]
                                            if Filt_Chk_Inp in Filt_Chk_Get:
                                                PassData = Filt_Chk_Get + "--" + FiltGetPass[4] + "--" + FiltGetPass[1]
                                                OrderList.append(PassData)
                                        
                                        #Unq_OrderList = list(set(OrderList))
                                        FinalList.append(OrderList)
                                    
                                    
                                    def reorder_connections(connections):
                                        if len(connections) < 2:
                                            return connections

                                        result = connections[:2]
                                        
                                        remaining = connections[2:]
                                        
                                        current_refdes = connections[1]
                                        
                                        find_by_refdes = True
                                        
                                        while remaining:
                                            if find_by_refdes:
                                                found = False
                                                for i, element in enumerate(remaining):
                                                    if '--' in element and element.split('--')[0] == current_refdes:
                                                        result.append(element)
                                                        remaining.pop(i)
                                                        current_netname = element.split('--')[2] if len(element.split('--')) > 2 else ''
                                                        found = True
                                                        break
                                                
                                                if not found:
                                                    result[-1] = result[-1] + '-- Connection Missing'
                                                    result.extend(remaining)
                                                    break
                                            else:
                                                found = False
                                                for i, element in enumerate(remaining):
                                                    if '--' in element and len(element.split('--')) > 2 and element.split('--')[2] == current_netname:
                                                        result.append(element)
                                                        remaining.pop(i)
                                                        current_refdes = element.split('--')[0]
                                                        found = True
                                                        break
                                                
                                                if not found:
                                                    result[-1] = result[-1] + '[ Connection Missing !]'
                                                    result.extend(remaining)
                                                    break
                                            
                                            find_by_refdes = not find_by_refdes
                                        
                                        return result

                                    def save_connection_report(sheet1_data, sheet2_data):
    
                                        timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
                                        filename = f"Connection Report {timestamp}.xlsx"

                                        wb = Workbook()

                                        yellow_fill = PatternFill(
                                            start_color="FFFF00",
                                            end_color="FFFF00",
                                            fill_type="solid"
                                        )

                                        # -------------------------------------------------
                                        # Find rows with error FROM SHEET 2 ONLY
                                        # -------------------------------------------------
                                        error_rows = set()

                                        for idx, row in enumerate(sheet2_data, start=1):
                                            if any("Connection Missing !" in str(x) for x in row):
                                                error_rows.add(idx)

                                        # -------------------------------------------------
                                        # Helper to write sheet
                                        # -------------------------------------------------
                                        def write_sheet(ws, data):

                                            for row_idx, row in enumerate(data, start=1):

                                                for col_idx, value in enumerate(row, start=1):
                                                    ws.cell(row=row_idx, column=col_idx, value=value)

                                                # highlight if row index exists in error list
                                                if row_idx in error_rows:
                                                    for col_idx in range(1, len(row) + 1):
                                                        ws.cell(row=row_idx, column=col_idx).fill = yellow_fill

                                            # auto column width
                                            for col in ws.columns:
                                                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                                                ws.column_dimensions[col[0].column_letter].width = max_len + 2

                                        # Sheet 1
                                        ws1 = wb.active
                                        ws1.title = "Connection Report"
                                        write_sheet(ws1, sheet1_data)

                                        # Sheet 2
                                        ws2 = wb.create_sheet("Detailed Connection Report")
                                        write_sheet(ws2, sheet2_data)

                                        save_path = os.path.join(folder_path, filename)
                                        
                                        wb.save(save_path)

                                        # auto open
                                        if sys.platform.startswith("win"):
                                            os.startfile(save_path)
                                        elif sys.platform.startswith("darwin"):
                                            subprocess.call(["open", save_path])
                                        else:
                                            subprocess.call(["xdg-open", save_path])
        
        
                                    Sheet2OutputList = []   
                                    Sheet1OutputList = []                           
                                    for FinalListlp in FinalList:
                                        output = reorder_connections(FinalListlp)
                                        Sheet2OutputList.append(output)
                                        
                                        PassList = []
                                        PassList.append(output[0])
                                        PassList.append(output[1])
                                        OutputListLen = len(output)
                                        for Passlp in range(2, OutputListLen - 1):
                                            PassDtPrse = output[Passlp].split("--")
                                            PassList.append(PassDtPrse[0])
                                        
                                        PassListLast = output[-1]
                                        PassListLastPrse = PassListLast.split("--")
                                        PassList.append(PassListLastPrse[-1])
                                        PassList = list(dict.fromkeys(PassList))
                                        Sheet1OutputList.append(PassList)
                                     
                                    save_connection_report(Sheet1OutputList, Sheet2OutputList)
                                    
                            except Exception as e:
                                print(f"Error opening the file: {e}")
                    NetL = tk.Tk()
                    NetL.withdraw()  
                    open_file()
                else:
                    messagebox.showerror("Error", f"Please Check.. The Required Input Field is Empty..")
                

            def create_button(text, relx, rely, command):
                btn = tk.Label(window, text=text, font=("Trebuchet MS", 20),
                            fg="#F6FAFC", bg="#020638", bd=0)
                btn.bind("<Enter>", lambda e: btn.config(fg="#1F5186"))
                btn.bind("<Leave>", lambda e: btn.config(fg="#F6FAFC"))
                btn.bind("<Button-1>", lambda e: command())
                btn.place(relx=relx, rely=rely, anchor="center")

            def create_label_entry_lr(text, rely):
                tk.Label(window, text=text, font=("Didot", 14), bg="#f0f0f0").place(relx=0.40, rely=rely, anchor="center")
                e = tk.Entry(window, font=("Didot", 13), width=25)
                e.place(relx=0.65, rely=rely, anchor="center")
                return e

            def create_label_entry_range(text, rely):
                tk.Label(window, text=text, font=("Didot", 14), bg="#f0f0f0").place(relx=0.40, rely=rely, anchor="center")
                e1 = tk.Entry(window, font=("Didot", 8), width=10)
                e1.place(relx=0.60, rely=rely, anchor="center")
                tk.Label(window, text="And", font=("Didot", 13), bg="#f0f0f0").place(relx=0.65, rely=rely, anchor="center")
                e2 = tk.Entry(window, font=("Didot", 8), width=10)
                e2.place(relx=0.70, rely=rely, anchor="center")
                return e1, e2

            form_y = 0.25
            sp = 0.10

            entry1 = create_label_entry_lr("[ POWER NET NAMES (Ex : +12V,+15V) ]", form_y)
            entry2 = create_label_entry_lr("[ POWER CON START LETTER ]", form_y + sp)
            entry6, entry7 = create_label_entry_range("[ COMMON PIN NUMBER ]", form_y + sp * 2)

            create_button("[ CLICK HERE ]", 0.5, form_y + sp * 4, ConnectionReport)
            create_button("[<-- BACK ]", 0.1, 0.05, SchFunc)
            
#------------------------------------------------------------------------------------------------------------------------------------------------------------------  

        def AutoConRepCBitFunc():
            UsedDt = "Schematic Menu - Used : Auto Connection Report By Netlist (CBIT)"
            PurposeList.append(UsedDt)
            
            for widget in window.winfo_children():
                if isinstance(widget, tk.Label) and hasattr(widget, 'image'):
                    continue
                widget.destroy()
            
            
            def ConnectionRepByCBIT():
                
                from tkinter import filedialog
                import numpy as np
                import csv
                from tkinter.filedialog import asksaveasfilename
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill
                from datetime import datetime
                import os
                import sys
                import subprocess
                
                InpNetNameUI = entry1.get()
                InpNetNameUIUp = InpNetNameUI.upper()
                
                StartPinNo = entry6.get()
                EndPinNo = entry7.get()
                
                if InpNetNameUI != "" and StartPinNo != "" and EndPinNo != "":
                
                    def open_file():
                        messagebox.showinfo("Connection", "Please Select The Netlist File..!")
                        file_path = filedialog.askopenfilename(title="Select a DAT File", filetypes=[("DAT Files", "*.dat")])
                        if file_path:  # If a file was selected
                            try:
                                with open(file_path, 'r') as file:
                                    content_list = file.readlines()  
                                    content_list = [line.strip() for line in content_list]
                                    
                                    my_array = np.array(content_list)
                                    LenVal = len(content_list)
                                    NetListData = []
                                    RefdesList = []
                                    for i in range(LenVal):
                                        if my_array[i] == "NET_NAME":
                                            j = i + 1
                                            PassNetName = my_array[i + 1][1:-1]
                                            PassNetNameUp = PassNetName.upper()
                                            StrNetName = "NET_NAME" + "--" + PassNetName
                                            # FilteredListInp.append(StrNetName)
                                            SubRefList = []
                                            while j <= LenVal - 1:
                                                parsed_list = my_array[j].split()
                                                FirstEle = parsed_list[0]
                                                if FirstEle == "NODE_NAME":
                                                    StrPassVal = StrNetName + "--" + "NODE_NAME" + "--" + parsed_list[1] + "--" + parsed_list[2] 
                                                    NetListData.append(StrPassVal)
                                                    if parsed_list[2] == StartPinNo or parsed_list[2] == EndPinNo:
                                                        if "+" not in PassNetName :
                                                            SubRefList.append(parsed_list[1])
                                                if my_array[j] == "NET_NAME":
                                                    break
                                                j += 1
                                            
                                            if SubRefList != []:
                                                SubRefList = list(set(SubRefList))
                                                RefdesList.append(SubRefList)
                                    
                                    CBitNetNameList = []
                                    for NetListDatalp in NetListData:
                                        InpSplitList = NetListDatalp.split("--")
                                        InpNetName = InpSplitList[1]
                                        InpNetNameUp = InpNetName.upper()
                                        if InpNetNameUIUp in InpNetNameUp:
                                            CBitNetNameList.append(InpNetNameUp)
                                    
                                    CBitNetNameList = sorted(set(CBitNetNameList))
                                    
                                    
                                    ResultList = []
                                    for CBitNetNameListlp in CBitNetNameList:
                                        SubList = []
                                        HeadPass = CBitNetNameListlp + " Connections :"
                                        SubList.append(HeadPass)
                                        for NetListDataPassLp in NetListData:
                                            InpSplitPassList = NetListDataPassLp.split("--")
                                            InpNetNamePass = InpSplitPassList[1]
                                            if CBitNetNameListlp == InpNetNamePass:
                                                SubList.append(InpSplitPassList[3])
                                        
                                        SubListLen = len(SubList)

                                        SubInc = 0
                                        while SubInc <= SubListLen - 1:
                                            ChkVal = SubList[SubInc] 
                                            for RefdesListlp in RefdesList:
                                                if ChkVal in RefdesListlp:
                                                    for addlp in RefdesListlp:
                                                        if addlp not in SubList:
                                                            SubList.append(addlp)
                                                    SubListLen = len(SubList)
                                            SubInc +=1 
                                        
                                        ResultList.append(SubList)
                                        
                                    folder_path = os.path.dirname(file_path)

                                    timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
                                    filename = f"Connection Report {timestamp}.xlsx"
                                    save_path = os.path.join(folder_path, filename)

                                    wb = Workbook()
                                    ws = wb.active
                                    ws.title = "Detailed Connection Report"
                                    
                                    for col_index, col_data in enumerate(ResultList, start=1):
                                        for row_index, value in enumerate(col_data, start=1):
                                            ws.cell(row=row_index, column=col_index, value=value)

                                    wb.save(save_path)
                                    os.startfile(save_path)
                                        
                            except Exception as e:
                                print(f"Error opening the file: {e}")
                    NetL = tk.Tk()
                    NetL.withdraw()  
                    open_file()
                else:
                    messagebox.showerror("Error", f"Please Check.. The Required Input Field is Empty..")
                
            
            def create_button(text, relx, rely, command):
                btn = tk.Label(window, text=text, font=("Trebuchet MS", 20),
                            fg="#F6FAFC", bg="#020638", bd=0)
                btn.bind("<Enter>", lambda e: btn.config(fg="#1F5186"))
                btn.bind("<Leave>", lambda e: btn.config(fg="#F6FAFC"))
                btn.bind("<Button-1>", lambda e: command())
                btn.place(relx=relx, rely=rely, anchor="center")

            def create_label_entry_lr(text, rely):
                tk.Label(window, text=text, font=("Didot", 14), bg="#f0f0f0").place(relx=0.40, rely=rely, anchor="center")
                e = tk.Entry(window, font=("Didot", 13), width=25)
                e.place(relx=0.65, rely=rely, anchor="center")
                return e

            def create_label_entry_range(text, rely):
                tk.Label(window, text=text, font=("Didot", 14), bg="#f0f0f0").place(relx=0.40, rely=rely, anchor="center")
                e1 = tk.Entry(window, font=("Didot", 8), width=10)
                e1.place(relx=0.60, rely=rely, anchor="center")
                tk.Label(window, text="And", font=("Didot", 13), bg="#f0f0f0").place(relx=0.65, rely=rely, anchor="center")
                e2 = tk.Entry(window, font=("Didot", 8), width=10)
                e2.place(relx=0.70, rely=rely, anchor="center")
                return e1, e2

            form_y = 0.25
            sp = 0.10

            entry1 = create_label_entry_lr("[ INPUT NET NAME (Ex : CBIT) ]", form_y)
            entry6, entry7 = create_label_entry_range("[ COMMON PIN NUMBER ]", form_y + sp)

            create_button("[ CLICK HERE ]", 0.5, form_y + sp * 3, ConnectionRepByCBIT)
            create_button("[<-- BACK ]", 0.1, 0.05, SchFunc)

#------------------------------------------------------------------------------------------------------------------------------------------------------------------ 

        def AutoSiteToSiteDiffFunc():
            UsedDt = "Schematic Menu - Used : Auto Site to Site Short Different Dut"
            PurposeList.append(UsedDt)
            
            for widget in window.winfo_children():
                if isinstance(widget, tk.Label) and hasattr(widget, 'image'):
                    continue
                widget.destroy()
                
                
            def SiteToSiteShortDiff():
                
                from tkinter import filedialog
                import numpy as np
                import csv
                from tkinter.filedialog import asksaveasfilename
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill
                from datetime import datetime
                import os
                import sys
                import subprocess
                
                InpDutName = entry1.get()
                InpDutNameUp = InpDutName.upper()
                
                if InpDutNameUp != "":
                    def open_file():
                        messagebox.showinfo("Connection", "Please Select The Netlist File..!")
                        file_path = filedialog.askopenfilename(title="Select a DAT File", filetypes=[("DAT Files", "*.dat")])
                        if file_path:  # If a file was selected
                            try:
                                with open(file_path, 'r') as file:
                                    content_list = file.readlines()  
                                    content_list = [line.strip() for line in content_list]
                                    
                                    my_array = np.array(content_list)
                                    LenVal = len(content_list)

                                    RefdesList = []
                                    for i in range(LenVal):
                                        if my_array[i] == "NET_NAME":
                                            j = i + 1
                                            PassNetName = my_array[i + 1][1:-1]
                                            PassNetNameUp = PassNetName.upper()
                                            StrNetName = "NET_NAME" + "--" + PassNetName
                                            SubRefList = []
                                            SubRefList.append(PassNetName)
                                            while j <= LenVal - 1:
                                                parsed_list = my_array[j].split()
                                                FirstEle = parsed_list[0]
                                                if FirstEle == "NODE_NAME":
                                                    StrPassVal = StrNetName + "--" + "NODE_NAME" + "--" + parsed_list[1] + "--" + parsed_list[2] 
                                                    SubRefList.append(parsed_list[1])
                                                if my_array[j] == "NET_NAME":
                                                    break
                                                j += 1
                                            
                                            if SubRefList != []:
                                                SubRefList = list(dict.fromkeys(SubRefList))
                                                RefdesList.append(SubRefList)
                                    
                                    ErrorList = [] 
                                    for RefdesListlp in RefdesList:
                                        ErrorSubList = []
                                        site_list = [item for item in RefdesListlp if InpDutNameUp in item]
                                        if len(site_list) >= 2:
                                            ErrorNetName = RefdesListlp[0]
                                            ErrorSubList.append(ErrorNetName)
                                            for site_listlp in site_list:
                                                ErrorSubList.append(site_listlp)

                                        if ErrorSubList != []:
                                            ErrorList.append(ErrorSubList)
                                    
                                    
                                    folder_path = os.path.dirname(file_path)

                                    # Create filename
                                    timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
                                    filename = f"Site to Site Short Diff Dut {timestamp}.xlsx"
                                    save_path = os.path.join(folder_path, filename)

                                    # Create workbook
                                    wb = Workbook()
                                    ws = wb.active
                                    ws.title = "Error Site to Site Short Diff Dut"

                                    # Write row-wise
                                    for row in ErrorList:
                                        ws.append(row)

                                    # Save file
                                    wb.save(save_path)

                                    # ✅ Automatically open the file (Windows)
                                    os.startfile(save_path)


                                            
                            except Exception as e:
                                print(f"Error opening the file: {e}")
                    NetL = tk.Tk()
                    NetL.withdraw()  
                    open_file()
                else:
                    messagebox.showerror("Error", f"Please Check.. The Required Input Field is Empty..")
                
            def create_button(text, relx, rely, command):
                btn = tk.Label(window, text=text, font=("Trebuchet MS", 20),
                            fg="#F6FAFC", bg="#020638", bd=0)
                btn.bind("<Enter>", lambda e: btn.config(fg="#1F5186"))
                btn.bind("<Leave>", lambda e: btn.config(fg="#F6FAFC"))
                btn.bind("<Button-1>", lambda e: command())
                btn.place(relx=relx, rely=rely, anchor="center")

            def create_label_entry_lr(text, rely):
                tk.Label(window, text=text, font=("Didot", 14), bg="#f0f0f0").place(relx=0.40, rely=rely, anchor="center")
                e = tk.Entry(window, font=("Didot", 13), width=25)
                e.place(relx=0.65, rely=rely, anchor="center")
                return e

            form_y = 0.25
            sp = 0.10

            entry1 = create_label_entry_lr("[ DUT COMMON NAME : ]", form_y)

            create_button("[ CLICK HERE ]", 0.5, form_y + sp * 2, SiteToSiteShortDiff)
            create_button("[<-- BACK ]", 0.1, 0.05, SchFunc)

#------------------------------------------------------------------------------------------------------------------------------------------------------------------ 

        def AutoSiteToSiteSameFunc():
            
            UsedDt = "Schematic Menu - Used : Auto Site Short Same Dut"
            PurposeList.append(UsedDt)
            
            for widget in window.winfo_children():
                if isinstance(widget, tk.Label) and hasattr(widget, 'image'):
                    continue
                widget.destroy()
                
                
            def SiteToSiteShortSame():
    
                import numpy as np
                import tkinter as tk
                from tkinter import filedialog, messagebox
                from openpyxl import load_workbook
                from collections import OrderedDict
                from openpyxl import Workbook

                InpDutName = entry1.get()
                InpDutNameUp = InpDutName.upper()
                
                if InpDutNameUp != "":
                    
                    def open_file():

                        RefdesList = []
                        ExcelList = []

                        # ================= DAT FILE =================
                        messagebox.showinfo("Connection", "Please Select The Netlist DAT File..!")
                        dat_path = filedialog.askopenfilename(
                            title="Select DAT File",
                            filetypes=[("DAT Files", "*.dat")]
                        )

                        if not dat_path:
                            return

                        try:
                            with open(dat_path, 'r') as file:
                                content_list = [line.strip() for line in file.readlines()]
                                my_array = np.array(content_list)
                                LenVal = len(content_list)

                                for i in range(LenVal):
                                    if my_array[i] == "NET_NAME":
                                        j = i + 1
                                        PassNetName = my_array[i + 1][1:-1]

                                        SubRefList = []
                                        SubRefList.append(PassNetName)

                                        while j <= LenVal - 1:

                                            if my_array[j] == "NET_NAME":
                                                break

                                            parsed_list = my_array[j].split()

                                            if len(parsed_list) > 0 and parsed_list[0] == "NODE_NAME":
                                                DutName = parsed_list[1].upper()
                                                if DutName == InpDutNameUp:
                                                    SubRefList.append(parsed_list[2])

                                            j += 1

                                        if SubRefList:
                                            SubRefList = list(dict.fromkeys(SubRefList))
                                            RefdesList.append(SubRefList)

                        except Exception as e:
                            print("DAT Error:", e)
                            return

                        # ================= EXCEL FILE =================
                        messagebox.showinfo("Connection", "Now Select The Input Excel File..!")
                        excel_path = filedialog.askopenfilename(
                            title="Select Excel File",
                            filetypes=[("Excel Files", "*.xlsx")]
                        )

                        if not excel_path:
                            return

                        try:
                            wb = load_workbook(excel_path)
                            ws = wb.active

                            for row in ws.iter_rows(min_row=2, values_only=True):
                                colA = row[0]   # Site
                                colB = row[1]   # Pin

                                if colA is None or colB is None:
                                    continue

                                ExcelList.append([colA, colB])

                        except Exception as e:
                            print("Excel Error:", e)
                            return

                        # ================= GROUP EXCEL BY SITE =================
                        GroupedExcelList = OrderedDict()

                        for site, pin in ExcelList:
                            site = str(site)

                            if site not in GroupedExcelList:
                                GroupedExcelList[site] = [site]

                            GroupedExcelList[site].append(pin)

                        FinalExcelGroupedList = list(GroupedExcelList.values())

                        # ================= BUILD PIN → SITE MAP =================
                        PinToSite = {}

                        for group in FinalExcelGroupedList:
                            site = group[0]
                            pins = group[1:]

                            for pin in pins:
                                PinToSite[pin] = site

                        # ================= MULTI-SITE CHECK =================
                        ErrorList = []

                        for RefdesListlp in RefdesList:

                            NetName = RefdesListlp[0]
                            Pins = RefdesListlp[1:]

                            FoundSites = set()

                            for pin in Pins:
                                if pin in PinToSite:
                                    FoundSites.add(PinToSite[pin])

                            if len(FoundSites) > 1:
                                ErrorList.append([NetName] + list(FoundSites))

                        if ErrorList:
                            messagebox.showwarning("Result", "Same-Site Short Errors Found")
                            
                            folder_path = os.path.dirname(dat_path)

                            timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
                            filename = f"Same Site Short {timestamp}.xlsx"
                            save_path = os.path.join(folder_path, filename)

                            wb = Workbook()
                            ws = wb.active
                            ws.title = "Same Site Short"

                            for row in ErrorList:
                                ws.append(row)

                            wb.save(save_path)

                            os.startfile(save_path)

                        else:
                            messagebox.showinfo("Result", "No Same-Site Short Errors Found!")

                    root = tk.Tk()
                    root.withdraw()
                    open_file()
                else:
                    messagebox.showerror("Error", f"Please Check.. The Required Input Field is Empty..")
            
            def create_button(text, relx, rely, command):
                btn = tk.Label(window, text=text, font=("Trebuchet MS", 20),
                            fg="#F6FAFC", bg="#020638", bd=0)
                btn.bind("<Enter>", lambda e: btn.config(fg="#1F5186"))
                btn.bind("<Leave>", lambda e: btn.config(fg="#F6FAFC"))
                btn.bind("<Button-1>", lambda e: command())
                btn.place(relx=relx, rely=rely, anchor="center")

            def create_label_entry_lr(text, rely):
                tk.Label(window, text=text, font=("Didot", 14), bg="#f0f0f0").place(relx=0.40, rely=rely, anchor="center")
                e = tk.Entry(window, font=("Didot", 13), width=25)
                e.place(relx=0.65, rely=rely, anchor="center")
                return e

            form_y = 0.25
            sp = 0.10

            entry1 = create_label_entry_lr("[ DUT COMMON NAME : ]", form_y)

            create_button("[ CLICK HERE ]", 0.5, form_y + sp * 2, SiteToSiteShortSame)
            create_button("[<-- BACK ]", 0.1, 0.05, SchFunc)

#------------------------------------------------------------------------------------------------------------------------------------------------------------------ 

        def PinSameNetNameFunc():
            UsedDt = "Schematic Menu - Used : Pin Same Net Check"
            PurposeList.append(UsedDt)
            
            for widget in window.winfo_children():
                if isinstance(widget, tk.Label) and hasattr(widget, 'image'):
                    continue
                widget.destroy()
            
            def PinSameNetShort():
                
                from tkinter import filedialog
                import numpy as np
                import csv
                from tkinter.filedialog import asksaveasfilename
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill
                from datetime import datetime
                import os
                import sys
                import subprocess
                
                StartPinNo = entry6.get()
                EndPinNo = entry7.get()
                
                if StartPinNo != "" and EndPinNo != "":
            
                    def open_file():
                        messagebox.showinfo("Connection", "Please Select The Netlist File..!")
                        file_path = filedialog.askopenfilename(title="Select a DAT File", filetypes=[("DAT Files", "*.dat")])
                        if file_path:  
                            try:
                                with open(file_path, 'r') as file:
                                    content_list = file.readlines()  
                                    content_list = [line.strip() for line in content_list]
                                    
                                    my_array = np.array(content_list)
                                    LenVal = len(content_list)

                                    ReportList = []
                                    TitleVal = f"REFDES,NETNAME - PIN NO : {StartPinNo},NETNAME - PIN NO : {EndPinNo}"
                                    ReportList.append(TitleVal)
                                    
                                    Val1List = []
                                    Val2List = []
                                    
                                    RefdesList = []
                                    for i in range(LenVal):
                                        if my_array[i] == "NET_NAME":
                                            j = i + 1
                                            PassNetName = my_array[i + 1][1:-1]
                                            PassNetNameUp = PassNetName.upper()
                                            StrNetName = "NET_NAME" + "--" + PassNetName
                                            SubRefList = []
                                            while j <= LenVal - 1:
                                                parsed_list = my_array[j].split()
                                                FirstEle = parsed_list[0]
                                                if FirstEle == "NODE_NAME":
                                                    StrPassVal = parsed_list[1] + "," + parsed_list[2] 
                                                    if parsed_list[2] == StartPinNo or parsed_list[2] == EndPinNo:
                                                        SubRefList.append(PassNetName)
                                                        SubRefList.append(StrPassVal)
                                                    
                                                    if parsed_list[2] == StartPinNo:
                                                        Val1 = f"{PassNetName},{parsed_list[1]}"
                                                        Val1List.append(Val1)
                                                    
                                                    if parsed_list[2] == EndPinNo:
                                                        Val2 = f"{PassNetName},{parsed_list[1]}"
                                                        Val2List.append(Val2)
                                                        
                                                if my_array[j] == "NET_NAME":
                                                    break
                                                j += 1
                                            
                                            if SubRefList != []:
                                                SubRefList = list(dict.fromkeys(SubRefList))
                                                RefdesList.append(SubRefList)                           
                                    
                                    
                                    for Val1Listlp in Val1List:
                                        Val1ListPass = Val1Listlp.split(',')
                                        
                                        for Val2Listlp in Val2List:
                                            Val2ListPass = Val2Listlp.split(',')
                                            
                                            if Val1ListPass[1] == Val2ListPass[1]:
                                                RepVal = f"{Val1ListPass[1]},{Val1ListPass[0]},{Val2ListPass[0]}"
                                                ReportList.append(RepVal)
                                        
                                    
                                    
                                    if ReportList != []:
                                        ReportList = list(dict.fromkeys(ReportList))
                                    
                                    ErrorList = []

                                    for RefdesListlp in RefdesList:
                                        RefdesListlpUp = RefdesListlp[1:]   
                                        SubRefListChk = []

                                        for RefdesListlpUplp in RefdesListlpUp:
                                            NodeNamePass = RefdesListlpUplp.split(',')
                                            SubRefListChk.append(NodeNamePass[0])

                                        duplicates = list(set([x for x in SubRefListChk if SubRefListChk.count(x) > 1]))
                                        if duplicates:   
                                            ErrorSubList = []
                                            ErrorSubList.append(RefdesListlp[0])   
                                            
                                            for duplicateslp in duplicates:
                                                ErrorSubList.append(duplicateslp)

                                            ErrorList.append(ErrorSubList)
                                    
                                    if ErrorList != [] or ReportList != []:
                                        folder_path = os.path.dirname(file_path)

                                        timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
                                        filename = f"Pins Same Net Name {timestamp}.xlsx"
                                        save_path = os.path.join(folder_path, filename)

                                        wb = Workbook()
                                        ws = wb.active
                                        ws.title = "Error Pin Short"

                                        TitleList = ("NETNAME", "REFDES")
                                        ws.append(TitleList)
                                        for row in ErrorList:
                                            ws.append(row)
                                        
                                        ws2 = wb.create_sheet("Overall Report")
                                        for row1 in ReportList:
                                            row1Str = row1.split(',')
                                            ws2.append(row1Str)

                                        wb.save(save_path)

                                        os.startfile(save_path)
                                    else:
                                        messagebox.showinfo("Pin Same Net", f"No Error Found..!")

                                            
                                                
                            except Exception as e:
                                print(f"Error opening the file: {e}")
                    
                    NetL = tk.Tk()
                    NetL.withdraw()  
                    open_file()
                
                else:
                    messagebox.showerror("Error", f"Please Check.. The Required Input Field is Empty..")
                
            
            def create_button(text, relx, rely, command):
                btn = tk.Label(window, text=text, font=("Trebuchet MS", 20),
                            fg="#F6FAFC", bg="#020638", bd=0)
                btn.bind("<Enter>", lambda e: btn.config(fg="#1F5186"))
                btn.bind("<Leave>", lambda e: btn.config(fg="#F6FAFC"))
                btn.bind("<Button-1>", lambda e: command())
                btn.place(relx=relx, rely=rely, anchor="center")

            def create_label_entry_range(text, rely):
                tk.Label(window, text=text, font=("Didot", 14), bg="#f0f0f0").place(relx=0.40, rely=rely, anchor="center")
                e1 = tk.Entry(window, font=("Didot", 8), width=10)
                e1.place(relx=0.60, rely=rely, anchor="center")
                tk.Label(window, text="And", font=("Didot", 13), bg="#f0f0f0").place(relx=0.65, rely=rely, anchor="center")
                e2 = tk.Entry(window, font=("Didot", 8), width=10)
                e2.place(relx=0.70, rely=rely, anchor="center")
                return e1, e2

            form_y = 0.25
            sp = 0.10

            entry6, entry7 = create_label_entry_range("[ COMMON PIN NUMBER ]", form_y + sp)

            create_button("[ CLICK HERE ]", 0.5, form_y + sp * 3, PinSameNetShort)
            create_button("[<-- BACK ]", 0.1, 0.05, SchFunc)

#------------------------------------------------------------------------------------------------------------------------------------------------------------------ 

        def ERCUnConAndUnusedFunc():
            
            UsedDt = "Schematic Menu - Used : ERC - Un Connected and Un Used Net"
            PurposeList.append(UsedDt)
            
            from tkinter import filedialog
            import numpy as np
            from tkinter.filedialog import asksaveasfilename
            from openpyxl import Workbook
            from datetime import datetime
            import os
            
            def open_file():
                messagebox.showinfo("Connection", "Please Select The ERC File..!")
                file_path = filedialog.askopenfilename(title="Select a TXT File", filetypes=[("TXT Files", "*.txt")])
                if file_path:  
                    try:
                        with open(file_path, 'r') as file:
                            content_list = file.readlines()  
                            content_list = [line.strip() for line in content_list]
                            
                            my_array = np.array(content_list)
                            LenVal = len(content_list)

                            ReportList = []
                            for i in range(LenVal):
                                Sublist = []
                                if "Single node net" in my_array[i] and "#" in my_array[i]:
                                    NetName = my_array[i + 1]
                                    NetNameFilt = NetName.split(':')
                                    NetNameStr = NetNameFilt[2]
                                    
                                    Sublist.append(NetNameStr)
                                    
                                    PinName = my_array[i + 2]
                                    PinNameFilt = PinName.split(':')
                                    PinNameStr = PinNameFilt[1]
                                    
                                    Sublist.append(PinNameStr)
                                    
                                    PageNum = my_array[i + 4]
                                    PageNumFilt = PageNum.split(':')
                                    PageNumStr = PageNumFilt[4]
                                    
                                    Sublist.append(PageNumStr)
                                
                                if Sublist != []:
                                    ReportList.append(Sublist)                                                                              
                            
                            if ReportList != []:
                                folder_path = os.path.dirname(file_path)

                                timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
                                filename = f"ERC UnConnected and Unused {timestamp}.xlsx"
                                save_path = os.path.join(folder_path, filename)

                                wb = Workbook()
                                ws = wb.active
                                ws.title = "ERC UnConnected and Unused"

                                for row in ReportList:
                                    ws.append(row)

                                wb.save(save_path)

                                os.startfile(save_path)
                            else:
                                messagebox.showinfo("ERC: UnConnected and Unused", f"No Data Found..!")
                                
                                
                    except Exception as e:
                        print(f"Error opening the file: {e}")
                    
            NetL = tk.Tk()
            NetL.withdraw()  
            open_file()

#------------------------------------------------------------------------------------------------------------------------------------------------------------------ 

        def NetnameRepByPageFunc():
            UsedDt = "Schematic Menu - Used : Netname Report By Page"
            PurposeList.append(UsedDt)
            
            for widget in window.winfo_children():
                if isinstance(widget, tk.Label) and hasattr(widget, 'image'):
                    continue
                widget.destroy()
            
            def NetNameRepByPage():
                import os
                import csv
                from tkinter import filedialog, messagebox
                import numpy as np
                from tkinter.filedialog import asksaveasfilename
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill
                from datetime import datetime

                messagebox.showinfo("NetName Report", "Please Select the Excel Files located Folder..")
                folder_path = filedialog.askdirectory(title="Select a Folder")

                if not folder_path:
                    return  # user cancelled

                StartPinNo = entry6.get().strip()
                EndPinNo = entry7.get().strip()

                # Validate input
                if not StartPinNo.isdigit() or not EndPinNo.isdigit():
                    messagebox.showerror("Error", "Please enter valid numeric page range.")
                    return

                start_page = int(StartPinNo)
                end_page = int(EndPinNo)

                if start_page > end_page:
                    messagebox.showerror("Error", "Start page should be less than or equal to End page.")
                    return

                # Loop through page range
                FinalList = []
                FinalList2 = []
                for page_num in range(start_page, end_page + 1):
                    SubList = []
                    file_name = f"page{page_num}.csv"
                    Page_Name = f"page{page_num}"
                    Page_Name = Page_Name.upper()
                    file_path = os.path.join(folder_path, file_name)
                    SubList.append(Page_Name)
                    #FinalList2.append(Page_Name)
                    if os.path.exists(file_path):
                        with open(file_path, newline='', encoding='utf-8') as csvfile:
                            reader = csv.reader(csvfile)

                            for row in reader:
                                if len(row) > 0:
                                    if "%" in row[0]:
                                        break
                                    
                                    if '"' in row[0]:
                                        NetNameData = row[0]
                                        NetNameDataFilt = NetNameData.split('"')
                                        SubList.append(NetNameDataFilt[1])
                                        FinalList2.append(NetNameDataFilt[1])
                                        
                    FinalList.append(SubList) 
                
                
                timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
                filename_out = f"Netname Report By Page {timestamp}.xlsx"
                save_path = os.path.join(folder_path, filename_out)

                # Create workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Netname Report By Page"

                # Write row-wise
                for col_index, col_data in enumerate(FinalList, start=1):
                    for row_index, value in enumerate(col_data, start=1):
                        ws.cell(row=row_index, column=col_index, value=value)
                
                # ---------- Sheet 2 ----------
                ws2 = wb.create_sheet("Summary")

                for row_index, value in enumerate(FinalList2, start=1):
                    ws2.cell(row=row_index, column=1, value=value)

                # Save file
                wb.save(save_path)

                # ✅ Automatically open the file (Windows)
                os.startfile(save_path)
                
                    
            def create_button(text, relx, rely, command):
                btn = tk.Label(window, text=text, font=("Trebuchet MS", 20),
                            fg="#F6FAFC", bg="#020638", bd=0)
                btn.bind("<Enter>", lambda e: btn.config(fg="#1F5186"))
                btn.bind("<Leave>", lambda e: btn.config(fg="#F6FAFC"))
                btn.bind("<Button-1>", lambda e: command())
                btn.place(relx=relx, rely=rely, anchor="center")

            def create_label_entry_range(text, rely):
                tk.Label(window, text=text, font=("Didot", 14), bg="#f0f0f0").place(relx=0.40, rely=rely, anchor="center")
                e1 = tk.Entry(window, font=("Didot", 8), width=10)
                e1.place(relx=0.60, rely=rely, anchor="center")
                tk.Label(window, text="To", font=("Didot", 13), bg="#f0f0f0").place(relx=0.65, rely=rely, anchor="center")
                e2 = tk.Entry(window, font=("Didot", 8), width=10)
                e2.place(relx=0.70, rely=rely, anchor="center")
                return e1, e2

            form_y = 0.25
            sp = 0.10

            entry6, entry7 = create_label_entry_range("[ ENTER THE PAGE RANGE ]", form_y + sp)

            create_button("[ CLICK HERE ]", 0.5, form_y + sp * 3, NetNameRepByPage)
            create_button("[<-- BACK ]", 0.1, 0.05, SchFunc)

#------------------------------------------------------------------------------------------------------------------------------------------------------------------ 
            
        def create_button(text, y_pixel_pos, command):
            btn = tk.Label(window, text=text, font=("Trebuchet MS", 15), fg="#F6FAFC",
                    bg="#020638", bd=0)
            btn.bind("<Enter>", lambda e: btn.config(fg="#1F5186"))
            btn.bind("<Leave>", lambda e: btn.config(fg="#F6FAFC"))
            btn.bind("<Button-1>", lambda e: command())
            btn.place(x=20, y=y_pixel_pos)   

        create_button("[<-- BACK ]", 20, show_main_menu)
        create_button("[ -- CH-MAPPING NETLIST -- ]", 100, NetLisDt)
        create_button("[ -- BGA PIN LIST CHECK -- ]", 140, BGAPinListFunc)
        create_button("[ -- PIN LIST RENAME -- ]", 180, PinListFunc)
        create_button("[ -- NETLIST COMPARE -- ]", 220, TempNetListComp)
        create_button("[ -- NETLIST UPDATE -- ]", 260, NetUpFunc)
        create_button("[ -- SITE NET REPORT BY NETLIST -- ]", 300, SiteChkByNLFunc)
        create_button("[ -- SOCKET SITE NET REPORT BY NETLIST -- ]", 340, SocSiteChkByNLFunc)
        create_button("[ -- CONNECTION REPORT BY NETLIST -- ]", 380, AutoConRepFunc)
        create_button("[ -- CONNECTION REPORT BY NETLIST(CBIT) -- ]", 420, AutoConRepCBitFunc)
        create_button("[ -- SITE TO SITE SHORT DIFF DUT -- ]", 460, AutoSiteToSiteDiffFunc)
        create_button("[ -- SITE TO SITE SHORT SAME DUT -- ]", 500, AutoSiteToSiteSameFunc)
        create_button("[ -- PIN : SAME NET NAME -- ]", 540, PinSameNetNameFunc)
        create_button("[ -- ERC : UNCONNECTED AND UNUSED  -- ]", 580, ERCUnConAndUnusedFunc)
        create_button("[ -- NETNAME REPORT BY PAGE  -- ]", 620, NetnameRepByPageFunc)

        def create_button1(text, y_pixel_pos1, command):
            btn1 = tk.Label(window, text=text, font=("Trebuchet MS", 15), fg="#F6FAFC",
                    bg="#020638", bd=0)
            btn1.bind("<Enter>", lambda e: btn1.config(fg="#1F5186"))
            btn1.bind("<Leave>", lambda e: btn1.config(fg="#F6FAFC"))
            btn1.bind("<Button-1>", lambda e: command())
            btn1.place(x=500, y=y_pixel_pos1)  

        create_button1("[|-- APPLICATION MENU --|]", 20, MenuPass)
        create_button1("[ -- EXCEL -- ]", 100, ExcelFunc)
        create_button1("[ -- APPS - 2 -- ]", 140, Apps2Func)
        create_button1("[ -- APPS - 4 -- ]", 180, Apps4Func)
        create_button1("[ -- TCL SCH - 2 -- ]", 220, TCLSch2Func)
        create_button1("[ -- TCL SCH - 4 -- ]", 260, TCLSch4Func)
 

#------------------------------------------------------------------------------------------------------------------------------------------------------------------


    def IntFunc():
        global  cap, after_id


        UsedDt = "Opened - Documentation Menu"
        PurposeList.append(UsedDt)
        
        if after_id:
            window.after_cancel(after_id)
            after_id = None
        if cap:
            cap.release()
        for widget in window.winfo_children():
            widget.destroy()

        cap_path = resource_path("VidMain1.mp4")
        cap = cv2.VideoCapture(cap_path)
        bg_label = tk.Label(window)
        bg_label.place(relwidth=1, relheight=1)
        update_video_frame(bg_label, cap_path)
        
        
        #------------------------------------------------------------------------------------------------------------------------------------------------------------------
        

        def AttManagementIdealhrs():
            UsedDt = "Documetation : Opened - Attendance Report Ideal Hrs"
            PurposeList.append(UsedDt)
            
            from tkinter import messagebox
            import pandas as pd
            from tkinter import Tk
            from tkinter.filedialog import askopenfilename

            messagebox.showinfo("Attendance Report", "Please Select The Input Excel File To Generate Updated Attendance Excel File..!")

            Tk().withdraw()

            file_path = askopenfilename(title="Select an Excel file", filetypes=[("Excel Files", "*.xlsx;*.xls")])

            df = pd.read_excel(file_path, index_col=None, dtype=str, header=None)

            outputlist = []


            for _, row in df.iterrows():
                row_values = []
                for value in row:
                    if isinstance(value, str) and '-' in value:  
                        try:
                            date_obj = pd.to_datetime(value, errors='raise')  
                            row_values.append(date_obj.strftime('%d-%b-%y'))  
                        except ValueError:
                            row_values.append(value)  
                    else:
                        row_values.append(str(value))  
                outputlist.append(",".join(row_values))  
                
            
            valid_combinations = []
            valid_combinations1 = []
            for i in range(1, 9):  
                for j in range(1, 9):  
                    for k in range(1, 9):  
                        if i + j + k == 8:  
                            PassStr = f"{i},{j},{k}"
                            valid_combinations.append(PassStr)
                        if i + j + k == 4: 
                            PassStr1 = f"{i},{j},{k}"
                            valid_combinations1.append(PassStr1)
            
            valid_combinationsKP = []
            valid_combinations1KP = []
            for i in range(1, 9):  
                for j in range(1, 9):  
                    if i + j == 8:  
                        PassStr2 = f"{i},{j}"
                        valid_combinationsKP.append(PassStr2)
                    if i + j == 4: 
                        PassStr3 = f"{i},{j}"
                        valid_combinations1KP.append(PassStr3)
            
            import random
            FinalList = []
            FinalListKP = []
            SNo = 0
            SNoKp = 0
            HeadDumAS = 0
            HeadDumKP = 0
            DummyVal = 1
            
            TeamNameListVal = []
            for TeamFiltlp in outputlist[1:]:
                TeamNameList = TeamFiltlp.split(",") 
                TeamName = TeamNameList[3]
                TeamName = TeamName.upper()
                TeamNameListVal.append(TeamName)
            
            TeamNameSort = sorted(set(TeamNameListVal))

            for TeamNameSortlp in TeamNameSort:
                for outputlistlp in outputlist:
                    DetailsList = outputlistlp.split(",") 
                    TeamInpName = DetailsList[3]
                    TeamInpName = TeamInpName.upper()
                    if DummyVal == 1:
                        if HeadDumAS == 0:
                            HeadDt = "S.No, EMP ID, NAME, DEPT/TEAM, BATCH, PAN NUMBER, INDIA %, EDA %, IDEAL HOURS %"
                            DumHeadList = []
                            DateHeadList = []
                            for k in range(6, len(DetailsList)): 
                                Val1 = "India"
                                Val2 = "EDA"
                                Val3 = "Ideal Hours"
                                DumHeadList.append(Val1)
                                DumHeadList.append(Val2)
                                DumHeadList.append(Val3)
                                
                                ValDate = f"-,{DetailsList[k]},-"
                                DateHeadList.append(ValDate)
                            
                            UpdatedDate =  ','.join(DateHeadList)
                            UpdatedDumAS = ','.join(DumHeadList)
                            
                            FinalDateAS = f"-,-,-,-,-,-,-,-,-,{UpdatedDate},-,-,-,-"
                            FinalList.append(FinalDateAS)
                            FinalHeadAS = f"{HeadDt},{UpdatedDumAS},Total Hours India,Total Hours EDA,Total Ideal Hours,Total Working Hours"
                            FinalList.append(FinalHeadAS)
                            HeadDumAS = HeadDumAS + 1
                        
                        if HeadDumKP == 0:
                            HeadDtKP = "S.No, EMP ID, NAME, DEPT/TEAM, BATCH, PAN NUMBER, INDIA %"
                            DumHeadListKP = []
                            DateHeadListKP = []
                            for l in range(6, len(DetailsList)): 
                                Val1KP = "India"
                                Val2KP = "EDA"
                                DumHeadListKP.append(Val1KP)
                                #DumHeadListKP.append(Val2KP)
                                
                                ValDateKP = f"{DetailsList[l]}"
                                DateHeadListKP.append(ValDateKP)
                            
                            UpdatedDateKP = ','.join(DateHeadListKP)
                            UpdatedDumKP = ','.join(DumHeadListKP)
                            
                            FinalDateKP = f"-,-,-,-,-,-,-,{UpdatedDateKP},-"
                            FinalListKP.append(FinalDateKP)
                            FinalHeadKP = f"{HeadDtKP},{UpdatedDumKP},Total Hours India"
                            FinalListKP.append(FinalHeadKP)
                            HeadDumKP = HeadDumKP + 1
                        DummyVal = DummyVal + 1
                    else:
                        if TeamInpName == TeamNameSortlp:       
                            if DetailsList[1] == "nan":
                                EmpID = "No ID Found"
                            else:
                                EmpID = DetailsList[1]
                            
                            if DetailsList[2] == "nan":
                                Name = "No Name Found"
                            else:
                                Name = DetailsList[2]
                            
                            if DetailsList[3] == "nan":
                                Team = "No Team Found"
                            else:
                                Team = DetailsList[3]
                        
                            if DetailsList[4] == "nan":
                                Branch = "No Branch Found"
                            else:
                                Branch = DetailsList[4]
                            
                            if DetailsList[5] == "nan":
                                PAN = "No PAN Num Found"
                            else:
                                PAN = DetailsList[5]
                        
                            DefaultDt = f"{EmpID},{Name},{Team},{Branch},{PAN}"
                            BranchUp = Branch.upper()
                            if "KARUR" in BranchUp or "PERUNDURAI" in BranchUp:
                                PassListKP = []
                                IndHrsKP = 0
                                PacHrsKP = 0
                                for i in range(6, len(DetailsList)):
                                    ValUpKP = DetailsList[i].upper()
                                    ValUpKP = ValUpKP.strip()
                                    if ValUpKP == "8" or ValUpKP == "WH" or ValUpKP == "4/WH" or ValUpKP == "WH/4" or ValUpKP == "OD" or ValUpKP == "OD/4" or ValUpKP == "4/OD" or ValUpKP == "OD/WH" or ValUpKP == "WH/OD":
                                        #random_valueKP = random.choice(valid_combinationsKP)
                                        #Ran_PrseKP = random_valueKP.split(",")
                                        #PassListKP.append(Ran_PrseKP[0]) 
                                        #IndHrsKP = IndHrsKP + int(Ran_PrseKP[0])
                                        #PassListKP.append(Ran_PrseKP[1])  
                                        #PacHrsKP = PacHrsKP + int(Ran_PrseKP[1])
                                        
                                        PassListKP.append("8") 
                                        IndHrsKP = IndHrsKP + 8
                                        
                                    elif ValUpKP == "4/A" or ValUpKP == "A/4" or ValUpKP == "WH/A" or ValUpKP == "A/WH" or ValUpKP == "4" or ValUpKP == "OD/A" or ValUpKP == "A/OD":
                                        #random_valueKP = random.choice(valid_combinations1KP)
                                        #Ran_PrseKP = random_valueKP.split(",")
                                        #PassListKP.append(Ran_PrseKP[0]) 
                                        #IndHrsKP = IndHrsKP + int(Ran_PrseKP[0])
                                        #PassListKP.append(Ran_PrseKP[1])  
                                        #PacHrsKP = PacHrsKP + int(Ran_PrseKP[1])
                                        
                                        PassListKP.append("4") 
                                        IndHrsKP = IndHrsKP + 4
                                        
                                    else:
                                        PassListKP.append("0")  
                                
                                UpdatedtKP = ','.join(PassListKP)
                                
                                #TotalHrsKP = IndHrsKP + PacHrsKP 
                                #IndPrctKP = (IndHrsKP / TotalHrsKP) * 100
                                #IndPrctKP = f"{IndPrctKP:.2f}"
                                #PacPrctKP = (PacHrsKP / TotalHrsKP) * 100
                                #PacPrctKP = f"{PacPrctKP:.2f}"
                                
                                
                                SNoKp = SNoKp + 1 
                                #FinalDtKP = f"{SNoKp},{DefaultDt},{IndPrctKP},{PacPrctKP},{UpdatedtKP},{IndHrsKP},{PacHrsKP},{TotalHrsKP}"
                                FinalDtKP = f"{SNoKp},{DefaultDt},100,{UpdatedtKP},{IndHrsKP}"
                                FinalListKP.append(FinalDtKP)
                            else:
                                PassList = []
                                IndHrs = 0
                                PacHrs = 0
                                IdlHrs = 0
                                for i in range(6, len(DetailsList)):
                                    ValUp = DetailsList[i].upper()
                                    ValUp = ValUp.strip()
                                    if ValUp == "8" or ValUp == "WH" or ValUp == "4/WH" or ValUp == "WH/4" or ValUp == "OD" or ValUp == "OD/4" or ValUp == "4/OD" or ValUp == "OD/WH" or ValUp == "WH/OD":
                                        random_value = random.choice(valid_combinations)
                                        Ran_Prse = random_value.split(",")
                                        PassList.append(Ran_Prse[0]) 
                                        IndHrs = IndHrs + int(Ran_Prse[0])
                                        PassList.append(Ran_Prse[1])  
                                        PacHrs = PacHrs + int(Ran_Prse[1])
                                        PassList.append(Ran_Prse[2])  
                                        IdlHrs = IdlHrs + int(Ran_Prse[2]) 
                                    elif ValUp == "4/A" or ValUp == "A/4" or ValUp == "WH/A" or ValUp == "A/WH" or ValUp == "4" or ValUp == "OD/A" or ValUp == "A/OD":
                                        random_value = random.choice(valid_combinations1)
                                        Ran_Prse = random_value.split(",")
                                        PassList.append(Ran_Prse[0]) 
                                        IndHrs = IndHrs + int(Ran_Prse[0])
                                        PassList.append(Ran_Prse[1])  
                                        PacHrs = PacHrs + int(Ran_Prse[1])
                                        PassList.append(Ran_Prse[2])  
                                        IdlHrs = IdlHrs + int(Ran_Prse[2]) 
                                        
                                    else:
                                        PassList.append("0") 
                                        PassList.append("0")  
                                        PassList.append("0")
                                
                                Updatedt = ','.join(PassList)
                                TotalHrs = IndHrs + PacHrs + IdlHrs
                                IndPrct = (IndHrs / TotalHrs) * 100
                                IndPrct = f"{IndPrct:.2f}"
                                PacPrct = (PacHrs / TotalHrs) * 100
                                PacPrct = f"{PacPrct:.2f}"
                                IdlPrct = (IdlHrs / TotalHrs) * 100
                                IdlPrct = f"{IdlPrct:.2f}"
                                SNo = SNo + 1 
                                FinalDt = f"{SNo},{DefaultDt},{IndPrct},{PacPrct},{IdlPrct},{Updatedt},{IndHrs},{PacHrs},{IdlHrs},{TotalHrs}"
                                FinalList.append(FinalDt)   
            
            
            import pandas as pd
            import tkinter as tk
            from tkinter import filedialog
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            def convert_to_numeric_or_date(value):
                try:
                    return pd.to_numeric(value, errors='raise')
                except ValueError:
                    pass  
                
                try:
                    return pd.to_datetime(value, errors='raise')
                except ValueError:
                    pass  
                
                return value

            def format_dates_in_dataframe(df):
                for col in df.columns:
                    df[col] = df[col].apply(lambda x: x.strftime('%d-%b-%Y').lstrip('0').replace('-0', '-') if isinstance(x, pd.Timestamp) else x)

            def center_cell_content(sheet):
                # Loop through every cell in the sheet and center-align the content
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

            # Assuming FinalList and FinalListKP are defined earlier in the script
            df1 = pd.DataFrame([item.split(',') for item in FinalList])
            df1 = df1.applymap(convert_to_numeric_or_date)  
            if FinalListKP != []:
                df2 = pd.DataFrame([item.split(',') for item in FinalListKP])
                df2 = df2.applymap(convert_to_numeric_or_date)  

            format_dates_in_dataframe(df1)  
            if FinalListKP != []:
                format_dates_in_dataframe(df2)  

            root = tk.Tk()
            root.withdraw()

            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

            if file_path:
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df1.to_excel(writer, sheet_name='AKS', index=False, header=True)
                    if FinalListKP != []:
                        df2.to_excel(writer, sheet_name='KARUR-PERUNDURAI', index=False, header=True)

                wb = load_workbook(file_path)

                def color_same_value_columns(sheet):
                    for col in range(1, sheet.max_column + 1):
                        values = [sheet.cell(row=row, column=col).value for row in range(3, sheet.max_row + 1)]

                        if len(set(values)) == 1 and (values[0] == "0" or values[0] == 0):
                            sheet.cell(row=2, column=col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                            sheet.cell(row=2, column=col).value = "HOLIDAY"
                            continue

                        for row in range(3, sheet.max_row + 1):
                            cell = sheet.cell(row=row, column=col)

                            if cell.value == 0:  
                                cell.fill = PatternFill(start_color="33f9ff", end_color="33f9ff", fill_type="solid")

                            if isinstance(cell.value, pd.Timestamp):  
                                cell.fill = PatternFill(start_color="ffcc99", end_color="ffcc99", fill_type="solid")

                    sheet1 = wb['AKS']
                    for col in range(1, sheet1.max_column):
                        cell_1 = sheet1.cell(row=2, column=col)
                        cell_2 = sheet1.cell(row=2, column=col+1)
                        cell_3 = sheet1.cell(row=2, column=col+2)

                        if cell_1.value == "HOLIDAY" and cell_2.value == "HOLIDAY" and cell_3.value == "HOLIDAY":
                            sheet1.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+2)
                            merged_cell = sheet1.cell(row=2, column=col)
                            merged_cell.value = "HOLIDAY"
                            merged_cell.alignment = merged_cell.alignment.copy(horizontal='center', vertical='center')
                    
                    #sheet2 = wb['KARUR-PERUNDURAI']
                    #for col in range(1, sheet2.max_column):
                        #cell_1 = sheet2.cell(row=2, column=col)
                        #cell_2 = sheet2.cell(row=2, column=col+1)

                        #if cell_1.value == "HOLIDAY" and cell_2.value == "HOLIDAY":
                            #sheet2.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+1)
                            #merged_cell = sheet2.cell(row=2, column=col)
                            #merged_cell.value = "HOLIDAY"
                            #merged_cell.alignment = merged_cell.alignment.copy(horizontal='center', vertical='center')

                def adjust_column_width(sheet):
                    for col in range(1, sheet.max_column + 1):
                        max_length = 0
                        column = get_column_letter(col) 
                        
                        for cell in sheet[column]:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        
                        adjusted_width = (max_length + 2)  
                        sheet.column_dimensions[column].width = adjusted_width

                sheet1 = wb['AKS']
                for row in sheet1.iter_rows(min_row=2, max_row=3): 
                    for cell in row:
                        cell.font = Font(bold=True)
                sheet1.delete_rows(1)
                color_same_value_columns(sheet1)
                center_cell_content(sheet1)  # Center-align the content of sheet1
                adjust_column_width(sheet1)

                if FinalListKP != []:
                    sheet2 = wb['KARUR-PERUNDURAI']
                    for row in sheet2.iter_rows(min_row=2, max_row=3):  
                        for cell in row:
                            cell.font = Font(bold=True)
                    sheet2.delete_rows(1)
                    color_same_value_columns(sheet2)
                    center_cell_content(sheet2)  # Center-align the content of sheet2
                    adjust_column_width(sheet2)

                wb.save(file_path)

                from tkinter import messagebox
                messagebox.showinfo("Attendance Report", f"Result successfully saved to {file_path}")
            else:
                print("No file selected, data not saved.")
                
                
        #------------------------------------------------------------------------------------------------------------------------------------------------------------------        
                
        def AttManagement():
            
            UsedDt = "Documetation : Opened - Attendance Report"
            PurposeList.append(UsedDt)
            
            from tkinter import messagebox
            import pandas as pd
            from tkinter import Tk
            from tkinter.filedialog import askopenfilename


            messagebox.showinfo("Attendance Report", "Please Select The Input Excel File To Generate Updated Attendance Excel File..!")

            Tk().withdraw()

            file_path = askopenfilename(title="Select an Excel file", filetypes=[("Excel Files", "*.xlsx;*.xls")])

            df = pd.read_excel(file_path, index_col=None, dtype=str, header=None)

            outputlist = []


            for _, row in df.iterrows():
                row_values = []
                for value in row:
                    if isinstance(value, str) and '-' in value:  
                        try:
                            date_obj = pd.to_datetime(value, errors='raise')  
                            row_values.append(date_obj.strftime('%d-%b-%y'))  
                        except ValueError:
                            row_values.append(value)  
                    else:
                        row_values.append(str(value))  
                outputlist.append(",".join(row_values))  
                
            
            valid_combinations = []
            valid_combinations1 = []
            for i in range(1, 9):  
                for j in range(1, 9):  
                    for k in range(1, 9):  
                        if i + j + k == 8:  
                            PassStr = f"{i},{j},{k}"
                            valid_combinations.append(PassStr)
                        if i + j + k == 4: 
                            PassStr1 = f"{i},{j},{k}"
                            valid_combinations1.append(PassStr1)
            
            valid_combinationsKP = []
            valid_combinations1KP = []
            for i in range(1, 9):  
                for j in range(1, 9):  
                    if i + j == 8:  
                        PassStr2 = f"{i},{j}"
                        valid_combinationsKP.append(PassStr2)
                    if i + j == 4: 
                        PassStr3 = f"{i},{j}"
                        valid_combinations1KP.append(PassStr3)
            
            import random
            FinalList = []
            FinalListKP = []
            SNo = 0
            SNoKp = 0
            HeadDumAS = 0
            HeadDumKP = 0
            DummyVal = 1
            
            TeamNameListVal = []
            for TeamFiltlp in outputlist[1:]:
                TeamNameList = TeamFiltlp.split(",") 
                TeamName = TeamNameList[3]
                TeamName = TeamName.upper()
                TeamNameListVal.append(TeamName)
            
            TeamNameSort = sorted(set(TeamNameListVal))

            for TeamNameSortlp in TeamNameSort:
                for outputlistlp in outputlist:
                    DetailsList = outputlistlp.split(",") 
                    TeamInpName = DetailsList[3]
                    TeamInpName = TeamInpName.upper()
                    if DummyVal == 1:
                        if HeadDumAS == 0:
                            #HeadDt = "S.No, EMP ID, NAME, DEPT/TEAM, BATCH, PAN NUMBER, INDIA %, EDA %, IDEAL HOURS %"
                            HeadDt = "S.No, EMP ID, NAME, DEPT/TEAM, BATCH, PAN NUMBER, INDIA %, EDA %"
                            DumHeadList = []
                            DateHeadList = []
                            for k in range(6, len(DetailsList)): 
                                Val1 = "India"
                                Val2 = "EDA"
                                #Val3 = "Ideal Hours"
                                DumHeadList.append(Val1)
                                DumHeadList.append(Val2)
                                #DumHeadList.append(Val3)
                                
                                ValDate = f"{DetailsList[k]},-"
                                DateHeadList.append(ValDate)
                            
                            UpdatedDate =  ','.join(DateHeadList)
                            UpdatedDumAS = ','.join(DumHeadList)
                            
                            #FinalDateAS = f"-,-,-,-,-,-,-,-,-,{UpdatedDate},-,-,-,-"
                            FinalDateAS = f"-,-,-,-,-,-,-,-,{UpdatedDate},-,-,-"
                            FinalList.append(FinalDateAS)
                            FinalHeadAS = f"{HeadDt},{UpdatedDumAS},Total Hours India,Total Hours EDA,Total Working Hours"
                            FinalList.append(FinalHeadAS)
                            HeadDumAS = HeadDumAS + 1
                        
                        if HeadDumKP == 0:
                            HeadDtKP = "S.No, EMP ID, NAME, DEPT/TEAM, BATCH, PAN NUMBER, INDIA %"
                            DumHeadListKP = []
                            DateHeadListKP = []
                            for l in range(6, len(DetailsList)): 
                                Val1KP = "India"
                                Val2KP = "EDA"
                                DumHeadListKP.append(Val1KP)
                                #DumHeadListKP.append(Val2KP)
                                
                                ValDateKP = f"{DetailsList[l]}"
                                DateHeadListKP.append(ValDateKP)
                            
                            UpdatedDateKP = ','.join(DateHeadListKP)
                            UpdatedDumKP = ','.join(DumHeadListKP)
                            
                            FinalDateKP = f"-,-,-,-,-,-,-,{UpdatedDateKP},-"
                            FinalListKP.append(FinalDateKP)
                            FinalHeadKP = f"{HeadDtKP},{UpdatedDumKP},Total Hours India"
                            FinalListKP.append(FinalHeadKP)
                            HeadDumKP = HeadDumKP + 1
                        DummyVal = DummyVal + 1
                    else:
                        if TeamInpName == TeamNameSortlp:       
                            if DetailsList[1] == "nan":
                                EmpID = "No ID Found"
                            else:
                                EmpID = DetailsList[1]
                            
                            if DetailsList[2] == "nan":
                                Name = "No Name Found"
                            else:
                                Name = DetailsList[2]
                            
                            if DetailsList[3] == "nan":
                                Team = "No Team Found"
                            else:
                                Team = DetailsList[3]
                        
                            if DetailsList[4] == "nan":
                                Branch = "No Branch Found"
                            else:
                                Branch = DetailsList[4]
                            
                            if DetailsList[5] == "nan":
                                PAN = "No PAN Num Found"
                            else:
                                PAN = DetailsList[5]
                        
                            DefaultDt = f"{EmpID},{Name},{Team},{Branch},{PAN}"
                            BranchUp = Branch.upper()
                            if "KARUR" in BranchUp or "PERUNDURAI" in BranchUp:
                                PassListKP = []
                                IndHrsKP = 0
                                PacHrsKP = 0
                                for i in range(6, len(DetailsList)):
                                    ValUpKP = DetailsList[i].upper()
                                    ValUpKP = ValUpKP.strip()
                                    if ValUpKP == "8" or ValUpKP == "WH" or ValUpKP == "4/WH" or ValUpKP == "WH/4" or ValUpKP == "OD" or ValUpKP == "OD/4" or ValUpKP == "4/OD" or ValUpKP == "OD/WH" or ValUpKP == "WH/OD":
                                        #random_valueKP = random.choice(valid_combinationsKP)
                                        #Ran_PrseKP = random_valueKP.split(",")
                                        #PassListKP.append(Ran_PrseKP[0]) 
                                        #IndHrsKP = IndHrsKP + int(Ran_PrseKP[0])
                                        #PassListKP.append(Ran_PrseKP[1])  
                                        #PacHrsKP = PacHrsKP + int(Ran_PrseKP[1])
                                        
                                        PassListKP.append("8") 
                                        IndHrsKP = IndHrsKP + 8
                                        
                                    elif ValUpKP == "4/A" or ValUpKP == "A/4" or ValUpKP == "WH/A" or ValUpKP == "A/WH" or ValUpKP == "4" or ValUpKP == "OD/A" or ValUpKP == "A/OD":
                                        #random_valueKP = random.choice(valid_combinations1KP)
                                        #Ran_PrseKP = random_valueKP.split(",")
                                        #PassListKP.append(Ran_PrseKP[0]) 
                                        #IndHrsKP = IndHrsKP + int(Ran_PrseKP[0])
                                        #PassListKP.append(Ran_PrseKP[1])  
                                        #PacHrsKP = PacHrsKP + int(Ran_PrseKP[1])
                                        
                                        PassListKP.append("4") 
                                        IndHrsKP = IndHrsKP + 4
                                        
                                    else:
                                        PassListKP.append("0")  
                                
                                UpdatedtKP = ','.join(PassListKP)
                                
                                #TotalHrsKP = IndHrsKP + PacHrsKP 
                                #IndPrctKP = (IndHrsKP / TotalHrsKP) * 100
                                #IndPrctKP = f"{IndPrctKP:.2f}"
                                #PacPrctKP = (PacHrsKP / TotalHrsKP) * 100
                                #PacPrctKP = f"{PacPrctKP:.2f}"
                                
                                
                                SNoKp = SNoKp + 1 
                                #FinalDtKP = f"{SNoKp},{DefaultDt},{IndPrctKP},{PacPrctKP},{UpdatedtKP},{IndHrsKP},{PacHrsKP},{TotalHrsKP}"
                                FinalDtKP = f"{SNoKp},{DefaultDt},100,{UpdatedtKP},{IndHrsKP}"
                                FinalListKP.append(FinalDtKP)
                            else:
                                PassList = []
                                IndHrs = 0
                                PacHrs = 0
                                IdlHrs = 0
                                for i in range(6, len(DetailsList)):
                                    ValUp = DetailsList[i].upper()
                                    ValUp = ValUp.strip()
                                    if ValUp == "8" or ValUp == "WH" or ValUp == "4/WH" or ValUp == "WH/4" or ValUp == "OD" or ValUp == "OD/4" or ValUp == "4/OD" or ValUp == "OD/WH" or ValUp == "WH/OD":
                                        random_value = random.choice(valid_combinationsKP)
                                        Ran_Prse = random_value.split(",")
                                        PassList.append(Ran_Prse[0]) 
                                        IndHrs = IndHrs + int(Ran_Prse[0])
                                        PassList.append(Ran_Prse[1])  
                                        PacHrs = PacHrs + int(Ran_Prse[1])
                                        #PassList.append(Ran_Prse[2])  
                                        #IdlHrs = IdlHrs + int(Ran_Prse[2]) 
                                    elif ValUp == "4/A" or ValUp == "A/4" or ValUp == "WH/A" or ValUp == "A/WH" or ValUp == "4" or ValUp == "OD/A" or ValUp == "A/OD":
                                        random_value = random.choice(valid_combinations1KP)
                                        Ran_Prse = random_value.split(",")
                                        PassList.append(Ran_Prse[0]) 
                                        IndHrs = IndHrs + int(Ran_Prse[0])
                                        PassList.append(Ran_Prse[1])  
                                        PacHrs = PacHrs + int(Ran_Prse[1])
                                        #PassList.append(Ran_Prse[2])  
                                        #IdlHrs = IdlHrs + int(Ran_Prse[2]) 
                                        
                                    else:
                                        PassList.append("0") 
                                        PassList.append("0")  
                                        #PassList.append("0")
                                
                                Updatedt = ','.join(PassList)
                                TotalHrs = IndHrs + PacHrs 
                                IndPrct = (IndHrs / TotalHrs) * 100
                                IndPrct = f"{IndPrct:.2f}"
                                PacPrct = (PacHrs / TotalHrs) * 100
                                PacPrct = f"{PacPrct:.2f}"
                                #IdlPrct = (IdlHrs / TotalHrs) * 100
                                #IdlPrct = f"{IdlPrct:.2f}"
                                SNo = SNo + 1 
                                #FinalDt = f"{SNo},{DefaultDt},{IndPrct},{PacPrct},{IdlPrct},{Updatedt},{IndHrs},{PacHrs},{IdlHrs},{TotalHrs}"
                                FinalDt = f"{SNo},{DefaultDt},{IndPrct},{PacPrct},{Updatedt},{IndHrs},{PacHrs},{TotalHrs}"
                                FinalList.append(FinalDt)   
            
            
            import pandas as pd
            import tkinter as tk
            from tkinter import filedialog
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            def convert_to_numeric_or_date(value):
                try:
                    return pd.to_numeric(value, errors='raise')
                except ValueError:
                    pass  
                
                try:
                    return pd.to_datetime(value, errors='raise')
                except ValueError:
                    pass  
                
                return value

            def format_dates_in_dataframe(df):
                for col in df.columns:
                    df[col] = df[col].apply(lambda x: x.strftime('%d-%b-%Y').lstrip('0').replace('-0', '-') if isinstance(x, pd.Timestamp) else x)

            def center_cell_content(sheet):
                # Loop through every cell in the sheet and center-align the content
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

            # Assuming FinalList and FinalListKP are defined earlier in the script
            df1 = pd.DataFrame([item.split(',') for item in FinalList])
            df1 = df1.applymap(convert_to_numeric_or_date)  
            if FinalListKP != []:
                df2 = pd.DataFrame([item.split(',') for item in FinalListKP])
                df2 = df2.applymap(convert_to_numeric_or_date)  

            format_dates_in_dataframe(df1)  
            if FinalListKP != []:
                format_dates_in_dataframe(df2)  

            root = tk.Tk()
            root.withdraw()

            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

            if file_path:
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df1.to_excel(writer, sheet_name='AKS', index=False, header=True)
                    if FinalListKP != []:
                        df2.to_excel(writer, sheet_name='KARUR-PERUNDURAI', index=False, header=True)

                wb = load_workbook(file_path)

                def color_same_value_columns(sheet):
                    for col in range(1, sheet.max_column + 1):
                        values = [sheet.cell(row=row, column=col).value for row in range(3, sheet.max_row + 1)]

                        if len(set(values)) == 1 and (values[0] == "0" or values[0] == 0):
                            sheet.cell(row=2, column=col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                            sheet.cell(row=2, column=col).value = "HOLIDAY"
                            continue

                        for row in range(3, sheet.max_row + 1):
                            cell = sheet.cell(row=row, column=col)

                            if cell.value == 0:  
                                cell.fill = PatternFill(start_color="33f9ff", end_color="33f9ff", fill_type="solid")

                            if isinstance(cell.value, pd.Timestamp):  
                                cell.fill = PatternFill(start_color="ffcc99", end_color="ffcc99", fill_type="solid")

                    sheet1 = wb['AKS']
                    for col in range(1, sheet1.max_column):
                        cell_1 = sheet1.cell(row=2, column=col)
                        cell_2 = sheet1.cell(row=2, column=col+1)

                        if cell_1.value == "HOLIDAY" and cell_2.value == "HOLIDAY":
                            sheet1.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+1)
                            merged_cell = sheet1.cell(row=2, column=col)
                            merged_cell.value = "HOLIDAY"
                            merged_cell.alignment = merged_cell.alignment.copy(horizontal='center', vertical='center')
                    

                def adjust_column_width(sheet):
                    for col in range(1, sheet.max_column + 1):
                        max_length = 0
                        column = get_column_letter(col) 
                        
                        for cell in sheet[column]:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        
                        adjusted_width = (max_length + 2)  
                        sheet.column_dimensions[column].width = adjusted_width

                sheet1 = wb['AKS']
                for row in sheet1.iter_rows(min_row=2, max_row=3): 
                    for cell in row:
                        cell.font = Font(bold=True)
                sheet1.delete_rows(1)
                color_same_value_columns(sheet1)
                center_cell_content(sheet1)  # Center-align the content of sheet1
                adjust_column_width(sheet1)

                if FinalListKP != []:
                    sheet2 = wb['KARUR-PERUNDURAI']
                    for row in sheet2.iter_rows(min_row=2, max_row=3):  
                        for cell in row:
                            cell.font = Font(bold=True)
                    sheet2.delete_rows(1)
                    color_same_value_columns(sheet2)
                    center_cell_content(sheet2)  # Center-align the content of sheet2
                    adjust_column_width(sheet2)

                wb.save(file_path)

                from tkinter import messagebox
                messagebox.showinfo("Attendance Report", f"Result successfully saved to {file_path}")
            else:
                print("No file selected, data not saved.")


        #------------------------------------------------------------------------------------------------------------------------------------------------------------------
        def create_button(text, y_pixel_pos, command):
            btn = tk.Label(window, text=text, font=("Trebuchet MS", 15), fg="#F6FAFC",
                    bg="#020638", bd=0)
            btn.bind("<Enter>", lambda e: btn.config(fg="#1F5186"))
            btn.bind("<Leave>", lambda e: btn.config(fg="#F6FAFC"))
            btn.bind("<Button-1>", lambda e: command())
            btn.place(x=20, y=y_pixel_pos)  

        create_button("[<-- BACK ]", 20, show_main_menu)
        create_button("[ -- Hours Report With Ideal Hrs -- ]", 100, AttManagementIdealhrs)
        create_button("[ -- Hours Report -- ]", 140, AttManagement)


#------------------------------------------------------------------------------------------------------------------------------------------------------------------


    def QCFunc():
        global  cap, after_id


        UsedDt = "Opened - QC Menu"
        PurposeList.append(UsedDt)
        
        if after_id:
            window.after_cancel(after_id)
            after_id = None
        if cap:
            cap.release()
        for widget in window.winfo_children():
            widget.destroy()

        cap_path = resource_path("VidMain1.mp4")
        cap = cv2.VideoCapture(cap_path)
        bg_label = tk.Label(window)
        bg_label.place(relwidth=1, relheight=1)
        update_video_frame(bg_label, cap_path)
        
        
#------------------------------------------------------------------------------------------------------------------------------------------------------------------


        def BrdFileCompFunc():
            UsedDt = "QC : Opened - Brd Compare"
            PurposeList.append(UsedDt)
            
            import os
            import re
            import subprocess
            import time
            from tkinter.filedialog import askopenfilename
            from tkinter import messagebox
            from openpyxl import Workbook

            def write_mismatch_to_excel(file_path, sheet1_name, sheet1_data, sheet2_name, sheet2_data, headers):
                wb = Workbook()

                # Sheet 1
                ws1 = wb.active
                ws1.title = sheet1_name
                ws1.append(headers)
                for row in sheet1_data:
                    ws1.append(row.split())

                # Sheet 2
                ws2 = wb.create_sheet(title=sheet2_name)
                ws2.append(headers)
                for row in sheet2_data:
                    ws2.append(row.split())

                wb.save(file_path)

            # File selection
            messagebox.showinfo("QC_Brd Compare", "Please Select The Current Design Brd File..!")
            selected1 = askopenfilename(title="Select First .brd File", filetypes=[("Board Files", "*.brd")])
            messagebox.showinfo("QC_Brd Compare", "Please Select The Template Design Brd File..!")
            selected2 = askopenfilename(title="Select Second .brd File", filetypes=[("Board Files", "*.brd")])

            if selected1 and selected2:
                FirstFile = selected1
                SecondFile = selected2
                

                cadence_path = r"C:\Cadence"

                all_folders = [name for name in os.listdir(cadence_path) 
                            if os.path.isdir(os.path.join(cadence_path, name))]

                tool_folders = []
                version_map = {}

                for folder in all_folders:
                    match = re.match(r"SPB_(\d+)\.(\d+)", folder)
                    if match:
                        major, minor = map(int, match.groups())
                        version_tuple = (major, minor)
                        tool_folders.append(folder)
                        version_map[folder] = version_tuple

                found_folders = []

                for folder in tool_folders:
                    full_path = os.path.join(cadence_path, folder)
                    
                    for root, dirs, files in os.walk(full_path):
                        if 'allegro.exe' in files:
                            found_folders.append(folder)
                            break  

                if found_folders:
                    highest_folder = max(found_folders, key=lambda f: version_map[f])
                    extracta_path = os.path.join(cadence_path, highest_folder, "tools", "bin", "extracta.exe")
                    extracta_path = extracta_path.replace("\\", "/")
                    #print(f'\nHighest version extracta path:\n"{extracta_path}"')
                

                    output_folder = os.path.dirname(FirstFile)
                    sym_output_path = os.path.join(output_folder, "QC_Brd_Compare_Sym.xlsx")
                    via_output_path = os.path.join(output_folder, "QC_Brd_Compare_Via.xlsx")

                    # Symbol fields
                    fields = [
                        "SYM_NAME", "SYM_MIRROR", "SYM_PIN_COUNT", "SYM_ROTATE",
                        "SYM_TYPE", "SYM_X", "SYM_Y", "PIN_ROTATION",
                        "PAD_STACK_NAME", "PIN_X", "PIN_Y"
                    ]
                    extracta_script_lines = (
                        ["COMPOSITE_PAD", ""] +
                        [f"{field} != ''" for field in fields] +
                        [""] +
                        fields +
                        ["", "END"]
                    )
                    extracta_script_content = "\n".join(extracta_script_lines)

                    # Via fields
                    fields_via = ["PAD_STACK_NAME", "VIA_ROTATION", "VIA_X", "VIA_Y", "VIA_MIRROR"]
                    extracta_script_lines_via = (
                        ["COMPOSITE_PAD", ""] +
                        [f"{field} != ''" for field in fields_via] +
                        [""] +
                        fields_via +
                        ["", "END"]
                    )
                    extracta_script_content_via = "\n".join(extracta_script_lines_via)

                    timestamp = int(time.time())

                    temp_script_path = os.path.join(output_folder, f"extracta_script_sym_{timestamp}.txt")
                    temp_script_path_via = os.path.join(output_folder, f"extracta_script_via_{timestamp}.txt")

                    with open(temp_script_path, "w") as f:
                        f.write(extracta_script_content)

                    with open(temp_script_path_via, "w") as f:
                        f.write(extracta_script_content_via)

                    CmdStr1 = f'{extracta_path} -s -q "{FirstFile}" "{temp_script_path}"'
                    CmdStr2 = f'{extracta_path} -s -q "{SecondFile}" "{temp_script_path}"'
                    CmdStr3 = f'{extracta_path} -s -q "{FirstFile}" "{temp_script_path_via}"'
                    CmdStr4 = f'{extracta_path} -s -q "{SecondFile}" "{temp_script_path_via}"'

                    try:
                        result1 = subprocess.run(CmdStr1, shell=True, check=True, stdout=subprocess.PIPE, text=True)
                        Brd1 = result1.stdout.splitlines()

                        result2 = subprocess.run(CmdStr2, shell=True, check=True, stdout=subprocess.PIPE, text=True)
                        Brd2 = result2.stdout.splitlines()
                        
                        result3 = subprocess.run(CmdStr3, shell=True, check=True, stdout=subprocess.PIPE, text=True)
                        Brd1_Via = result3.stdout.splitlines()

                        result4 = subprocess.run(CmdStr4, shell=True, check=True, stdout=subprocess.PIPE, text=True)
                        Brd2_Via = result4.stdout.splitlines()

                        if "Revising design for compatibility with current software." in Brd1:
                            Brd1.remove("Revising design for compatibility with current software.")
                        
                        if "Revising design for compatibility with current software." in Brd2:
                            Brd2.remove("Revising design for compatibility with current software.")
                            
                        if "Revising design for compatibility with current software." in Brd1_Via:
                            Brd1_Via.remove("Revising design for compatibility with current software.")
                        
                        if "Revising design for compatibility with current software." in Brd2_Via:
                            Brd2_Via.remove("Revising design for compatibility with current software.")
                        
                        Brd1Details = Brd1 + Brd1_Via
                        Brd2Details = Brd2 + Brd2_Via

                        if Brd1Details == Brd2Details:
                            messagebox.showinfo("QC_Brd Compare", "No Mismatch found. Good to go!")
                        else:
                            if Brd1 != Brd2:
                                missing_in_brd1 = [line for line in Brd2 if line not in Brd1]
                                missing_in_brd2 = [line for line in Brd1 if line not in Brd2]
                                write_mismatch_to_excel(
                                    sym_output_path,
                                    "Missing in Main Brd", missing_in_brd1,
                                    "Missing in Template", missing_in_brd2,
                                    fields
                                )
                                messagebox.showwarning("QC_Brd Compare", f"Mismatch in Symbol data.\nSee:\n{sym_output_path}")

                            if Brd1_Via != Brd2_Via:
                                missing_in_brd1_via = [line for line in Brd2_Via if line not in Brd1_Via]
                                missing_in_brd2_via = [line for line in Brd1_Via if line not in Brd2_Via]
                                write_mismatch_to_excel(
                                    via_output_path,
                                    "Missing in Main Brd", missing_in_brd1_via,
                                    "Missing in Template", missing_in_brd2_via,
                                    fields_via
                                )
                                messagebox.showwarning("QC_Brd Compare", f"Mismatch in Via data.\nSee:\n{via_output_path}")

                    except FileNotFoundError:
                        messagebox.showerror("QC_Brd Compare", "[ERROR] 'extracta' not found. Please ensure it is installed and in your PATH.")
                    except subprocess.CalledProcessError as e:
                        messagebox.showerror("QC_Brd Compare", f"[ERROR] extracta command failed:\n{e}")
                    finally:
                        try:
                            if os.path.exists(temp_script_path):
                                os.remove(temp_script_path)
                            if os.path.exists(temp_script_path_via):
                                os.remove(temp_script_path_via)
                        except Exception as cleanup_error:
                            messagebox.showwarning("QC_Brd Compare", f"Warning: Could not delete temporary files.\n{cleanup_error}")
                else:
                    messagebox.showerror("QC_Brd Compare", "[INFO] Allegro Tool Not Found.")
            else:
                messagebox.showerror("QC_Brd Compare", "[INFO] File selection cancelled.")


#------------------------------------------------------------------------------------------------------------------------------------------------------------------


        def MenuPass():
            from tkinter import messagebox
            messagebox.showinfo("Allegro Menu", "Please Choose Allegro Skill Menu According To Your Team..")


#------------------------------------------------------------------------------------------------------------------------------------------------------------------


        def ExcelFunc():
            UsedDt = "QC : Added - Excel Menu"
            PurposeList.append(UsedDt)
            import os
            import shutil
            import getpass
            import time
            import win32com.client
            from tkinter import messagebox

            SOURCE_PATH = r"\\sng-psrvr06\Cadence_Skills\Menu-enc\MyCustomMenu.xlam"

            USERNAME = getpass.getuser()

            DEST_FOLDER = rf"C:\Users\{USERNAME}\AppData\Roaming\Microsoft\AddIns\Menu"
            DEST_PATH = os.path.join(DEST_FOLDER, os.path.basename(SOURCE_PATH))

            if not os.path.exists(SOURCE_PATH):
                raise FileNotFoundError(f"Source add-in not found: {SOURCE_PATH}")
            print(f"✅ Source file found: {SOURCE_PATH}")

            os.makedirs(DEST_FOLDER, exist_ok=True)
            print(f"✅ Destination folder ready: {DEST_FOLDER}")

            if os.path.exists(DEST_PATH):
                print(f"⚠️ Existing add-in found, replacing: {DEST_PATH}")
                os.remove(DEST_PATH)  
            shutil.copy2(SOURCE_PATH, DEST_PATH)
            print(f"✅ Copied new add-in to: {DEST_PATH}")

            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False 

            excel.Workbooks.Add()
            time.sleep(1)

            addin_name = os.path.basename(DEST_PATH).lower()
            existing_addin = None

            for a in excel.AddIns:
                if os.path.basename(a.FullName).lower() == addin_name:
                    existing_addin = a
                    break

            if existing_addin:
                try:
                    print(f"🔄 Disabling old version of add-in: {existing_addin.FullName}")
                    existing_addin.Installed = False
                    del existing_addin
                    time.sleep(0.5)
                except Exception as e:
                    print(f"⚠️ Warning: could not disable old version: {e}")


            try:
                print(f"➕ Registering new add-in: {DEST_PATH}")
                new_addin = excel.AddIns.Add(DEST_PATH)
                new_addin.Installed = True
                print(f"✅ '{new_addin.Name}' successfully installed and enabled in Excel!")

            except Exception as e:
                print(f"❌ Failed to add or enable add-in: {e}")
                print("➡️ Check if the add-in is in a Trusted Location or not blocked (Properties → Unblock).")

            finally:
                excel.Quit()

            messagebox.showinfo("Excel Update", "✅ EDA Excel Menu.\n -- Updated Successfully --")


#------------------------------------------------------------------------------------------------------------------------------------------------------------------


        def QCSkillFunc():
            UsedDt = "QC : Added - QC Menu"
            PurposeList.append(UsedDt)
            
            from pathlib import Path
            from datetime import datetime
            import shutil
            import getpass
            import os
            import tkinter as tk
            from tkinter import messagebox

            # --- Initialize Tkinter safely (to avoid 'no default root' errors) ---
            root = tk.Tk()
            root.withdraw()  # Hide the main window

            base_path = Path(r"C:\Cadence")

            insert_text1 = """POPUP "&QC"
        BEGIN
            MENUITEM "qc_menu",           "qc_menu"
        END"""

            # --- Update Allegro menu definitions ---
            for folder in base_path.iterdir():
                if folder.is_dir() and "SPB_" in folder.name:
                    link_val = folder / "share" / "pcb" / "text" / "cuimenus"
                    allegro_file = link_val / "allegro.men"

                    if allegro_file.exists():
                        content = allegro_file.read_text(encoding="utf-8")
                        has_ate1 = 'POPUP "&QC"' in content

                        if has_ate1:
                            continue  # Skip already updated files

                        # Backup original
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        backup_file = link_val / f"allegro_Backup_{timestamp}.txt"
                        shutil.copy2(allegro_file, backup_file)

                        # Find insertion point
                        lines = content.splitlines()
                        insert_index = next((i for i in range(len(lines)-1, -1, -1) if lines[i].strip() == "END"), len(lines))

                        updated_lines = lines[:insert_index]

                        if not has_ate1:
                            updated_lines += insert_text1.splitlines()

                        updated_lines.append("END")

                        allegro_file.write_text("\n".join(updated_lines) + "\n", encoding="utf-8")

            # --- Source and destination paths ---
            source_path = Path(r"\\sng-psrvr06\Cadence_Skills\Menu-enc")

            allegro_env = os.environ.get("ALLEGRO_PCBENV")
            home_env = os.environ.get("HOME") or os.environ.get("USERPROFILE")

            # Ensure we have a valid destination path
            if allegro_env:
                dest_path = Path(allegro_env)
            elif home_env:
                dest_path = Path(home_env) / "pcbenv"
            else:
                messagebox.showerror("Menu Update", "Cannot determine user environment path.")
                return

            files_to_copy = ["allegro.ilinit", "env"]

            allowed_lines = [
                'load("//sng-psrvr06/QC-Cadence_Skills/Menu-enc/qc1_menu1.ile")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align.ile" "EncFileKar")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align_do.ile" "EncFileKar")',
                'load("//sng-psrvr06/Cadence_Skills/EDA-enc/align_utils.ile" "EncFileKar")',
            ]

            # --- Copy and clean up files ---
            if not source_path.exists():
                messagebox.showerror("Menu Update", "Error: Source path not found. Contact Automation Team.")
                return

            dest_path.mkdir(parents=True, exist_ok=True)

            for file_name in files_to_copy:
                src_file = source_path / file_name
                dest_file = dest_path / file_name

                if src_file.exists():
                    # Backup existing file
                    if dest_file.exists():
                        backup_file = dest_file.with_suffix(dest_file.suffix + ".bak")
                        shutil.copy2(dest_file, backup_file)

                    # Copy new file
                    shutil.copy2(src_file, dest_file)

                    # Clean up allegro.ilinit after copy
                    if file_name.lower() == "allegro.ilinit":
                        try:
                            lines = dest_file.read_text(encoding="utf-8").splitlines()
                            filtered_lines = [line for line in lines if line.strip() in allowed_lines]
                            dest_file.write_text("\n".join(filtered_lines) + "\n", encoding="utf-8")
                        except Exception as e:
                            messagebox.showerror("Menu Update", f"Error updating allegro.ilinit: {e}")
                else:
                    messagebox.showerror("Menu Update", f"Missing file: {src_file}. Contact Automation Team.")

            messagebox.showinfo("Menu Update", "✅ Please restart Allegro.\nIf SKILL was not updated, contact Automation Team.")

            root.destroy()


#------------------------------------------------------------------------------------------------------------------------------------------------------------------


        def BrdAndBomCompFunc():
            UsedDt = "QC : Used - Brd and Bom Compare"
            PurposeList.append(UsedDt)

            import os
            import re
            import subprocess
            import time
            import pandas as pd
            from datetime import datetime
            from tkinter.filedialog import askopenfilename
            from tkinter import messagebox
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment


            # -------------------------------------------------
            # Expand reference ranges
            # -------------------------------------------------
            def expand_reference(ref):
                ref = str(ref).strip()
                if not ref:
                    return []
                if ref.startswith("+") or ref.startswith("-"):
                    return [ref]
                m = re.match(r"(.+?)(\d+)([A-Za-z]*)-(.+?)(\d+)([A-Za-z]*)$", ref)
                if m:
                    p1, start, s1, p2, end, s2 = m.groups()
                    if p1 == p2 and s1 == s2:
                        start = int(start)
                        end = int(end)
                        if start <= end:
                            return [f"{p1}{i}{s1}" for i in range(start, end + 1)]
                        else:
                            return [f"{p1}{i}{s1}" for i in range(start, end - 1, -1)]
                m = re.match(r"([A-Za-z_+]+)(\d+)-(\d+)$", ref)
                if m:
                    prefix, start, end = m.groups()
                    start = int(start)
                    end = int(end)
                    if start <= end:
                        return [f"{prefix}{i}" for i in range(start, end + 1)]
                    else:
                        return [f"{prefix}{i}" for i in range(start, end - 1, -1)]
                return [ref]


            # -------------------------------------------------
            # FIX #3: Validate a RefDes token
            # Accepted formats:
            #   - Purely numeric          : 1, 10, 205
            #   - Letter-led alphanumeric : R1, C10, U5, TP3, J1A
            #   - Power rails             : +3V3, -12V (start with + or -)
            # Rejected: tokens containing spaces, @, #, $, !, etc.
            # -------------------------------------------------
            def is_valid_refdes(ref):
                ref = str(ref).strip()
                if not ref:
                    return False
                if ref.startswith("+") or ref.startswith("-"):
                    return True
                # Purely numeric (e.g. 1, 10, 205)
                if re.match(r"^\d+$", ref):
                    return True
                # Letter-led: letters, digits, underscores, dots, hyphens
                return bool(re.match(r"^[A-Za-z][A-Za-z0-9_.\-]*$", ref))


            # -------------------------------------------------
            # Select BRD file
            # -------------------------------------------------
            messagebox.showinfo("QC_Brd And Bom Compare", "Please Select The Current Design BRD File")

            SelectedFile = askopenfilename(
                title="Select BRD File",
                filetypes=[("Board Files", "*.brd")]
            )

            if not SelectedFile:
                messagebox.showerror("QC_Brd And Bom Compare", "File selection cancelled.")
                return


            # -------------------------------------------------
            # Find Cadence installation
            # -------------------------------------------------
            cadence_path = r"C:\Cadence"

            if not os.path.exists(cadence_path):
                messagebox.showerror("QC_Brd And Bom Compare", f"Cadence path not found:\n{cadence_path}")
                return

            all_folders = [
                f for f in os.listdir(cadence_path)
                if os.path.isdir(os.path.join(cadence_path, f))
            ]

            version_map = {}
            tool_folders = []

            for folder in all_folders:
                m = re.match(r"SPB_(\d+)\.(\d+)", folder)
                if m:
                    version_map[folder] = tuple(map(int, m.groups()))
                    tool_folders.append(folder)

            found_folders = []
            for folder in tool_folders:
                full_path = os.path.join(cadence_path, folder)
                for root, dirs, files in os.walk(full_path):
                    if "allegro.exe" in files:
                        found_folders.append(folder)
                        break

            if not found_folders:
                messagebox.showerror("QC_Brd And Bom Compare", "Allegro tool not found.")
                return

            highest_folder = max(found_folders, key=lambda f: version_map[f])
            extracta_path = os.path.join(
                cadence_path, highest_folder, "tools", "bin", "extracta.exe"
            ).replace("\\", "/")

            if not os.path.exists(extracta_path.replace("/", "\\")):
                messagebox.showerror("QC_Brd And Bom Compare", f"extracta.exe not found:\n{extracta_path}")
                return


            # -------------------------------------------------
            # Create extracta script
            # -------------------------------------------------
            script_lines = [
                "COMPONENT", "",
                "REFDES != ''",
                "SYM_NAME != ''",
                "SYM_CENTER_X != ''",
                "",
                "REFDES",
                "SYM_NAME",
                "SYM_CENTER_X",
                "",
                "END"
            ]

            output_folder = os.path.dirname(SelectedFile)
            temp_timestamp = int(time.time())
            script_path = os.path.join(output_folder, f"extracta_script_{temp_timestamp}.txt")

            with open(script_path, "w") as f:
                f.write("\n".join(script_lines))


            # -------------------------------------------------
            # Run extracta
            # -------------------------------------------------
            cmd = f'{extracta_path} -s -q "{SelectedFile}" "{script_path}"'

            try:
                result = subprocess.run(
                    cmd, shell=True, check=True,
                    stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True
                )
                Brd_Data = result.stdout.splitlines()
                if "Revising design for compatibility with current software." in Brd_Data:
                    Brd_Data.remove("Revising design for compatibility with current software.")

            except subprocess.CalledProcessError as e:
                err_msg = e.stderr.strip() if e.stderr else str(e)
                messagebox.showerror("QC_Brd And Bom Compare", f"extracta failed:\n{err_msg}")
                return
            except Exception as e:
                messagebox.showerror("QC_Brd And Bom Compare", str(e))
                return
            finally:
                if os.path.exists(script_path):
                    try:
                        os.remove(script_path)
                    except:
                        pass


            # -------------------------------------------------
            # Parse BRD Data  →  brd_refs = {REFDES_upper: (original_REFDES, SYM_NAME)}
            # FIX #1: Store original case for case-mismatch detection later
            # -------------------------------------------------
            # brd_refs keyed by UPPERCASE refdes for case-insensitive lookup
            # value = (original_refdes_from_brd, sym_name)
            brd_refs_upper = {}   # key: upper → value: (orig_ref, sym)

            for line in Brd_Data:
                parts = line.split()
                if len(parts) >= 2:
                    ref = parts[0].strip()
                    sym = parts[1].strip()
                    brd_refs_upper[ref.upper()] = (ref, sym)


            # -------------------------------------------------
            # Select BOM file
            # -------------------------------------------------
            messagebox.showinfo("QC_Brd And Bom Compare", "Please Select BOM Excel File")

            bom_file = askopenfilename(
                title="Select BOM Excel",
                filetypes=[("Excel Files", "*.xlsx *.xls")]
            )

            if not bom_file:
                messagebox.showerror("QC_Brd And Bom Compare", "BOM file not selected.")
                return


            # -------------------------------------------------
            # Read BOM Data
            # -------------------------------------------------
            try:
                df = pd.read_excel(bom_file, header=0)
            except Exception as e:
                messagebox.showerror("QC_Brd And Bom Compare", f"Unable to read BOM file:\n{e}")
                return


            # -------------------------------------------------
            # Tracking structures
            # -------------------------------------------------
            # FIX #1: case mismatch  → [(bom_ref, brd_ref, bom_sym, brd_sym)]
            case_mismatch_rows = []

            # FIX #2: duplicate refdes → [(excel_row, ref, sym, "same_row" / "row_X")]
            seen_refdes = {}    # upper_ref → (first_excel_row, original_ref, sym)
            duplicate_rows = []

            # FIX #3: invalid refdes  → [(excel_row, ref_token, sym)]
            invalid_refdes_rows = []

            # FIX #5: missing symbol or quantity → [(excel_row, ref_text, "Symbol"/"Quantity")]
            missing_field_rows = []

            # Qty mismatch
            qty_mismatch_rows = []

            # Main BOM map keyed by upper for matching
            bom_refs_upper = {}   # key: upper → value: (orig_ref, sym)

            last_sym = None

            for idx, row in df.iterrows():
                excel_row_no = idx + 2

                ref_cell = row.iloc[0] if len(row) > 0 else None
                sym_cell = row.iloc[1] if len(row) > 1 else None
                qty_cell = row.iloc[2] if len(row) > 2 else None

                if pd.isna(ref_cell):
                    continue

                # -----------------------------------------------
                # FIX #5: Detect missing Symbol (footprint) field
                # -----------------------------------------------
                sym_missing = pd.isna(sym_cell) or str(sym_cell).strip() == ""
                qty_missing = pd.isna(qty_cell) or str(qty_cell).strip() == ""

                if sym_missing:
                    sym_value = last_sym   # carry forward for BOM matching (existing behavior)
                else:
                    sym_value = str(sym_cell).strip()
                    last_sym = sym_value

                ref_text = str(ref_cell).replace("\n", "").strip()

                # FIX #4: Support both comma and semicolon as RefDes separators
                raw_refs = [x.strip() for x in re.split(r"[,;]", ref_text) if str(x).strip()]

                expanded_refs_for_this_row = []
                invalid_in_this_row = []

                for ref in raw_refs:
                    expanded = expand_reference(ref)
                    for e_ref in expanded:
                        # FIX #3: Validate each expanded token
                        if not is_valid_refdes(e_ref):
                            invalid_in_this_row.append(e_ref)
                            invalid_refdes_rows.append([excel_row_no, e_ref, sym_value or ""])
                        else:
                            expanded_refs_for_this_row.append(e_ref)

                # FIX #5: Record missing fields (after we know ref_text)
                if sym_missing:
                    missing_field_rows.append([excel_row_no, ref_text, sym_value or "", "Symbol (Footprint)"])
                if qty_missing:
                    missing_field_rows.append([excel_row_no, ref_text, sym_value or "", "Quantity"])

                # -----------------------------------------------
                # FIX #2: Detect duplicates within this row and across rows
                # -----------------------------------------------
                for r in expanded_refs_for_this_row:
                    r_up = r.upper()
                    if r_up in seen_refdes:
                        first_row, first_ref, first_sym = seen_refdes[r_up]
                        if first_row == excel_row_no:
                            location = "Same row (repeated in RefDes list)"
                        else:
                            location = f"Row {first_row}"
                        duplicate_rows.append([
                            excel_row_no, r, sym_value or "",
                            first_row, first_ref, first_sym,
                            location
                        ])
                    else:
                        seen_refdes[r_up] = (excel_row_no, r, sym_value or "")
                        bom_refs_upper[r_up] = (r, sym_value or "")

                # Quantity check
                if not qty_missing:
                    try:
                        input_qty = int(float(qty_cell))
                    except:
                        input_qty = None

                    actual_qty = len(expanded_refs_for_this_row)
                    if input_qty is not None and input_qty != actual_qty:
                        qty_mismatch_rows.append([
                            excel_row_no, ref_text,
                            sym_value if sym_value is not None else "",
                            input_qty, actual_qty
                        ])


            # -------------------------------------------------
            # Compare BRD vs BOM  (case-insensitive)
            # FIX #1: Detect case mismatches separately from true missing
            # -------------------------------------------------
            brd_keys_upper = set(brd_refs_upper.keys())
            bom_keys_upper = set(bom_refs_upper.keys())

            truly_missing_in_bom = []    # in BRD but not in BOM at all
            truly_missing_in_brd = []    # in BOM but not in BRD at all

            for up_key in sorted(brd_keys_upper - bom_keys_upper):
                orig_brd_ref, brd_sym = brd_refs_upper[up_key]
                truly_missing_in_bom.append((orig_brd_ref, brd_sym))

            for up_key in sorted(bom_keys_upper - brd_keys_upper):
                orig_bom_ref, bom_sym = bom_refs_upper[up_key]
                truly_missing_in_brd.append((orig_bom_ref, bom_sym))

            # FIX #1: Case mismatches = refs present in both (case-insensitive) but differ in case
            for up_key in brd_keys_upper & bom_keys_upper:
                orig_brd_ref, brd_sym = brd_refs_upper[up_key]
                orig_bom_ref, bom_sym = bom_refs_upper[up_key]
                if orig_brd_ref != orig_bom_ref:
                    case_mismatch_rows.append([orig_bom_ref, orig_brd_ref, bom_sym, brd_sym])


            # -------------------------------------------------
            # Always generate report  (FIX #6: always create file, not just on error)
            # -------------------------------------------------
            timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
            filename_out = f"QC_Brd_Bom_Compare {timestamp}.xlsx"
            output_file = os.path.join(output_folder, filename_out)

            wb = Workbook()

            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            red_fill    = PatternFill("solid", start_color="C00000", end_color="C00000")
            orange_fill = PatternFill("solid", start_color="ED7D31", end_color="ED7D31")
            yellow_fill = PatternFill("solid", start_color="FFD966", end_color="FFD966")
            blue_fill   = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
            green_fill  = PatternFill("solid", start_color="375623", end_color="375623")
            grey_fill   = PatternFill("solid", start_color="595959", end_color="595959")
            ok_fill     = PatternFill("solid", start_color="70AD47", end_color="70AD47")

            def write_header(ws, headers, fill):
                ws.append(headers)
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            def auto_col_width(ws):
                for col in ws.columns:
                    max_len = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            val = str(cell.value) if cell.value is not None else ""
                            max_len = max(max_len, len(val))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


            # --------------------------------------------------
            # Sheet 1 – Summary
            # --------------------------------------------------
            ws_sum = wb.active
            ws_sum.title = "Summary"

            summary_data = [
                ["Check", "Count", "Status"],
                ["Missing in BOM (BRD has, BOM doesn't)", len(truly_missing_in_bom),
                "OK" if not truly_missing_in_bom else "MISMATCH"],
                ["Missing in BRD (BOM has, BRD doesn't)", len(truly_missing_in_brd),
                "OK" if not truly_missing_in_brd else "MISMATCH"],
                ["Case Mismatch (same ref, different case)", len(case_mismatch_rows),
                "OK" if not case_mismatch_rows else "WARNING"],
                ["Duplicate RefDes in BOM", len(duplicate_rows),
                "OK" if not duplicate_rows else "WARNING"],
                ["Invalid RefDes in BOM", len(invalid_refdes_rows),
                "OK" if not invalid_refdes_rows else "WARNING"],
                ["Missing Symbol / Quantity in BOM", len(missing_field_rows),
                "OK" if not missing_field_rows else "WARNING"],
                ["Quantity Count Mismatch", len(qty_mismatch_rows),
                "OK" if not qty_mismatch_rows else "MISMATCH"],
            ]

            status_fill_map = {
                "OK":       ok_fill,
                "MISMATCH": red_fill,
                "WARNING":  orange_fill,
            }

            write_header(ws_sum, summary_data[0], blue_fill)
            for row_data in summary_data[1:]:
                ws_sum.append(row_data)
                status = row_data[2]
                row_idx = ws_sum.max_row
                for cell in ws_sum[row_idx]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                ws_sum.cell(row=row_idx, column=3).fill = status_fill_map.get(status, ok_fill)
                ws_sum.cell(row=row_idx, column=3).font = Font(bold=True, color="FFFFFF")

            ws_sum.column_dimensions["A"].width = 50
            ws_sum.column_dimensions["B"].width = 12
            ws_sum.column_dimensions["C"].width = 14

            has_error = bool(
                truly_missing_in_bom or truly_missing_in_brd or
                case_mismatch_rows or duplicate_rows or
                invalid_refdes_rows or missing_field_rows or qty_mismatch_rows
            )

            overall_status = "All Checks Passed – BRD and BOM Match" if not has_error else "Issues Found – See Sheets Below"
            ws_sum.insert_rows(1)
            ws_sum["A1"] = overall_status
            ws_sum["A1"].font = Font(bold=True, size=13, color="FFFFFF")
            ws_sum["A1"].fill = ok_fill if not has_error else red_fill
            ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws_sum.merge_cells("A1:C1")
            ws_sum.row_dimensions[1].height = 24


            # --------------------------------------------------
            # Sheet 2 – Missing in BOM
            # --------------------------------------------------
            ws1 = wb.create_sheet("Missing in BOM")
            write_header(ws1, ["Reference (BRD)", "BRD Symbol"], red_fill)
            for ref, sym in truly_missing_in_bom:
                ws1.append([ref, sym])
            auto_col_width(ws1)


            # --------------------------------------------------
            # Sheet 3 – Missing in BRD
            # --------------------------------------------------
            ws2 = wb.create_sheet("Missing in BRD")
            write_header(ws2, ["Reference (BOM)", "BOM Symbol"], red_fill)
            for ref, sym in truly_missing_in_brd:
                ws2.append([ref, sym])
            auto_col_width(ws2)


            # --------------------------------------------------
            # Sheet 4 – Case Mismatch  (FIX #1)
            # These refs exist in both but differ in uppercase/lowercase.
            # They are NOT treated as errors but are flagged for review.
            # --------------------------------------------------
            ws_case = wb.create_sheet("Case Mismatch")
            write_header(
                ws_case,
                ["BOM RefDes", "BRD RefDes", "BOM Symbol", "BRD Symbol", "Note"],
                orange_fill
            )
            for row_data in case_mismatch_rows:
                ws_case.append(row_data + ["Case difference only – not treated as error"])
            auto_col_width(ws_case)


            # --------------------------------------------------
            # Sheet 5 – Duplicate RefDes  (FIX #2)
            # --------------------------------------------------
            ws_dup = wb.create_sheet("Duplicate RefDes")
            write_header(
                ws_dup,
                ["Excel Row", "Duplicate Ref", "Symbol", "First Seen Row", "First Ref", "First Symbol", "Location"],
                orange_fill
            )
            for row_data in duplicate_rows:
                ws_dup.append(row_data)
            auto_col_width(ws_dup)


            # --------------------------------------------------
            # Sheet 6 – Invalid RefDes  (FIX #3)
            # --------------------------------------------------
            ws_inv = wb.create_sheet("Invalid RefDes")
            write_header(
                ws_inv,
                ["Excel Row", "Invalid Token", "Symbol", "Reason"],
                orange_fill
            )
            for row_data in invalid_refdes_rows:
                ws_inv.append(row_data + ["Special character or invalid format in RefDes"])
            auto_col_width(ws_inv)


            # --------------------------------------------------
            # Sheet 7 – Missing Fields  (FIX #5)
            # --------------------------------------------------
            ws_mf = wb.create_sheet("Missing Fields")
            write_header(
                ws_mf,
                ["Excel Row", "RefDes", "Symbol (if available)", "Missing Field"],
                yellow_fill
            )
            # yellow header needs dark font
            for cell in ws_mf[1]:
                cell.font = Font(bold=True, color="000000")
            for row_data in missing_field_rows:
                ws_mf.append(row_data)
            auto_col_width(ws_mf)


            # --------------------------------------------------
            # Sheet 8 – Quantity Mismatch
            # --------------------------------------------------
            ws3 = wb.create_sheet("Quantity Mismatch")
            write_header(
                ws3,
                ["Excel Row", "RefDes List", "BOM Symbol", "Input Qty", "Actual Expanded Qty"],
                grey_fill
            )
            for row_data in qty_mismatch_rows:
                ws3.append(row_data)
            auto_col_width(ws3)


            wb.save(output_file)

            # --------------------------------------------------
            # FIX #6: Always open the report; message reflects result
            # --------------------------------------------------
            if not has_error:
                messagebox.showinfo(
                    "QC_Brd And Bom Compare",
                    f"No mismatch found.\nBRD and BOM are matching.\nComparison report has been saved for reference:\n{output_file}"
                )
            else:
                messagebox.showwarning(
                    "QC_Brd And Bom Compare",
                    f"Issues Found.\nReport Generated:\n{output_file}"
                )

            os.startfile(output_file)


#------------------------------------------------------------------------------------------------------------------------------------------------------------------

        def create_button(text, y_pixel_pos, command):
            btn = tk.Label(window, text=text, font=("Trebuchet MS", 15), fg="#F6FAFC",
                    bg="#020638", bd=0)
            btn.bind("<Enter>", lambda e: btn.config(fg="#1F5186"))
            btn.bind("<Leave>", lambda e: btn.config(fg="#F6FAFC"))
            btn.bind("<Button-1>", lambda e: command())
            btn.place(x=20, y=y_pixel_pos)  

        create_button("[<-- BACK ]", 20, show_main_menu)
        create_button("[ -- Board XY Compare -- ]", 100, BrdFileCompFunc)
        create_button("[ -- Board & BOM Compare -- ]", 140, BrdAndBomCompFunc)
        
        def create_button1(text, y_pixel_pos1, command):
            btn1 = tk.Label(window, text=text, font=("Trebuchet MS", 15), fg="#F6FAFC",
                    bg="#020638", bd=0)
            btn1.bind("<Enter>", lambda e: btn1.config(fg="#1F5186"))
            btn1.bind("<Leave>", lambda e: btn1.config(fg="#F6FAFC"))
            btn1.bind("<Button-1>", lambda e: command())
            btn1.place(x=500, y=y_pixel_pos1)  

        create_button1("[|-- QC SKILL MENU --|]", 20, MenuPass)
        create_button1("[ -- EXCEL -- ]", 100, ExcelFunc)
        create_button1("[ -- QC -- ]", 140, QCSkillFunc)


#------------------------------------------------------------------------------------------------------------------------------------------------------------------


    def show_main_menu():
        global cap, after_id
        if after_id:
            window.after_cancel(after_id)
            after_id = None
        if cap:
            cap.release()
        for widget in window.winfo_children():
            widget.destroy()

        cap_path = resource_path("VidMainComp.mp4")
        cap = cv2.VideoCapture(cap_path)
        if not cap.isOpened():
            print(f"Error: Could not open video file {cap_path}")
            messagebox.showerror("Video Error", "Could not load the background video.")
            return

        bg_label = tk.Label(window, bg='black')
        bg_label.place(relwidth=1, relheight=1)
        update_video_frame(bg_label, cap_path)

        def create_button(text, y, command_func, border_style="circuit"):
            button = MainPage(window, text, y, command_func, border_style=border_style)
            return button
        
        def button_click_handler(command_func):
            if hasattr(window, 'button_references'):
                for btn in window.button_references:
                    btn.stop_animation()
            command_func()
        
        buttons = []
        
        design_btn = MainPage(window, "[ -- DESIGN -- ]", 0.2, 
                            lambda: button_click_handler(DesignFunc), 
                            border_style="circuit")
        buttons.append(design_btn)
        
        sim_btn = MainPage(window, "[ -- SIMULATION -- ]", 0.35, 
                        lambda: button_click_handler(SimFunc), 
                        border_style="corner")
        buttons.append(sim_btn)
        
        sch_btn = MainPage(window, "[ -- SCHEMATICS -- ]", 0.50, 
                        lambda: button_click_handler(SchFunc), 
                        border_style="segmented")
        buttons.append(sch_btn)
        
        doc_btn = MainPage(window, "[ -- DOCUMENTATION -- ]", 0.65, 
                        lambda: button_click_handler(IntFunc), 
                        border_style="corner")
        buttons.append(doc_btn)
        
        qc_btn = MainPage(window, "[ -- QC -- ]", 0.8, 
                        lambda: button_click_handler(QCFunc), 
                        border_style="circuit")
        buttons.append(qc_btn)
        
        window.button_references = buttons


#------------------------------------------------------------------------------------------------------------------------------------------------------------------

    show_main_menu()
    window.protocol("WM_DELETE_WINDOW", on_close)
    window.mainloop()

    if cap:
        cap.release()

else:
    messagebox.showerror("Access Denied", f"Hello {username}, access to EDA Application is denied.")